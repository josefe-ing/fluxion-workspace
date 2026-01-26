"""
Router para Pedidos Inter-CEDI (CEDI → CEDI)
Sistema de reposición de CEDI Caracas desde CEDIs Valencia (Seco, Frio, Verde)

Características:
- Demanda agregada de todas las tiendas de la región destino
- Stock de seguridad en CEDI destino
- Cobertura configurable por clase ABC (A, B, C, D)
- Múltiples CEDIs origen consolidados en un pedido
- Exportación Excel por CEDI origen
"""

from fastapi import APIRouter, HTTPException, Depends, Response
from typing import List, Optional, Any, Dict
import uuid
from datetime import datetime, date
from decimal import Decimal
import math
import logging
import io

from models.pedidos_inter_cedi import (
    # Request/Response models
    CalcularPedidoInterCediRequest,
    CalcularPedidoInterCediResponse,
    GuardarPedidoInterCediRequest,
    PedidoInterCediGuardadoResponse,
    ActualizarPedidoInterCediRequest,
    CambiarEstadoPedidoRequest,
    ExportarPedidoRequest,
    # Entity models
    ProductoInterCediCalculado,
    ProductoInterCediAjustado,
    PedidoInterCediCompleto,
    PedidoInterCediResumen,
    HistorialEstadoPedido,
    ConfiguracionRutaInterCedi,
    P75PorTienda,
    StockPorTienda,
    # Enums
    EstadoPedidoInterCedi,
    CediOrigen,
)
from db_manager import get_db_connection, get_db_connection_write

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/pedidos-inter-cedi", tags=["Pedidos Inter-CEDI"])


def get_db():
    """Get database connection (read-only)"""
    with get_db_connection() as conn:
        yield conn


def get_db_write():
    """Get database connection (read-write)"""
    with get_db_connection_write() as conn:
        yield conn


# =====================================================================================
# CONSTANTES
# =====================================================================================

# Z-scores por clase ABC para cálculo de stock de seguridad
Z_SCORES_ABC = {
    'A': Decimal('2.33'),  # 99% nivel servicio
    'B': Decimal('1.88'),  # 97% nivel servicio
    'C': Decimal('1.28'),  # 90% nivel servicio
    'D': Decimal('0.0'),   # Sin stock de seguridad estadístico
}

# Lead time default Valencia -> Caracas (días)
LEAD_TIME_DEFAULT = Decimal('2.0')

# Variabilidad estimada de demanda (30% conservador)
VARIABILIDAD_DEMANDA_DEFAULT = Decimal('0.30')


# =====================================================================================
# LISTAR PEDIDOS
# =====================================================================================

@router.get("/", response_model=List[PedidoInterCediResumen])
async def listar_pedidos_inter_cedi(
    estado: Optional[str] = None,
    cedi_destino_id: Optional[str] = None,
    limit: int = 50,
    conn: Any = Depends(get_db)
):
    """
    Lista pedidos inter-CEDI con filtros opcionales

    Filtros:
    - estado: borrador, confirmado, despachado, recibido, cancelado
    - cedi_destino_id: filtrar por CEDI destino
    """
    try:
        cursor = conn.cursor()

        query = """
            SELECT * FROM v_pedidos_inter_cedi_resumen
            WHERE 1=1
        """
        params = []

        if estado:
            query += " AND estado = %s"
            params.append(estado)

        if cedi_destino_id:
            query += " AND cedi_destino_id = %s"
            params.append(cedi_destino_id)

        query += " ORDER BY fecha_creacion DESC LIMIT %s"
        params.append(limit)

        cursor.execute(query, params)
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        cursor.close()

        pedidos = []
        for row in rows:
            row_dict = dict(zip(columns, row))
            pedidos.append(PedidoInterCediResumen(**row_dict))

        return pedidos

    except Exception as e:
        logger.error(f"Error listando pedidos inter-CEDI: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error listando pedidos: {str(e)}")


# =====================================================================================
# CALCULAR PEDIDO
# =====================================================================================

@router.post("/calcular", response_model=CalcularPedidoInterCediResponse)
async def calcular_pedido_inter_cedi(
    request: CalcularPedidoInterCediRequest,
    conn: Any = Depends(get_db)
):
    """
    Calcula pedido inter-CEDI basado en demanda regional agregada

    Proceso:
    1. Obtiene todas las tiendas de la región del CEDI destino
    2. Calcula demanda agregada (suma de P75 de cada tienda)
    3. Calcula stock de seguridad en CEDI destino
    4. Determina cantidad a pedir según clase ABC y días de cobertura
    5. Agrupa productos por CEDI origen (Seco/Frio/Verde)
    """
    try:
        logger.info(f"📦 Calculando pedido Inter-CEDI para: {request.cedi_destino_id}")
        cursor = conn.cursor()

        # 1. Obtener región del CEDI destino
        cursor.execute("""
            SELECT id, nombre, region FROM ubicaciones WHERE id = %s
        """, [request.cedi_destino_id])
        cedi_row = cursor.fetchone()

        if not cedi_row:
            raise HTTPException(status_code=404, detail="CEDI destino no encontrado")

        cedi_destino_nombre = cedi_row[1] or request.cedi_destino_id
        region = cedi_row[2] or 'CARACAS'

        # 2. Obtener todas las tiendas de la región
        cursor.execute("""
            SELECT id, nombre FROM ubicaciones
            WHERE region = %s
              AND tipo = 'tienda'
              AND activo = true
        """, [region])
        tiendas_region = cursor.fetchall()
        tiendas_ids = [t[0] for t in tiendas_region]

        if not tiendas_ids:
            raise HTTPException(
                status_code=400,
                detail=f"No hay tiendas activas en la región {region}"
            )

        logger.info(f"📍 Región: {region}, Tiendas: {len(tiendas_region)} ({', '.join([t[1] for t in tiendas_region])})")

        # 3. Calcular demanda regional agregada por producto
        # Obtener productos con ventas en la región
        query_demanda = """
            WITH ventas_30d AS (
                -- Ventas de los últimos 30 días de las tiendas de la región
                SELECT
                    producto_id,
                    ubicacion_id,
                    fecha_venta::date as fecha,
                    SUM(cantidad_vendida) as cantidad_dia
                FROM ventas
                WHERE ubicacion_id = ANY(%s)
                  AND fecha_venta >= CURRENT_DATE - INTERVAL '30 days'
                  AND fecha_venta < CURRENT_DATE
                  -- Excluir día de inauguración de Paraíso
                  AND NOT (ubicacion_id = 'tienda_18' AND fecha_venta::date = '2025-12-06')
                GROUP BY producto_id, ubicacion_id, fecha_venta::date
            ),
            p75_por_tienda AS (
                -- P75 por producto por tienda
                SELECT
                    producto_id,
                    ubicacion_id,
                    PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY cantidad_dia) as p75_tienda,
                    AVG(cantidad_dia) as promedio_tienda,
                    STDDEV(cantidad_dia) as sigma_tienda,
                    COUNT(DISTINCT fecha) as dias_con_venta
                FROM ventas_30d
                GROUP BY producto_id, ubicacion_id
            ),
            demanda_regional AS (
                -- Suma de P75 de todas las tiendas = demanda regional total
                SELECT
                    producto_id,
                    SUM(COALESCE(p75_tienda, 0)) as p75_regional,
                    AVG(COALESCE(p75_tienda, 0)) as p75_promedio_tienda,
                    -- Sigma regional: raíz de suma de varianzas
                    SQRT(SUM(COALESCE(sigma_tienda * sigma_tienda, 0))) as sigma_regional,
                    COUNT(DISTINCT ubicacion_id) as num_tiendas
                FROM p75_por_tienda
                GROUP BY producto_id
                HAVING SUM(COALESCE(p75_tienda, 0)) > 0
            ),
            abc_ranking AS (
                -- Clasificación ABC por ranking de cantidad vendida en la región
                SELECT
                    producto_id,
                    SUM(cantidad_dia) as cantidad_total,
                    ROW_NUMBER() OVER (ORDER BY SUM(cantidad_dia) DESC) as rank_cantidad
                FROM ventas_30d
                GROUP BY producto_id
            ),
            stock_cedi_destino AS (
                -- Stock actual en CEDI destino
                SELECT
                    producto_id,
                    SUM(cantidad) as stock_actual
                FROM inventario_actual
                WHERE ubicacion_id = %s
                GROUP BY producto_id
            )
            SELECT
                p.id as producto_id,
                p.codigo,
                p.codigo_barras,
                COALESCE(p.nombre, p.descripcion) as descripcion,
                p.categoria,
                p.grupo,
                p.marca,
                p.presentacion,
                p.cuadrante,
                COALESCE(p.unidades_por_bulto, 1) as unidades_por_bulto,
                COALESCE(p.unidad_pedido, 'Bulto') as unidad_pedido,
                p.peso_unitario,
                p.cedi_origen_id,
                -- Demanda regional
                COALESCE(dr.p75_regional, 0) as p75_regional,
                COALESCE(dr.p75_promedio_tienda, 0) as p75_promedio,
                COALESCE(dr.sigma_regional, dr.p75_regional * 0.3) as sigma_regional,
                COALESCE(dr.num_tiendas, 0) as num_tiendas,
                -- Stock CEDI destino
                COALESCE(scd.stock_actual, 0) as stock_cedi_destino,
                -- Clasificación ABC
                CASE
                    WHEN abc.rank_cantidad <= 50 THEN 'A'
                    WHEN abc.rank_cantidad <= 200 THEN 'B'
                    WHEN abc.rank_cantidad <= 800 THEN 'C'
                    ELSE 'D'
                END as clase_abc,
                abc.rank_cantidad
            FROM productos p
            INNER JOIN demanda_regional dr ON p.id = dr.producto_id
            LEFT JOIN stock_cedi_destino scd ON p.id = scd.producto_id
            LEFT JOIN abc_ranking abc ON p.id = abc.producto_id
            WHERE p.activo = true
              AND p.cedi_origen_id IS NOT NULL
            ORDER BY dr.p75_regional DESC
        """

        cursor.execute(query_demanda, [tiendas_ids, request.cedi_destino_id])
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()

        logger.info(f"📊 Encontrados {len(rows)} productos con demanda regional")

        # 3.0.1 Cargar productos excluidos para este CEDI destino
        try:
            cursor.execute("""
                SELECT codigo_producto
                FROM productos_excluidos_inter_cedi
                WHERE cedi_destino_id = %s AND activo = TRUE
            """, [request.cedi_destino_id])
            codigos_excluidos_intercedi = {row[0] for row in cursor.fetchall()}
            if codigos_excluidos_intercedi:
                logger.info(f"🚫 {len(codigos_excluidos_intercedi)} productos excluidos de inter-CEDI")
        except Exception as e:
            logger.warning(f"⚠️ No se pudieron cargar exclusiones inter-CEDI (tabla puede no existir): {e}")
            codigos_excluidos_intercedi = set()

        # 3.1 Obtener desglose de P75 por tienda para cada producto
        query_p75_tienda = """
            WITH ventas_30d AS (
                SELECT
                    producto_id,
                    ubicacion_id,
                    fecha_venta::date as fecha,
                    SUM(cantidad_vendida) as cantidad_dia
                FROM ventas
                WHERE ubicacion_id = ANY(%s)
                  AND fecha_venta >= CURRENT_DATE - INTERVAL '30 days'
                  AND fecha_venta < CURRENT_DATE
                  AND NOT (ubicacion_id = 'tienda_18' AND fecha_venta::date = '2025-12-06')
                GROUP BY producto_id, ubicacion_id, fecha_venta::date
            )
            SELECT
                v.producto_id,
                v.ubicacion_id as tienda_id,
                u.nombre as tienda_nombre,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY v.cantidad_dia) as p75_tienda
            FROM ventas_30d v
            LEFT JOIN ubicaciones u ON v.ubicacion_id = u.id
            GROUP BY v.producto_id, v.ubicacion_id, u.nombre
        """
        cursor.execute(query_p75_tienda, [tiendas_ids])
        p75_rows = cursor.fetchall()

        # Crear diccionario de P75 por producto y tienda
        p75_por_producto: Dict[str, List[Dict]] = {}
        for p75_row in p75_rows:
            producto_id, tienda_id, tienda_nombre, p75_valor = p75_row
            if producto_id not in p75_por_producto:
                p75_por_producto[producto_id] = []
            p75_por_producto[producto_id].append({
                'tienda_id': tienda_id,
                'tienda_nombre': tienda_nombre or tienda_id,
                'p75_unidades': Decimal(str(p75_valor)) if p75_valor else Decimal('0')
            })

        # 3.2 Obtener stock en tiendas de la región
        query_stock_tiendas = """
            SELECT
                ia.producto_id,
                ia.ubicacion_id as tienda_id,
                u.nombre as tienda_nombre,
                ia.cantidad as stock_unidades
            FROM inventario_actual ia
            LEFT JOIN ubicaciones u ON ia.ubicacion_id = u.id
            WHERE ia.ubicacion_id = ANY(%s)
        """
        cursor.execute(query_stock_tiendas, [tiendas_ids])
        stock_tiendas_rows = cursor.fetchall()

        # Crear diccionario de stock por producto y tienda
        stock_tiendas_por_producto: Dict[str, List[Dict]] = {}
        for stock_row in stock_tiendas_rows:
            producto_id, tienda_id, tienda_nombre, stock_valor = stock_row
            if producto_id not in stock_tiendas_por_producto:
                stock_tiendas_por_producto[producto_id] = []
            stock_tiendas_por_producto[producto_id].append({
                'tienda_id': tienda_id,
                'tienda_nombre': tienda_nombre or tienda_id,
                'stock_unidades': Decimal(str(stock_valor)) if stock_valor else Decimal('0')
            })

        # 4. Obtener stock disponible en CEDIs origen
        cursor.execute("""
            SELECT
                producto_id,
                ubicacion_id as cedi_origen_id,
                SUM(cantidad) as stock_disponible
            FROM inventario_actual
            WHERE ubicacion_id IN ('cedi_seco', 'cedi_frio', 'cedi_verde')
            GROUP BY producto_id, ubicacion_id
        """)
        stock_origen_rows = cursor.fetchall()
        cursor.close()

        # Crear diccionario de stock por producto y CEDI origen
        stock_por_producto_cedi: Dict[str, Dict[str, Decimal]] = {}
        for row in stock_origen_rows:
            producto_id, cedi_id, stock = row
            if producto_id not in stock_por_producto_cedi:
                stock_por_producto_cedi[producto_id] = {}
            stock_por_producto_cedi[producto_id][cedi_id] = Decimal(str(stock)) if stock else Decimal('0')

        # 5. Calcular cantidades sugeridas
        productos_calculados: List[ProductoInterCediCalculado] = []
        productos_por_cedi: Dict[str, List[ProductoInterCediCalculado]] = {
            'cedi_seco': [],
            'cedi_frio': [],
            'cedi_verde': []
        }

        lead_time = request.lead_time_dias or LEAD_TIME_DEFAULT

        for row in rows:
            row_dict = dict(zip(columns, row))

            producto_id = row_dict['producto_id']
            codigo = row_dict['codigo']

            # Saltar productos excluidos
            if codigo in codigos_excluidos_intercedi:
                continue

            cedi_origen_id = row_dict['cedi_origen_id'] or 'cedi_seco'
            unidades_por_bulto = Decimal(str(row_dict['unidades_por_bulto'] or 1))
            if unidades_por_bulto <= 0:
                unidades_por_bulto = Decimal('1')

            p75_regional = Decimal(str(row_dict['p75_regional'] or 0))
            sigma_regional = Decimal(str(row_dict['sigma_regional'] or p75_regional * VARIABILIDAD_DEMANDA_DEFAULT))
            stock_cedi_destino = Decimal(str(row_dict['stock_cedi_destino'] or 0))
            clase_abc = row_dict['clase_abc'] or 'D'

            # Obtener stock en CEDI origen
            stock_cedi_origen = Decimal('0')
            if producto_id in stock_por_producto_cedi:
                stock_cedi_origen = stock_por_producto_cedi[producto_id].get(cedi_origen_id, Decimal('0'))

            # Obtener categoría del producto
            categoria = (row_dict['categoria'] or '').upper()

            # Determinar días de cobertura según tipo de producto
            # Perecederos usan días fijos (ignoran ABC)
            es_fruver = cedi_origen_id == 'cedi_verde' or 'FRUVER' in categoria or 'FRUT' in categoria or 'VERDUR' in categoria
            es_panaderia = 'PANAD' in categoria or 'PAN ' in categoria or categoria.startswith('PAN')

            if es_fruver:
                dias_cobertura = request.dias_cobertura_fruver
            elif es_panaderia:
                dias_cobertura = request.dias_cobertura_panaderia
            else:
                # Productos normales usan días según ABC
                dias_cobertura = {
                    'A': request.dias_cobertura_a,
                    'B': request.dias_cobertura_b,
                    'C': request.dias_cobertura_c,
                    'D': request.dias_cobertura_d
                }.get(clase_abc, 18)

            # FÓRMULA INTER-CEDI:
            # SS_CEDI = Z × σ_regional × √Lead_Time
            # Stock_Min = Demanda_Regional × Lead_Time + SS_CEDI
            # Stock_Max = Stock_Min + (Demanda_Regional × Dias_Cobertura)
            # Cantidad_Sugerida = max(0, Stock_Max - Stock_Actual)

            z_score = Z_SCORES_ABC.get(clase_abc, Decimal('1.28'))

            # Stock de seguridad
            sqrt_lead_time = Decimal(str(math.sqrt(float(lead_time))))
            stock_seguridad = z_score * sigma_regional * sqrt_lead_time

            # Para clase D, usar método "Padre Prudente" (mínimo 30% de demanda durante LT)
            if clase_abc == 'D':
                demanda_ciclo = p75_regional * lead_time
                stock_seguridad = max(demanda_ciclo * Decimal('0.3'), stock_seguridad)

            # Demanda durante lead time
            demanda_ciclo = p75_regional * lead_time

            # Stock mínimo (punto de reorden)
            stock_minimo = demanda_ciclo + stock_seguridad

            # Stock máximo
            stock_maximo = stock_minimo + (p75_regional * Decimal(str(dias_cobertura)))

            # Cantidad ideal (sin limitar por stock origen)
            if stock_cedi_destino <= stock_minimo:
                cantidad_ideal_unid = max(Decimal('0'), stock_maximo - stock_cedi_destino)
            else:
                cantidad_ideal_unid = Decimal('0')

            # Cantidad sugerida (limitada por stock disponible en origen)
            cantidad_sugerida_unid = min(cantidad_ideal_unid, stock_cedi_origen)

            # Redondear a bultos completos (hacia arriba)
            cantidad_sugerida_bultos = Decimal(str(math.ceil(float(cantidad_sugerida_unid / unidades_por_bulto))))

            # Solo incluir si hay cantidad sugerida > 0
            if cantidad_sugerida_bultos > 0:
                # Obtener desglose de P75 por tienda para este producto
                p75_tiendas = []
                if producto_id in p75_por_producto:
                    for p75_data in p75_por_producto[producto_id]:
                        p75_tiendas.append(P75PorTienda(
                            tienda_id=p75_data['tienda_id'],
                            tienda_nombre=p75_data['tienda_nombre'],
                            p75_unidades=p75_data['p75_unidades']
                        ))

                # Obtener stock por tienda para este producto
                stock_tiendas = []
                stock_tiendas_total = Decimal('0')
                if producto_id in stock_tiendas_por_producto:
                    for stock_data in stock_tiendas_por_producto[producto_id]:
                        stock_tiendas.append(StockPorTienda(
                            tienda_id=stock_data['tienda_id'],
                            tienda_nombre=stock_data['tienda_nombre'],
                            stock_unidades=stock_data['stock_unidades']
                        ))
                        stock_tiendas_total += stock_data['stock_unidades']

                producto = ProductoInterCediCalculado(
                    codigo_producto=codigo,
                    codigo_barras=row_dict['codigo_barras'],
                    descripcion_producto=row_dict['descripcion'] or codigo,
                    categoria=row_dict['categoria'],
                    grupo=row_dict['grupo'],
                    marca=row_dict['marca'],
                    presentacion=row_dict['presentacion'],
                    cuadrante=row_dict.get('cuadrante'),
                    clasificacion_abc=clase_abc,

                    # CEDI Origen
                    cedi_origen_id=cedi_origen_id,
                    cedi_origen_nombre=CediOrigen.nombre(cedi_origen_id),
                    stock_cedi_origen=stock_cedi_origen,

                    # Cantidades físicas
                    unidades_por_bulto=unidades_por_bulto,
                    unidad_pedido=row_dict.get('unidad_pedido', 'Bulto'),
                    peso_unitario_kg=Decimal(str(row_dict['peso_unitario'] or 0)) if row_dict['peso_unitario'] else None,

                    # Demanda regional
                    demanda_regional_p75=p75_regional,
                    demanda_regional_promedio=Decimal(str(row_dict['p75_promedio'] or 0)),
                    num_tiendas_region=int(row_dict['num_tiendas'] or 0),
                    p75_por_tienda=p75_tiendas,

                    # Stock en tiendas de la región
                    stock_tiendas_total=stock_tiendas_total,
                    stock_por_tienda=stock_tiendas,

                    # Stock CEDI destino
                    stock_actual_cedi=stock_cedi_destino,
                    stock_en_transito=Decimal('0'),

                    # Parámetros calculados
                    stock_minimo_cedi=stock_minimo,
                    stock_seguridad_cedi=stock_seguridad,
                    stock_maximo_cedi=stock_maximo,
                    punto_reorden_cedi=stock_minimo,

                    # Cantidades sugeridas
                    cantidad_sugerida_unidades=cantidad_sugerida_unid,
                    cantidad_sugerida_bultos=cantidad_sugerida_bultos,
                    cantidad_ideal_unidades=cantidad_ideal_unid,

                    # Metadata
                    dias_cobertura_objetivo=dias_cobertura,
                    razon_pedido=f"Reposición CEDI - Clase {clase_abc} ({dias_cobertura}d cobertura)"
                )

                productos_calculados.append(producto)
                if cedi_origen_id in productos_por_cedi:
                    productos_por_cedi[cedi_origen_id].append(producto)

        # 6. Calcular totales
        total_productos = len(productos_calculados)
        total_bultos = sum(p.cantidad_sugerida_bultos for p in productos_calculados)
        total_unidades = sum(p.cantidad_sugerida_unidades for p in productos_calculados)
        total_cedis_origen = len([k for k, v in productos_por_cedi.items() if len(v) > 0])

        # Totales por CEDI origen
        totales_por_cedi = {}
        for cedi_id, productos in productos_por_cedi.items():
            if productos:
                totales_por_cedi[cedi_id] = {
                    'productos': len(productos),
                    'bultos': sum(p.cantidad_sugerida_bultos for p in productos),
                    'unidades': sum(p.cantidad_sugerida_unidades for p in productos)
                }

        logger.info(f"✅ Calculados {total_productos} productos, {total_bultos} bultos desde {total_cedis_origen} CEDIs")

        return CalcularPedidoInterCediResponse(
            productos=productos_calculados,
            productos_por_cedi_origen=productos_por_cedi,
            total_cedis_origen=total_cedis_origen,
            total_productos=total_productos,
            total_bultos=total_bultos,
            total_unidades=total_unidades,
            cedi_destino_id=request.cedi_destino_id,
            cedi_destino_nombre=cedi_destino_nombre,
            region=region,
            num_tiendas_region=len(tiendas_region),
            totales_por_cedi=totales_por_cedi,
            dias_cobertura_a=request.dias_cobertura_a,
            dias_cobertura_b=request.dias_cobertura_b,
            dias_cobertura_c=request.dias_cobertura_c,
            dias_cobertura_d=request.dias_cobertura_d,
            fecha_calculo=datetime.now(),
            mensaje=f"Pedido Inter-CEDI calculado: {total_productos} productos desde {total_cedis_origen} CEDIs para región {region}"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error calculando pedido Inter-CEDI: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error calculando pedido: {str(e)}")


# =====================================================================================
# GUARDAR PEDIDO
# =====================================================================================

@router.post("/", response_model=PedidoInterCediGuardadoResponse)
async def guardar_pedido_inter_cedi(
    request: GuardarPedidoInterCediRequest,
    conn: Any = Depends(get_db_write)
):
    """
    Guarda pedido Inter-CEDI en estado BORRADOR
    """
    try:
        logger.info(f"💾 Guardando pedido Inter-CEDI para {request.cedi_destino_id}")
        cursor = conn.cursor()

        pedido_id = str(uuid.uuid4())

        # Calcular totales
        productos_incluidos = [p for p in request.productos if p.incluido and p.cantidad_pedida_bultos > 0]
        total_productos = len(productos_incluidos)
        total_bultos = sum(p.cantidad_pedida_bultos for p in productos_incluidos)
        total_unidades = sum(p.cantidad_pedida_bultos * p.unidades_por_bulto for p in productos_incluidos)

        # Contar CEDIs origen únicos
        cedis_origen = set(p.cedi_origen_id for p in productos_incluidos)
        total_cedis_origen = len(cedis_origen)

        # Obtener nombre del CEDI destino
        cursor.execute("SELECT nombre, region FROM ubicaciones WHERE id = %s", [request.cedi_destino_id])
        cedi_row = cursor.fetchone()
        cedi_destino_nombre = request.cedi_destino_nombre or (cedi_row[0] if cedi_row else request.cedi_destino_id)
        cedi_destino_region = cedi_row[1] if cedi_row else 'CARACAS'

        # Insertar pedido (el trigger genera el número automáticamente)
        cursor.execute("""
            INSERT INTO pedidos_inter_cedi (
                id, cedi_destino_id, cedi_destino_nombre, cedi_destino_region,
                fecha_pedido, estado, prioridad,
                dias_cobertura_a, dias_cobertura_b, dias_cobertura_c, dias_cobertura_d,
                frecuencia_viajes_dias, lead_time_dias,
                total_cedis_origen, total_productos, total_lineas, total_bultos, total_unidades,
                observaciones, notas_logistica, usuario_creador
            ) VALUES (
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s
            )
            RETURNING numero_pedido, fecha_creacion
        """, [
            pedido_id, request.cedi_destino_id, cedi_destino_nombre, cedi_destino_region,
            request.fecha_pedido or date.today(), EstadoPedidoInterCedi.BORRADOR, request.prioridad,
            request.dias_cobertura_a, request.dias_cobertura_b, request.dias_cobertura_c, request.dias_cobertura_d,
            request.frecuencia_viajes_dias, float(request.lead_time_dias),
            total_cedis_origen, total_productos, total_productos, float(total_bultos), float(total_unidades),
            request.observaciones, request.notas_logistica, 'sistema'
        ])

        result = cursor.fetchone()
        numero_pedido = result[0]
        fecha_creacion = result[1]

        # Insertar detalle de productos
        for i, producto in enumerate(productos_incluidos, 1):
            cursor.execute("""
                INSERT INTO pedidos_inter_cedi_detalle (
                    pedido_id, linea_numero,
                    cedi_origen_id, cedi_origen_nombre,
                    codigo_producto, codigo_barras, descripcion_producto,
                    categoria, grupo, marca, presentacion, clasificacion_abc,
                    unidades_por_bulto, peso_unitario_kg,
                    demanda_regional_p75, demanda_regional_promedio, num_tiendas_region,
                    stock_actual_cedi, stock_en_transito,
                    stock_minimo_cedi, stock_seguridad_cedi, stock_maximo_cedi, punto_reorden_cedi,
                    stock_cedi_origen,
                    cantidad_sugerida_unidades, cantidad_sugerida_bultos,
                    cantidad_pedida_unidades, cantidad_pedida_bultos, total_unidades,
                    razon_pedido, dias_cobertura_objetivo, incluido, observaciones
                ) VALUES (
                    %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s, %s, %s,
                    %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s
                )
            """, [
                pedido_id, i,
                producto.cedi_origen_id, CediOrigen.nombre(producto.cedi_origen_id),
                producto.codigo_producto, producto.codigo_barras, producto.descripcion_producto,
                producto.categoria, producto.grupo, producto.marca, producto.presentacion, producto.clasificacion_abc,
                float(producto.unidades_por_bulto), float(producto.peso_unitario_kg) if producto.peso_unitario_kg else None,
                float(producto.demanda_regional_p75), float(producto.demanda_regional_promedio), producto.num_tiendas_region,
                float(producto.stock_actual_cedi), float(producto.stock_en_transito),
                float(producto.stock_minimo_cedi), float(producto.stock_seguridad_cedi), float(producto.stock_maximo_cedi), float(producto.punto_reorden_cedi),
                float(producto.stock_cedi_origen),
                float(producto.cantidad_sugerida_unidades), float(producto.cantidad_sugerida_bultos),
                float(producto.cantidad_pedida_unidades), float(producto.cantidad_pedida_bultos),
                float(producto.cantidad_pedida_bultos * producto.unidades_por_bulto),
                producto.razon_pedido, producto.dias_cobertura_objetivo, producto.incluido, producto.observaciones
            ])

        conn.commit()
        cursor.close()

        logger.info(f"✅ Pedido Inter-CEDI guardado: {numero_pedido}")

        return PedidoInterCediGuardadoResponse(
            id=pedido_id,
            numero_pedido=numero_pedido,
            estado=EstadoPedidoInterCedi.BORRADOR,
            total_cedis_origen=total_cedis_origen,
            total_productos=total_productos,
            total_bultos=total_bultos,
            total_unidades=total_unidades,
            fecha_creacion=fecha_creacion,
            mensaje=f"Pedido {numero_pedido} guardado exitosamente"
        )

    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error guardando pedido Inter-CEDI: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error guardando pedido: {str(e)}")


# =====================================================================================
# OBTENER PEDIDO
# =====================================================================================

@router.get("/{pedido_id}", response_model=PedidoInterCediCompleto)
async def obtener_pedido_inter_cedi(
    pedido_id: str,
    conn: Any = Depends(get_db)
):
    """
    Obtiene pedido Inter-CEDI completo con todos los productos
    """
    try:
        cursor = conn.cursor()

        # Obtener encabezado
        cursor.execute("""
            SELECT * FROM pedidos_inter_cedi WHERE id = %s
        """, [pedido_id])
        columns = [desc[0] for desc in cursor.description]
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        pedido_dict = dict(zip(columns, row))

        # Obtener detalle de productos
        cursor.execute("""
            SELECT * FROM pedidos_inter_cedi_detalle
            WHERE pedido_id = %s
            ORDER BY cedi_origen_id, linea_numero
        """, [pedido_id])
        det_columns = [desc[0] for desc in cursor.description]
        det_rows = cursor.fetchall()
        cursor.close()

        productos = []
        for det_row in det_rows:
            det_dict = dict(zip(det_columns, det_row))
            productos.append(ProductoInterCediAjustado(**det_dict))

        # Calcular días desde creación
        dias_desde_creacion = None
        if pedido_dict.get('fecha_creacion'):
            delta = datetime.now() - pedido_dict['fecha_creacion']
            dias_desde_creacion = delta.days

        # Contar productos por CEDI
        productos_cedi_seco = len([p for p in productos if p.cedi_origen_id == 'cedi_seco'])
        productos_cedi_frio = len([p for p in productos if p.cedi_origen_id == 'cedi_frio'])
        productos_cedi_verde = len([p for p in productos if p.cedi_origen_id == 'cedi_verde'])

        return PedidoInterCediCompleto(
            **pedido_dict,
            productos=productos,
            dias_desde_creacion=dias_desde_creacion,
            productos_cedi_seco=productos_cedi_seco,
            productos_cedi_frio=productos_cedi_frio,
            productos_cedi_verde=productos_cedi_verde
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo pedido Inter-CEDI: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error obteniendo pedido: {str(e)}")


# =====================================================================================
# ELIMINAR PEDIDO
# =====================================================================================

@router.delete("/{pedido_id}")
async def eliminar_pedido_inter_cedi(
    pedido_id: str,
    conn: Any = Depends(get_db_write)
):
    """
    Elimina un pedido Inter-CEDI (solo en estado BORRADOR)
    """
    try:
        cursor = conn.cursor()

        # Verificar estado
        cursor.execute("SELECT estado, numero_pedido FROM pedidos_inter_cedi WHERE id = %s", [pedido_id])
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        estado, numero_pedido = row

        if estado != EstadoPedidoInterCedi.BORRADOR:
            raise HTTPException(
                status_code=400,
                detail=f"Solo se pueden eliminar pedidos en estado BORRADOR. Estado actual: {estado}"
            )

        # Eliminar (cascade eliminará el detalle)
        cursor.execute("DELETE FROM pedidos_inter_cedi WHERE id = %s", [pedido_id])
        conn.commit()
        cursor.close()

        logger.info(f"🗑️ Pedido Inter-CEDI eliminado: {numero_pedido}")

        return {"mensaje": f"Pedido {numero_pedido} eliminado exitosamente"}

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error eliminando pedido Inter-CEDI: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error eliminando pedido: {str(e)}")


# =====================================================================================
# CAMBIOS DE ESTADO
# =====================================================================================

@router.post("/{pedido_id}/confirmar")
async def confirmar_pedido_inter_cedi(
    pedido_id: str,
    request: CambiarEstadoPedidoRequest = None,
    conn: Any = Depends(get_db_write)
):
    """Cambia estado de BORRADOR a CONFIRMADO"""
    return await _cambiar_estado(
        pedido_id, EstadoPedidoInterCedi.BORRADOR, EstadoPedidoInterCedi.CONFIRMADO,
        request, conn, campo_fecha='fecha_confirmacion', campo_usuario='usuario_confirmador'
    )


@router.post("/{pedido_id}/despachar")
async def despachar_pedido_inter_cedi(
    pedido_id: str,
    request: CambiarEstadoPedidoRequest = None,
    conn: Any = Depends(get_db_write)
):
    """Cambia estado de CONFIRMADO a DESPACHADO"""
    return await _cambiar_estado(
        pedido_id, EstadoPedidoInterCedi.CONFIRMADO, EstadoPedidoInterCedi.DESPACHADO,
        request, conn, campo_fecha='fecha_despacho', campo_usuario='usuario_despachador'
    )


@router.post("/{pedido_id}/recibir")
async def recibir_pedido_inter_cedi(
    pedido_id: str,
    request: CambiarEstadoPedidoRequest = None,
    conn: Any = Depends(get_db_write)
):
    """Cambia estado de DESPACHADO a RECIBIDO"""
    return await _cambiar_estado(
        pedido_id, EstadoPedidoInterCedi.DESPACHADO, EstadoPedidoInterCedi.RECIBIDO,
        request, conn, campo_fecha='fecha_recepcion', campo_usuario='usuario_receptor'
    )


@router.post("/{pedido_id}/cancelar")
async def cancelar_pedido_inter_cedi(
    pedido_id: str,
    request: CambiarEstadoPedidoRequest = None,
    conn: Any = Depends(get_db_write)
):
    """Cancela el pedido (solo desde BORRADOR)"""
    return await _cambiar_estado(
        pedido_id, EstadoPedidoInterCedi.BORRADOR, EstadoPedidoInterCedi.CANCELADO,
        request, conn
    )


async def _cambiar_estado(
    pedido_id: str,
    estado_esperado: str,
    nuevo_estado: str,
    request: CambiarEstadoPedidoRequest,
    conn: Any,
    campo_fecha: str = None,
    campo_usuario: str = None
):
    """Función interna para cambiar estado del pedido"""
    try:
        cursor = conn.cursor()

        # Verificar estado actual
        cursor.execute("SELECT estado, numero_pedido FROM pedidos_inter_cedi WHERE id = %s", [pedido_id])
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        estado_actual, numero_pedido = row

        if estado_actual != estado_esperado:
            raise HTTPException(
                status_code=400,
                detail=f"El pedido debe estar en estado {estado_esperado}. Estado actual: {estado_actual}"
            )

        # Construir query de actualización
        updates = ["estado = %s", "fecha_modificacion = CURRENT_TIMESTAMP"]
        params = [nuevo_estado]

        if campo_fecha:
            updates.append(f"{campo_fecha} = CURRENT_TIMESTAMP")

        if campo_usuario and request and request.usuario:
            updates.append(f"{campo_usuario} = %s")
            params.append(request.usuario)

        if request and request.notas:
            updates.append("observaciones = COALESCE(observaciones, '') || %s")
            params.append(f"\n[{nuevo_estado}] {request.notas}")

        params.append(pedido_id)

        cursor.execute(f"""
            UPDATE pedidos_inter_cedi
            SET {', '.join(updates)}
            WHERE id = %s
        """, params)

        conn.commit()
        cursor.close()

        logger.info(f"📋 Pedido {numero_pedido}: {estado_actual} → {nuevo_estado}")

        return {
            "mensaje": f"Pedido {numero_pedido} actualizado a {nuevo_estado}",
            "numero_pedido": numero_pedido,
            "estado_anterior": estado_actual,
            "estado_nuevo": nuevo_estado
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error cambiando estado: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error actualizando estado: {str(e)}")


# =====================================================================================
# HISTORIAL
# =====================================================================================

@router.get("/{pedido_id}/historial", response_model=List[HistorialEstadoPedido])
async def obtener_historial_pedido(
    pedido_id: str,
    conn: Any = Depends(get_db)
):
    """Obtiene historial de cambios de estado del pedido"""
    try:
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id, pedido_id, estado_anterior, estado_nuevo, motivo_cambio, usuario, fecha_cambio
            FROM pedidos_inter_cedi_historial
            WHERE pedido_id = %s
            ORDER BY fecha_cambio DESC
        """, [pedido_id])

        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        cursor.close()

        return [HistorialEstadoPedido(**dict(zip(columns, row))) for row in rows]

    except Exception as e:
        logger.error(f"Error obteniendo historial: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error obteniendo historial: {str(e)}")


# =====================================================================================
# CONFIGURACIÓN DE RUTAS
# =====================================================================================

@router.get("/config/rutas", response_model=List[ConfiguracionRutaInterCedi])
async def obtener_rutas_inter_cedi(
    activo: bool = True,
    conn: Any = Depends(get_db)
):
    """Obtiene configuración de rutas inter-CEDI"""
    try:
        cursor = conn.cursor()

        query = "SELECT * FROM config_rutas_inter_cedi WHERE 1=1"
        params = []

        if activo is not None:
            query += " AND activo = %s"
            params.append(activo)

        query += " ORDER BY cedi_origen_id"

        cursor.execute(query, params)
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        cursor.close()

        return [ConfiguracionRutaInterCedi(**dict(zip(columns, row))) for row in rows]

    except Exception as e:
        logger.error(f"Error obteniendo rutas: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error obteniendo rutas: {str(e)}")


# =====================================================================================
# EXPORTACIÓN EXCEL
# =====================================================================================

@router.get("/{pedido_id}/exportar")
async def exportar_pedido_excel(
    pedido_id: str,
    cedi_origen: Optional[str] = None,
    conn: Any = Depends(get_db)
):
    """
    Exporta pedido Inter-CEDI a Excel

    Query params:
    - cedi_origen: Filtrar por CEDI origen (cedi_seco, cedi_frio, cedi_verde)
                   Si no se especifica, exporta todos
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        cursor = conn.cursor()

        # Obtener pedido
        cursor.execute("""
            SELECT numero_pedido, cedi_destino_nombre, fecha_pedido, estado
            FROM pedidos_inter_cedi WHERE id = %s
        """, [pedido_id])
        pedido_row = cursor.fetchone()

        if not pedido_row:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        numero_pedido, cedi_destino, fecha_pedido, estado = pedido_row

        # Obtener productos
        query = """
            SELECT
                codigo_producto, descripcion_producto, categoria, clasificacion_abc,
                demanda_regional_p75, num_tiendas_region,
                stock_actual_cedi, stock_cedi_origen,
                cantidad_sugerida_bultos, cantidad_pedida_bultos,
                unidades_por_bulto, total_unidades,
                cedi_origen_id, cedi_origen_nombre
            FROM pedidos_inter_cedi_detalle
            WHERE pedido_id = %s
        """
        params = [pedido_id]

        if cedi_origen:
            query += " AND cedi_origen_id = %s"
            params.append(cedi_origen)

        query += " ORDER BY cedi_origen_id, clasificacion_abc, descripcion_producto"

        cursor.execute(query, params)
        rows = cursor.fetchall()
        cursor.close()

        if not rows:
            raise HTTPException(status_code=404, detail="No hay productos para exportar")

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Título
        cedi_filtro = CediOrigen.nombre(cedi_origen) if cedi_origen else "Todos"
        ws.title = f"Pedido {numero_pedido}"
        ws.merge_cells('A1:L1')
        ws['A1'] = f"Pedido Inter-CEDI: {numero_pedido} - {cedi_destino}"
        ws['A1'].font = Font(bold=True, size=14)

        ws.merge_cells('A2:L2')
        ws['A2'] = f"Fecha: {fecha_pedido} | Estado: {estado} | CEDI Origen: {cedi_filtro}"

        # Headers
        headers = [
            'Código', 'Descripción', 'Categoría', 'ABC', 'CEDI Origen',
            'Demanda P75', 'Tiendas', 'Stock CEDI Destino', 'Stock CEDI Origen',
            'Sugerido (Bultos)', 'Pedido (Bultos)', 'Unidades'
        ]

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        cedi_colors = {
            'cedi_seco': 'FFFDE7',   # Amarillo claro
            'cedi_frio': 'E3F2FD',   # Azul claro
            'cedi_verde': 'E8F5E9'   # Verde claro
        }

        for row_idx, row in enumerate(rows, 5):
            cedi_id = row[12]
            fill_color = cedi_colors.get(cedi_id, 'FFFFFF')
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            data = [
                row[0],   # Código
                row[1],   # Descripción
                row[2],   # Categoría
                row[3],   # ABC
                row[13],  # CEDI Origen nombre
                float(row[4]) if row[4] else 0,   # Demanda P75
                row[5],   # Tiendas
                float(row[6]) if row[6] else 0,   # Stock CEDI Destino
                float(row[7]) if row[7] else 0,   # Stock CEDI Origen
                float(row[8]) if row[8] else 0,   # Sugerido
                float(row[9]) if row[9] else 0,   # Pedido
                float(row[11]) if row[11] else 0  # Unidades
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill

        # Ajustar anchos de columna
        column_widths = [12, 40, 15, 6, 12, 12, 8, 15, 15, 15, 15, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Nombre del archivo
        filename = f"{numero_pedido}"
        if cedi_origen:
            filename += f"_{cedi_origen.upper()}"
        filename += ".xlsx"

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")
    except Exception as e:
        logger.error(f"Error exportando pedido: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exportando pedido: {str(e)}")


# =====================================================================================
# HISTORIAL DE VENTAS REGIONAL
# =====================================================================================

@router.get("/historial-ventas/{codigo_producto}")
async def obtener_historial_ventas_regional(
    codigo_producto: str,
    cedi_destino_id: str = "cedi_caracas",
    dias: int = 30,
    conn: Any = Depends(get_db)
):
    """
    Obtiene historial de ventas regional para un producto.
    Suma las ventas de todas las tiendas de la región del CEDI destino.

    Args:
        codigo_producto: Código del producto
        cedi_destino_id: ID del CEDI destino (para determinar la región)
        dias: Número de días de historial (default 30)

    Returns:
        ventas_diarias: Lista con fecha, día semana, y ventas por tienda + total
        totales: Resumen de totales
    """
    try:
        cursor = conn.cursor()

        # 1. Obtener región del CEDI destino
        cursor.execute("""
            SELECT region FROM ubicaciones WHERE id = %s
        """, [cedi_destino_id])
        region_row = cursor.fetchone()
        region = region_row[0] if region_row else 'CARACAS'

        # 2. Obtener tiendas de la región
        cursor.execute("""
            SELECT id, nombre FROM ubicaciones
            WHERE region = %s AND tipo = 'tienda' AND activo = true
            ORDER BY nombre
        """, [region])
        tiendas = cursor.fetchall()
        tiendas_ids = [t[0] for t in tiendas]
        tiendas_nombres = {t[0]: t[1] for t in tiendas}

        if not tiendas_ids:
            cursor.close()
            return {"ventas_diarias": [], "tiendas": [], "totales": {}}

        # 3. Obtener producto_id desde código
        cursor.execute("""
            SELECT id FROM productos WHERE codigo = %s
        """, [codigo_producto])
        prod_row = cursor.fetchone()
        if not prod_row:
            cursor.close()
            raise HTTPException(status_code=404, detail="Producto no encontrado")

        producto_id = prod_row[0]

        # 4. Obtener ventas diarias por tienda
        cursor.execute("""
            SELECT
                fecha_venta::date as fecha,
                TO_CHAR(fecha_venta::date, 'Dy') as dia_semana,
                ubicacion_id,
                SUM(cantidad_vendida) as cantidad
            FROM ventas
            WHERE producto_id = %s
              AND ubicacion_id = ANY(%s)
              AND fecha_venta::date >= CURRENT_DATE - INTERVAL '%s days'
              AND fecha_venta::date < CURRENT_DATE
              AND NOT (ubicacion_id = 'tienda_18' AND fecha_venta::date = '2025-12-06')
            GROUP BY fecha_venta::date, ubicacion_id
            ORDER BY fecha_venta::date DESC
        """, [producto_id, tiendas_ids, dias])

        ventas_rows = cursor.fetchall()
        cursor.close()

        # 5. Organizar datos por fecha
        ventas_por_fecha: Dict[str, Dict] = {}
        for fecha, dia_semana, ubicacion_id, cantidad in ventas_rows:
            fecha_str = fecha.isoformat()
            if fecha_str not in ventas_por_fecha:
                ventas_por_fecha[fecha_str] = {
                    'fecha': fecha_str,
                    'dia_semana': dia_semana,
                    'por_tienda': {},
                    'total': 0
                }
            ventas_por_fecha[fecha_str]['por_tienda'][ubicacion_id] = float(cantidad)
            ventas_por_fecha[fecha_str]['total'] += float(cantidad)

        # 6. Convertir a lista ordenada
        ventas_diarias = sorted(
            ventas_por_fecha.values(),
            key=lambda x: x['fecha'],
            reverse=True
        )

        # 7. Calcular totales y promedios
        total_vendido = sum(v['total'] for v in ventas_diarias)
        dias_con_venta = len(ventas_diarias)
        promedio_diario = total_vendido / dias_con_venta if dias_con_venta > 0 else 0

        # P75 por tienda
        p75_por_tienda = {}
        for tienda_id in tiendas_ids:
            ventas_tienda = [v['por_tienda'].get(tienda_id, 0) for v in ventas_diarias]
            if ventas_tienda:
                sorted_ventas = sorted(ventas_tienda)
                idx = int(len(sorted_ventas) * 0.75)
                p75_por_tienda[tienda_id] = sorted_ventas[min(idx, len(sorted_ventas) - 1)]
            else:
                p75_por_tienda[tienda_id] = 0

        # P75 regional
        totales_diarios = [v['total'] for v in ventas_diarias]
        if totales_diarios:
            sorted_totales = sorted(totales_diarios)
            idx = int(len(sorted_totales) * 0.75)
            p75_regional = sorted_totales[min(idx, len(sorted_totales) - 1)]
        else:
            p75_regional = 0

        return {
            "codigo_producto": codigo_producto,
            "region": region,
            "tiendas": [{"id": t[0], "nombre": t[1]} for t in tiendas],
            "ventas_diarias": ventas_diarias,
            "totales": {
                "total_vendido": total_vendido,
                "dias_con_venta": dias_con_venta,
                "promedio_diario": promedio_diario,
                "p75_regional": p75_regional,
                "p75_por_tienda": {
                    tiendas_nombres[tid]: p75
                    for tid, p75 in p75_por_tienda.items()
                }
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo historial ventas regional: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# =====================================================================================
# HISTORIAL DE INVENTARIO CEDI
# =====================================================================================

@router.get("/historial-inventario/{codigo_producto}")
async def obtener_historial_inventario_cedi(
    codigo_producto: str,
    ubicacion_id: str = "cedi_caracas",
    dias: int = 30,
    conn: Any = Depends(get_db)
):
    """
    Obtiene historial de inventario de un producto en un CEDI.
    Usa los snapshots de inventario_historico.

    Args:
        codigo_producto: Código del producto
        ubicacion_id: ID del CEDI (cedi_caracas, cedi_seco, cedi_frio, cedi_verde)
        dias: Número de días de historial (default 30)

    Returns:
        snapshots: Lista con fecha y cantidad
        estadisticas: min, max, promedio, actual
    """
    try:
        cursor = conn.cursor()

        # 1. Obtener producto_id desde código
        cursor.execute("""
            SELECT id, descripcion FROM productos WHERE codigo = %s
        """, [codigo_producto])
        prod_row = cursor.fetchone()
        if not prod_row:
            cursor.close()
            raise HTTPException(status_code=404, detail="Producto no encontrado")

        producto_id = prod_row[0]
        descripcion = prod_row[1]

        # 2. Obtener nombre de ubicación
        cursor.execute("""
            SELECT nombre FROM ubicaciones WHERE id = %s
        """, [ubicacion_id])
        ubi_row = cursor.fetchone()
        ubicacion_nombre = ubi_row[0] if ubi_row else ubicacion_id

        # 3. Obtener snapshots históricos
        cursor.execute("""
            SELECT
                fecha_snapshot::date as fecha,
                MAX(cantidad) as cantidad
            FROM inventario_historico
            WHERE producto_id = %s
              AND ubicacion_id = %s
              AND fecha_snapshot >= CURRENT_DATE - INTERVAL '%s days'
            GROUP BY fecha_snapshot::date
            ORDER BY fecha_snapshot::date DESC
        """, [producto_id, ubicacion_id, dias])

        snapshots_rows = cursor.fetchall()

        # 4. Obtener stock actual
        cursor.execute("""
            SELECT COALESCE(SUM(cantidad), 0) as stock_actual
            FROM inventario_actual
            WHERE producto_id = %s AND ubicacion_id = %s
        """, [producto_id, ubicacion_id])
        stock_actual_row = cursor.fetchone()
        stock_actual = float(stock_actual_row[0]) if stock_actual_row else 0

        cursor.close()

        # 5. Formatear snapshots
        snapshots = []
        for fecha, cantidad in snapshots_rows:
            snapshots.append({
                'fecha': fecha.isoformat(),
                'cantidad': float(cantidad)
            })

        # 5b. Añadir stock actual como punto de hoy si no existe
        hoy = date.today().isoformat()
        if snapshots:
            # Si el último snapshot no es de hoy, añadir el stock actual
            if snapshots[0]['fecha'] != hoy:
                snapshots.insert(0, {
                    'fecha': hoy,
                    'cantidad': stock_actual
                })
        else:
            # Si no hay snapshots, crear uno con el stock actual
            snapshots.append({
                'fecha': hoy,
                'cantidad': stock_actual
            })

        # 6. Calcular estadísticas
        cantidades = [s['cantidad'] for s in snapshots]
        if cantidades:
            estadisticas = {
                'max': max(cantidades),
                'min': min(cantidades),
                'promedio': sum(cantidades) / len(cantidades),
                'actual': stock_actual,
                'dias_con_datos': len(cantidades)
            }
        else:
            estadisticas = {
                'max': stock_actual,
                'min': stock_actual,
                'promedio': stock_actual,
                'actual': stock_actual,
                'dias_con_datos': 0
            }

        return {
            "codigo_producto": codigo_producto,
            "descripcion_producto": descripcion,
            "ubicacion_id": ubicacion_id,
            "ubicacion_nombre": ubicacion_nombre,
            "snapshots": snapshots,
            "estadisticas": estadisticas
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo historial inventario CEDI: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
