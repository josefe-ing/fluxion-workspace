"""
Router para Pedidos Multi-Tienda con algoritmo DPD+U

Este módulo permite:
- Calcular pedidos para múltiples tiendas desde un CEDI
- Detectar conflictos de stock (escasez)
- Aplicar distribución DPD+U (Demanda + Urgencia)
- Guardar múltiples pedidos en una sola transacción
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import JSONResponse
from typing import List, Optional, Dict, Any, Tuple
import uuid
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from decimal import Decimal
import logging
import math
import asyncio
from concurrent.futures import ThreadPoolExecutor
import time as _time

from models.pedidos_multitienda import (
    CalcularMultiTiendaRequest,
    CalcularMultiTiendaResponse,
    ConflictoProductoResponse,
    AsignacionTiendaResponse,
    ConfigDPDUResponse,
    PedidoTiendaResponse,
    ProductoPedidoSimplificado,
    GuardarMultiTiendaRequest,
    GuardarMultiTiendaResponse,
    PedidoGuardadoInfo,
)
from services.algoritmo_dpdu import (
    ConfigDPDU,
    DatosTiendaProducto,
    calcular_distribucion_dpdu,
    detectar_conflicto,
    to_dict as asignacion_to_dict,
)
from services.calculo_inventario_abc import (
    calcular_inventario_simple,
    ConfigTiendaABC,
    set_config_tienda,
    LEAD_TIME_DEFAULT
)
from db_manager import get_db_connection, get_db_connection_write, get_db_connection_resilient

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/pedidos-multitienda", tags=["Pedidos Multi-Tienda"])

# Timezone de Venezuela (UTC-4)
VENEZUELA_TZ = ZoneInfo("America/Caracas")


def get_db():
    """Get database connection (read-only, resilient to replica conflicts)"""
    with get_db_connection_resilient() as conn:
        yield conn


def get_db_write():
    """Get database connection (read-write)"""
    with get_db_connection_write() as conn:
        yield conn


# =====================================================================================
# FUNCIONES AUXILIARES
# =====================================================================================

async def obtener_config_dpdu(conn) -> ConfigDPDU:
    """Obtiene la configuración DPD+U desde la base de datos."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            COALESCE(dpdu_peso_demanda, 0.60) as peso_demanda,
            COALESCE(dpdu_peso_urgencia, 0.40) as peso_urgencia,
            COALESCE(dpdu_dias_minimo_urgencia, 0.5) as dias_minimo
        FROM config_inventario_global
        LIMIT 1
    """)
    row = cursor.fetchone()
    cursor.close()

    if row:
        return ConfigDPDU(
            peso_demanda=float(row[0]),
            peso_urgencia=float(row[1]),
            dias_minimo_urgencia=float(row[2])
        )

    # Valores por defecto si no existe configuración
    return ConfigDPDU()


async def cargar_config_tienda(conn, tienda_id: str, set_global: bool = True) -> ConfigTiendaABC:
    """
    Carga la configuración ABC específica de una tienda desde config_parametros_abc_tienda.
    Si no existe configuración, retorna valores por defecto.

    Args:
        set_global: Si True, aplica la configuración globalmente via set_config_tienda().
                    Pasar False para ejecución paralela (thread-safe).
    """
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT lead_time_override, dias_cobertura_a, dias_cobertura_b,
                   dias_cobertura_c, clase_d_dias_cobertura
            FROM config_parametros_abc_tienda
            WHERE tienda_id = %s AND activo = true
        """, [tienda_id])
        config_row = cursor.fetchone()

        if config_row:
            config = ConfigTiendaABC(
                lead_time=float(config_row[0]) if config_row[0] else LEAD_TIME_DEFAULT,
                dias_cobertura_a=int(config_row[1]) if config_row[1] else 7,
                dias_cobertura_b=int(config_row[2]) if config_row[2] else 14,
                dias_cobertura_c=int(config_row[3]) if config_row[3] else 21,
                dias_cobertura_d=int(config_row[4]) if config_row[4] else 30
            )
            if set_global:
                set_config_tienda(config)
            logger.info(f"📋 Config tienda {tienda_id} aplicada: LT={config.lead_time}, A={config.dias_cobertura_a}d, B={config.dias_cobertura_b}d, C={config.dias_cobertura_c}d, D={config.dias_cobertura_d}d")
            return config
        else:
            # Usar defaults
            config = ConfigTiendaABC()
            if set_global:
                set_config_tienda(config)
            logger.info(f"📋 Usando configuración ABC por defecto para {tienda_id}")
            return config
    except Exception as e:
        logger.warning(f"No se pudo cargar config tienda {tienda_id}: {e}. Usando defaults.")
        config = ConfigTiendaABC()
        if set_global:
            set_config_tienda(config)
        return config
    finally:
        cursor.close()


async def calcular_transito_tienda(conn, cedi_origen: str, tienda_destino: str) -> Dict[str, Dict]:
    """
    Calcula productos en tránsito hacia una tienda.

    Detecta el tránsito real analizando incrementos en inventario_historico
    (misma lógica que verificar-llegada). NO se basa en el estado del pedido.

    Algoritmo:
    1. Buscar pedidos recientes (10 días) del CEDI hacia la tienda
    2. Para cada producto, detectar incrementos en inventario_historico desde fecha_pedido
    3. transito = cantidad_pedida - llegadas_detectadas (si > 0)

    Returns: {codigo_producto: {transito_bultos: float, desglose: [...]}}
    """
    cursor = conn.cursor()

    # 1. Obtener productos de pedidos recientes (no cancelados/rechazados)
    cursor.execute("""
        SELECT
            ps.id, ps.numero_pedido, ps.fecha_pedido, ps.estado,
            psd.codigo_producto, psd.descripcion_producto,
            psd.cantidad_pedida_bultos,
            COALESCE(p.unidades_por_bulto, 1) as unidades_por_bulto
        FROM pedidos_sugeridos ps
        JOIN pedidos_sugeridos_detalle psd ON psd.pedido_id = ps.id
        JOIN productos p ON psd.codigo_producto = p.codigo
        WHERE ps.tienda_destino_id = %s
          AND ps.cedi_origen_id = %s
          AND ps.fecha_pedido >= CURRENT_DATE - INTERVAL '10 days'
          AND ps.estado NOT IN ('cancelado', 'rechazado')
          AND psd.cantidad_pedida_bultos > 0
          AND psd.incluido = true
    """, [tienda_destino, cedi_origen])

    pedidos_productos = cursor.fetchall()
    transito_por_producto: Dict[str, Dict] = {}

    # 2. Para cada producto/pedido, detectar llegadas via inventario_historico
    for row in pedidos_productos:
        (pedido_id, numero_pedido, fecha_pedido, estado,
         codigo, desc, cantidad_pedida, unidades_por_bulto) = row

        # Query de detección de llegada (igual que verificar-llegada línea 2347)
        # Suma todos los incrementos positivos entre snapshots consecutivos
        cursor.execute("""
            WITH snapshots AS (
                SELECT
                    fecha_snapshot,
                    SUM(cantidad) as cantidad,
                    LAG(SUM(cantidad)) OVER (ORDER BY fecha_snapshot) as cantidad_anterior
                FROM inventario_historico
                WHERE producto_id = %s
                  AND ubicacion_id = %s
                  AND fecha_snapshot >= %s
                GROUP BY fecha_snapshot
                ORDER BY fecha_snapshot
            )
            SELECT COALESCE(SUM(
                CASE
                    WHEN cantidad_anterior IS NOT NULL AND cantidad > cantidad_anterior
                    THEN cantidad - cantidad_anterior
                    ELSE 0
                END
            ), 0) as total_llegadas
            FROM snapshots
        """, [codigo, tienda_destino, fecha_pedido])

        llegadas_unidades = float(cursor.fetchone()[0] or 0)
        llegadas_bultos = llegadas_unidades / (unidades_por_bulto or 1)
        pendiente = max(0, float(cantidad_pedida) - llegadas_bultos)

        if pendiente > 0:
            if codigo not in transito_por_producto:
                transito_por_producto[codigo] = {'transito_bultos': 0, 'desglose': []}

            transito_por_producto[codigo]['transito_bultos'] += pendiente
            transito_por_producto[codigo]['desglose'].append({
                'numero_pedido': numero_pedido,
                'fecha_pedido': str(fecha_pedido),
                'estado': estado,
                'pedido_bultos': float(cantidad_pedida),
                'llegadas_bultos': round(llegadas_bultos, 2),
                'pendiente_bultos': round(pendiente, 2)
            })

    cursor.close()

    if transito_por_producto:
        logger.info(f"🚚 Tránsito calculado para {tienda_destino} desde {cedi_origen}: {len(transito_por_producto)} productos en tránsito")

    return transito_por_producto


async def obtener_productos_tienda(
    conn,
    cedi_origen: str,
    tienda_destino: str,
    dias_cobertura: int = 3,
    filtros: Optional[Dict[str, Any]] = None,
    config_tienda_abc: Optional[ConfigTiendaABC] = None
) -> List[Dict]:
    """
    Obtiene TODOS los productos para una tienda, marcando cuáles necesitan reposición.

    LÓGICA:
    - Retorna todos los productos que tienen ventas, stock en tienda o stock en CEDI
    - Marca como 'es_sugerido=True' los que están por debajo del ROP (necesitan reposición)
    - Marca como 'es_sugerido=False' los que están bien abastecidos (para revisión manual)

    Esto permite al usuario:
    - Ver productos sugeridos automáticamente
    - Buscar productos no sugeridos y agregarlos manualmente si lo considera necesario

    Niveles de stock por ABC:
    - SS (Stock Seguridad): A=1.5d, B=2d, C=3d, D=4d
    - ROP (Punto Reorden): A=3d, B=4d, C=6d, D=8d
    - MAX (Stock Máximo): A=5d, B=8d, C=12d, D=15d
    """
    cursor = conn.cursor()

    # Query principal: obtiene productos con ventas, stock, clasificación ABC y estadísticas
    # IMPORTANTE: Aplica los mismos filtros que el wizard de una sola tienda:
    # 1. Solo productos activos (p.activo = true)
    # 2. Excluye productos en productos_excluidos_tienda
    # 3. Solo productos con clasificación ABC válida (A, B, C, D)
    # 4. Calcula sigma_demanda y demanda_maxima para cálculo estadístico de inventario
    query = """
        WITH ventas_diarias AS (
            -- Ventas diarias por producto (base para todos los cálculos)
            SELECT
                producto_id,
                DATE(fecha_venta) as fecha,
                SUM(cantidad_vendida) as cantidad_vendida
            FROM ventas
            WHERE ubicacion_id = %(tienda)s
              AND fecha_venta >= CURRENT_DATE - INTERVAL '30 days'
              AND fecha_venta < CURRENT_DATE
            GROUP BY producto_id, DATE(fecha_venta)
        ),
        ventas_20dias AS (
            SELECT
                producto_id,
                AVG(cantidad_vendida) as prom_20d,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY cantidad_vendida) as p75
            FROM ventas_diarias
            WHERE fecha >= CURRENT_DATE - INTERVAL '20 days'
            GROUP BY producto_id
        ),
        estadisticas_30d AS (
            -- Sigma y demanda máxima para cálculo estadístico (igual que single-store)
            SELECT
                producto_id,
                CASE
                    WHEN COUNT(*) <= 1 THEN 0
                    ELSE COALESCE(STDDEV(cantidad_vendida), 0)
                END as sigma_demanda,
                COALESCE(MAX(cantidad_vendida), 0) as demanda_maxima
            FROM ventas_diarias
            GROUP BY producto_id
        ),
        stock_tienda AS (
            SELECT producto_id, SUM(cantidad) as stock
            FROM inventario_actual
            WHERE ubicacion_id = %(tienda)s
            GROUP BY producto_id
        ),
        stock_cedi AS (
            SELECT producto_id, SUM(cantidad) as stock
            FROM inventario_actual
            WHERE ubicacion_id = %(cedi)s
            GROUP BY producto_id
        ),
        abc_tienda AS (
            SELECT producto_id, clase_abc, venta_30d
            FROM productos_abc_tienda
            WHERE ubicacion_id = %(tienda)s
        ),
        productos_excluidos AS (
            SELECT producto_id
            FROM productos_excluidos_tienda
            WHERE tienda_id = %(tienda)s AND activo = TRUE
        )
        SELECT
            p.id as producto_id,
            p.codigo,
            p.descripcion,
            p.categoria,
            p.cuadrante,
            COALESCE(p.unidades_por_bulto, 1) as unidades_por_bulto,
            COALESCE(v.prom_20d, 0) as prom_20d,
            COALESCE(v.p75, 0) as p75,
            COALESCE(st.stock, 0) as stock_tienda,
            COALESCE(sc.stock, 0) as stock_cedi,
            COALESCE(abc.clase_abc, 'D') as clase_abc,
            COALESCE(abc.venta_30d, 0) as venta_30d,
            COALESCE(est.sigma_demanda, 0) as sigma_demanda,
            COALESCE(est.demanda_maxima, v.p75 * 2, 0) as demanda_maxima,
            COALESCE(p.es_generador_trafico, false) as es_generador_trafico,
            p.peso_unitario
        FROM productos p
        LEFT JOIN ventas_20dias v ON p.id = v.producto_id
        LEFT JOIN estadisticas_30d est ON p.id = est.producto_id
        LEFT JOIN stock_tienda st ON p.id = st.producto_id
        LEFT JOIN stock_cedi sc ON p.id = sc.producto_id
        LEFT JOIN abc_tienda abc ON p.id = abc.producto_id
        LEFT JOIN productos_excluidos pe ON p.id = pe.producto_id
        WHERE p.activo = true
          AND pe.producto_id IS NULL
          AND (abc.clase_abc IN ('A', 'B', 'C', 'D') OR abc.clase_abc IS NULL)
          AND (COALESCE(v.p75, 0) > 0 OR COALESCE(st.stock, 0) > 0 OR COALESCE(sc.stock, 0) > 0)
    """

    # Construir parámetros
    params = {'tienda': tienda_destino, 'cedi': cedi_origen}

    # Agregar filtro de cuadrantes si se proporciona
    if filtros and filtros.get('cuadrantes'):
        query += " AND p.cuadrante = ANY(%(cuadrantes)s)"
        params['cuadrantes'] = filtros['cuadrantes']

    # Agregar ORDER BY al final
    query += " ORDER BY COALESCE(abc.venta_30d, 0) DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()

    # Cargar límites de capacidad para esta tienda (igual que single-store)
    limites_capacidad = {}  # {codigo_producto: {capacidad_maxima, minimo_exhibicion}}
    try:
        cursor.execute("""
            SELECT producto_codigo, capacidad_maxima_unidades, minimo_exhibicion_unidades,
                   tipo_restriccion
            FROM capacidad_almacenamiento_producto
            WHERE tienda_id = %s AND activo = true
        """, [tienda_destino])
        for lim_row in cursor.fetchall():
            limites_capacidad[lim_row[0]] = {
                'capacidad_maxima': float(lim_row[1]) if lim_row[1] else None,
                'minimo_exhibicion': float(lim_row[2]) if lim_row[2] else None,
                'tipo_restriccion': lim_row[3] or 'espacio_fisico'
            }
        if limites_capacidad:
            logger.info(f"📦 Límites de capacidad cargados para {tienda_destino}: {len(limites_capacidad)} productos")
    except Exception as e:
        logger.warning(f"No se pudo cargar límites de capacidad para {tienda_destino}: {e}")

    # Cargar configuración de cobertura por categoría (perecederos, etc.)
    config_cobertura_categoria = {}
    try:
        cursor.execute("""
            SELECT categoria_normalizada, dias_cobertura_a, dias_cobertura_b,
                   dias_cobertura_c, dias_cobertura_d
            FROM config_cobertura_categoria
            WHERE activo = true
        """)
        for cat_row in cursor.fetchall():
            cat_norm = (cat_row[0] or '').strip().upper()
            config_cobertura_categoria[cat_norm] = {
                'A': cat_row[1] if cat_row[1] is not None else 7,
                'B': cat_row[2] if cat_row[2] is not None else 14,
                'C': cat_row[3] if cat_row[3] is not None else 21,
                'D': cat_row[4] if cat_row[4] is not None else 30
            }
        if config_cobertura_categoria:
            logger.info(f"🥬 Coberturas por categoría cargadas: {list(config_cobertura_categoria.keys())}")
    except Exception as e:
        logger.warning(f"No se pudo cargar config cobertura por categoría: {e}")

    cursor.close()

    # Calcular tránsito para esta tienda ANTES del loop de productos
    # para que podamos usar stock efectivo (stock + tránsito) al calcular SUG
    transito_data = await calcular_transito_tienda(conn, cedi_origen, tienda_destino)

    # Pre-calcular P75 de referencia para TODOS los productos candidatos a envío prueba
    # (p75=0 y stock_cedi>0) en UNA SOLA query batch, en vez de N queries individuales
    candidatos_envio_prueba = [
        row[0]  # producto_id
        for row in rows
        if float(row[7]) == 0 and float(row[9]) > 0  # p75==0 and stock_cedi>0
    ]
    p75_referencia_batch: Dict[str, float] = {}  # {producto_id: p75_promedio}
    if candidatos_envio_prueba:
        cursor_ep = conn.cursor()
        cursor_ep.execute("""
            WITH tiendas_mismo_cedi AS (
                SELECT DISTINCT ps.tienda_destino_id
                FROM pedidos_sugeridos ps
                WHERE ps.cedi_origen_id = %s
                  AND ps.tienda_destino_id != %s
                LIMIT 10
            ),
            ventas_diarias_tiendas AS (
                SELECT
                    v.producto_id,
                    v.ubicacion_id,
                    DATE(v.fecha_venta) as fecha,
                    SUM(v.cantidad_vendida) as cantidad_vendida
                FROM ventas v
                WHERE v.ubicacion_id IN (SELECT tienda_destino_id FROM tiendas_mismo_cedi)
                  AND v.producto_id = ANY(%s)
                  AND v.fecha_venta >= CURRENT_DATE - INTERVAL '20 days'
                  AND v.fecha_venta < CURRENT_DATE
                GROUP BY v.producto_id, v.ubicacion_id, DATE(v.fecha_venta)
            ),
            p75_por_tienda AS (
                SELECT
                    producto_id,
                    ubicacion_id,
                    PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY cantidad_vendida) as p75
                FROM ventas_diarias_tiendas
                GROUP BY producto_id, ubicacion_id
            )
            SELECT producto_id, AVG(p75) as p75_promedio
            FROM p75_por_tienda
            WHERE p75 > 0
            GROUP BY producto_id
        """, [cedi_origen, tienda_destino, candidatos_envio_prueba])
        for ep_row in cursor_ep.fetchall():
            if ep_row[1] and float(ep_row[1]) > 0:
                p75_referencia_batch[ep_row[0]] = float(ep_row[1])
        cursor_ep.close()
        logger.info(f"📦 Envío prueba batch: {len(candidatos_envio_prueba)} candidatos, {len(p75_referencia_batch)} con P75 referencia ({tienda_destino})")

    # Pre-calcular mediana de P75 por categoría (para envío prueba sin referencia)
    from collections import defaultdict
    import statistics as stats_module
    p75_por_cat: dict = defaultdict(list)
    for row in rows:
        cat = (row[3] or '').strip().upper()
        p75_val = float(row[7])
        if p75_val > 0 and cat:
            p75_por_cat[cat].append(p75_val)
    p75_mediana_cat = {
        cat: stats_module.median(vals) for cat, vals in p75_por_cat.items() if vals
    }
    all_p75_vals = [v for vals in p75_por_cat.values() for v in vals]
    p75_fallback = stats_module.median(all_p75_vals) if all_p75_vals else 1.0

    productos = []
    for row in rows:
        producto_id = row[0]
        codigo = row[1]
        descripcion = row[2]
        categoria = row[3]
        cuadrante = row[4]
        unidades_por_bulto = int(row[5]) or 1
        prom_20d = float(row[6])
        p75 = float(row[7])
        stock_tienda = float(row[8])
        stock_cedi = float(row[9])
        clase_abc = row[10]
        venta_30d = float(row[11])
        sigma_demanda = float(row[12])
        demanda_maxima = float(row[13])
        es_generador_trafico = bool(row[14]) if row[14] else False
        peso_unitario = float(row[15]) if row[15] else None

        # Obtener tránsito para este producto
        transito_info = transito_data.get(codigo, {'transito_bultos': 0, 'desglose': []})
        transito_bultos = transito_info['transito_bultos']
        transito_unidades = transito_bultos * unidades_por_bulto

        # Calcular stock efectivo (stock actual + tránsito)
        stock_efectivo = stock_tienda + transito_unidades

        # Lógica de envío prueba: Si no ha vendido localmente pero hay stock en CEDI
        # usar P75 de referencia pre-calculado en batch
        p75_usado = p75
        clase_abc_usada = clase_abc if clase_abc else 'D'
        es_envio_prueba = False

        if p75 == 0 and stock_cedi > 0:
            p75_ref = p75_referencia_batch.get(producto_id)
            if p75_ref and p75_ref > 0:
                p75_usado = p75_ref
                clase_abc_usada = 'D'  # Conservador para productos nuevos
                es_envio_prueba = True
            else:
                # Sin ventas en ninguna tienda pero con stock en CEDI -> envío prueba
                # Usar mediana de P75 de la categoría como proxy de demanda
                cat_norm = (categoria or '').strip().upper()
                p75_usado = p75_mediana_cat.get(cat_norm, p75_fallback)
                clase_abc_usada = 'D'
                es_envio_prueba = True

        # Determinar override de días de cobertura por categoría (perecederos)
        dias_cobertura_override = None
        if categoria:
            cat_norm = categoria.strip().upper()
            if cat_norm in config_cobertura_categoria:
                # Usar días de cobertura específicos para esta categoría según clase ABC
                dias_cobertura_override = config_cobertura_categoria[cat_norm].get(clase_abc_usada)

        # Usar el mismo cálculo estadístico que single-store (calcular_inventario_simple)
        # Esto garantiza que ambos wizards usen la misma lógica de ROP, SS y MAX
        # IMPORTANTE: Usar stock_efectivo (stock + tránsito) para calcular SUG correctamente
        # Si es envío prueba, usar p75_usado (de referencia) y clase D
        sigma_usada = sigma_demanda
        if es_envio_prueba and sigma_demanda == 0:
            sigma_usada = p75_usado * 0.3  # Estimar 30% de variabilidad para productos nuevos

        resultado = calcular_inventario_simple(
            demanda_p75=p75_usado,  # ← Usar P75 de referencia si es envío prueba
            sigma_demanda=sigma_usada,
            demanda_maxima=demanda_maxima if demanda_maxima > 0 else p75_usado * 2,
            unidades_por_bulto=unidades_por_bulto,
            stock_actual=stock_efectivo,  # ← CAMBIO: usar stock efectivo (stock + tránsito)
            stock_cedi=stock_cedi,
            clase_abc=clase_abc_usada,  # ← Usar clase D si es envío prueba
            es_generador_trafico=es_generador_trafico,
            dias_cobertura_override=dias_cobertura_override,
            config_tienda=config_tienda_abc  # Thread-safe: pasar config explícitamente
        )

        # Extraer valores del resultado estadístico (en unidades)
        punto_reorden_unid = resultado.punto_reorden_unid
        stock_maximo_unid = resultado.stock_maximo_unid
        stock_seguridad_unid = resultado.stock_seguridad_unid
        cantidad_sugerida_unid = resultado.cantidad_sugerida_unid
        cantidad_sugerida_bultos = resultado.cantidad_sugerida_bultos
        tiene_sobrestock = resultado.tiene_sobrestock

        # Calcular días de stock usando stock efectivo (incluye tránsito)
        # Si es envío prueba, usar p75_usado para el cálculo de días
        if p75_usado > 0:
            dias_stock = stock_efectivo / p75_usado  # ← CAMBIO: usar stock efectivo y p75_usado
            dias_ss = stock_seguridad_unid / p75_usado
            dias_rop = punto_reorden_unid / p75_usado
            dias_max = stock_maximo_unid / p75_usado
        else:
            dias_stock = 999 if stock_efectivo > 0 else 0
            dias_ss = 0
            dias_rop = 0
            dias_max = 0

        # Determinar si necesita reposición usando el cálculo estadístico
        # Similar a single-store: sugiere si hay cantidad > 0 y hay stock en CEDI
        necesita_reposicion = (
            cantidad_sugerida_bultos > 0 and
            stock_cedi > 0 and
            not tiene_sobrestock
        )

        # Si no necesita reposición, forzar cantidades a 0
        # Excepción: envío prueba garantiza mínimo 1 bulto
        if not necesita_reposicion:
            if es_envio_prueba and not tiene_sobrestock and stock_cedi > 0:
                cantidad_sugerida_bultos = 1
                cantidad_sugerida_unid = float(unidades_por_bulto)
            else:
                cantidad_sugerida_unid = 0
                cantidad_sugerida_bultos = 0

        # Aplicar límites de inventario (igual que single-store)
        # Orden: 1) Mínimo exhibición, 2) Capacidad máxima
        limite_info = limites_capacidad.get(codigo)
        if limite_info and not tiene_sobrestock:
            # 1. MÍNIMO DE EXHIBICIÓN: elevar cantidad si es necesario para exhibición
            minimo_exhibicion = limite_info.get('minimo_exhibicion')
            if minimo_exhibicion and minimo_exhibicion > 0:
                # Calcular cuántas unidades necesitamos para alcanzar el mínimo de exhibición
                unidades_necesarias_exhibicion = max(0, minimo_exhibicion - stock_tienda)

                if unidades_necesarias_exhibicion > cantidad_sugerida_unid:
                    cantidad_antes_minimo = cantidad_sugerida_unid
                    cantidad_sugerida_unid = unidades_necesarias_exhibicion
                    cantidad_sugerida_bultos = math.ceil(unidades_necesarias_exhibicion / unidades_por_bulto) if unidades_por_bulto > 0 else 0

                    logger.info(
                        f"📊 {codigo} ({tienda_destino}): Elevado por mínimo exhibición. "
                        f"Original: {cantidad_antes_minimo:.0f} → Elevado: {cantidad_sugerida_unid:.0f} unid "
                        f"(Mín exhibición: {minimo_exhibicion:.0f}, Stock: {stock_tienda:.0f})"
                    )

            # 2. CAPACIDAD MÁXIMA: no exceder el espacio disponible
            capacidad_maxima = limite_info.get('capacidad_maxima')
            if capacidad_maxima and capacidad_maxima > 0:
                # Espacio disponible = capacidad máxima - stock actual
                espacio_disponible = max(0, capacidad_maxima - stock_tienda)

                # Si la cantidad sugerida excede el espacio disponible, ajustar
                if cantidad_sugerida_unid > espacio_disponible:
                    cantidad_antes_cap = cantidad_sugerida_unid
                    cantidad_sugerida_unid = espacio_disponible
                    cantidad_sugerida_bultos = math.ceil(espacio_disponible / unidades_por_bulto) if unidades_por_bulto > 0 else 0

                    tipo_restriccion = limite_info.get('tipo_restriccion', 'espacio_fisico')
                    logger.info(
                        f"⚠️ {codigo} ({tienda_destino}): Ajustado por capacidad de {tipo_restriccion}. "
                        f"Original: {cantidad_antes_cap:.0f} → Ajustado: {cantidad_sugerida_unid:.0f} unid "
                        f"(Cap. máx: {capacidad_maxima:.0f}, Stock: {stock_tienda:.0f}, Disponible: {espacio_disponible:.0f})"
                    )

        # Incluir TODOS los productos (sugeridos y no sugeridos)
        productos.append({
            'producto_id': producto_id,
            'codigo_producto': codigo,
            'descripcion_producto': descripcion,
            'categoria': categoria,
            'cuadrante': cuadrante,
            'clasificacion_abc': clase_abc,
            'unidades_por_bulto': unidades_por_bulto,
            'prom_p75_unid': p75,
            'prom_20dias_unid': prom_20d,
            'stock_tienda': stock_tienda,
            'stock_cedi_origen': stock_cedi,
            'dias_stock': round(dias_stock, 2),
            # Niveles en DÍAS (como single-store)
            'dias_ss': round(dias_ss, 1),
            'dias_rop': round(dias_rop, 1),
            'dias_max': round(dias_max, 1),
            # También incluir en unidades por si se necesita
            'stock_seguridad_unid': stock_seguridad_unid,
            'punto_reorden_unid': punto_reorden_unid,
            'stock_maximo_unid': stock_maximo_unid,
            'cantidad_necesaria_unid': cantidad_sugerida_unid,
            'cantidad_sugerida_bultos': int(cantidad_sugerida_bultos),
            'cantidad_sugerida_unid': cantidad_sugerida_unid,
            'es_sugerido': necesita_reposicion,
            # Tránsito (ya calculado antes del loop)
            'transito_bultos': transito_bultos,
            'transito_desglose': transito_info['desglose'],
            # Peso
            'peso_kg': peso_unitario,
        })

    return productos


# =====================================================================================
# CEDI CARACAS: Cálculo de productos usando demanda regional agregada
# =====================================================================================

async def obtener_productos_cedi_caracas(
    conn,
    cedi_origen: str,
    dias_cobertura: int = 3,
    filtros: Optional[Dict[str, Any]] = None,
    config_tienda_abc: Optional[ConfigTiendaABC] = None
) -> List[Dict]:
    """
    Calcula productos para CEDI Caracas como destino en multi-tienda.

    La "demanda" de CEDI Caracas = SUM(P75 de todas las tiendas de la región Caracas).
    El stock es el inventario actual en cedi_caracas.
    Usa calcular_inventario_simple() con config ABC específica de cedi_caracas.

    Output: Mismo formato dict que obtener_productos_tienda() para integración
    transparente con el algoritmo DPD+U.
    """
    cursor = conn.cursor()

    # 1. Obtener tiendas de la región Caracas
    cursor.execute("""
        SELECT id, nombre FROM ubicaciones
        WHERE region = 'CARACAS'
          AND tipo = 'tienda'
          AND activo = true
    """)
    tiendas_region = cursor.fetchall()
    tiendas_ids = [t[0] for t in tiendas_region]

    if not tiendas_ids:
        logger.warning("⚠️ No hay tiendas activas en región CARACAS para cálculo multi-tienda")
        cursor.close()
        return []

    logger.info(f"📍 CEDI Caracas multi-tienda: {len(tiendas_region)} tiendas en región CARACAS")

    # 2. Query principal: demanda regional + stock CEDI Caracas + stock CEDI origen
    query = """
        WITH ventas_30d AS (
            SELECT
                producto_id,
                ubicacion_id,
                fecha_venta::date as fecha,
                SUM(cantidad_vendida) as cantidad_dia
            FROM ventas
            WHERE ubicacion_id = ANY(%(tiendas_ids)s)
              AND fecha_venta >= CURRENT_DATE - INTERVAL '30 days'
              AND fecha_venta < CURRENT_DATE
              AND NOT (ubicacion_id = 'tienda_18' AND fecha_venta::date = '2025-12-06')
            GROUP BY producto_id, ubicacion_id, fecha_venta::date
        ),
        p75_por_tienda AS (
            SELECT
                producto_id,
                ubicacion_id,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY cantidad_dia) as p75_tienda,
                STDDEV(cantidad_dia) as sigma_tienda,
                MAX(cantidad_dia) as max_tienda
            FROM ventas_30d
            GROUP BY producto_id, ubicacion_id
        ),
        demanda_regional AS (
            SELECT
                producto_id,
                SUM(COALESCE(p75_tienda, 0)) as p75_regional,
                SQRT(SUM(COALESCE(sigma_tienda * sigma_tienda, 0))) as sigma_regional,
                SUM(COALESCE(max_tienda, 0)) as demanda_maxima_regional,
                COUNT(DISTINCT ubicacion_id) as num_tiendas
            FROM p75_por_tienda
            GROUP BY producto_id
            HAVING SUM(COALESCE(p75_tienda, 0)) > 0
        ),
        stock_cedi_caracas AS (
            SELECT producto_id, SUM(cantidad) as stock
            FROM inventario_actual
            WHERE ubicacion_id = 'cedi_caracas'
            GROUP BY producto_id
        ),
        stock_cedi_origen AS (
            SELECT producto_id, SUM(cantidad) as stock
            FROM inventario_actual
            WHERE ubicacion_id = %(cedi_origen)s
            GROUP BY producto_id
        ),
        stock_tiendas_region AS (
            SELECT producto_id, SUM(cantidad) as stock_total
            FROM inventario_actual
            WHERE ubicacion_id = ANY(%(tiendas_ids)s)
            GROUP BY producto_id
        ),
        abc_tiendas_ccs AS (
            -- ABC de cada tienda de Caracas (Artigas + Paraíso)
            SELECT
                producto_id,
                ubicacion_id,
                clase_abc
            FROM productos_abc_tienda
            WHERE ubicacion_id = ANY(%(tiendas_ids)s)
        ),
        abc_cedi_ccs AS (
            -- ABC más crítico para CEDI Caracas (si alguna tienda es A → CEDI es A)
            SELECT
                producto_id,
                -- Tomar el ABC con mayor criticidad (A=1, B=2, C=3, D=4)
                CASE
                    WHEN MIN(
                        CASE clase_abc
                            WHEN 'A' THEN 1
                            WHEN 'B' THEN 2
                            WHEN 'C' THEN 3
                            WHEN 'D' THEN 4
                            ELSE 5
                        END
                    ) = 1 THEN 'A'
                    WHEN MIN(
                        CASE clase_abc
                            WHEN 'A' THEN 1
                            WHEN 'B' THEN 2
                            WHEN 'C' THEN 3
                            WHEN 'D' THEN 4
                            ELSE 5
                        END
                    ) = 2 THEN 'B'
                    WHEN MIN(
                        CASE clase_abc
                            WHEN 'A' THEN 1
                            WHEN 'B' THEN 2
                            WHEN 'C' THEN 3
                            WHEN 'D' THEN 4
                            ELSE 5
                        END
                    ) = 3 THEN 'C'
                    ELSE 'D'
                END as clase_abc_cedi
            FROM abc_tiendas_ccs
            GROUP BY producto_id
        )
        SELECT
            p.id as producto_id,
            p.codigo,
            p.descripcion,
            p.categoria,
            p.cuadrante,
            COALESCE(p.unidades_por_bulto, 1) as unidades_por_bulto,
            COALESCE(dr.p75_regional, 0) as p75_regional,
            COALESCE(dr.sigma_regional, dr.p75_regional * 0.3) as sigma_regional,
            COALESCE(dr.demanda_maxima_regional, 0) as demanda_maxima_regional,
            COALESCE(dr.num_tiendas, 0) as num_tiendas,
            COALESCE(scc.stock, 0) as stock_cedi_caracas,
            COALESCE(sco.stock, 0) as stock_cedi_origen,
            COALESCE(str.stock_total, 0) as stock_tiendas_region,
            p.cedi_origen_id,
            p.codigo_barras,
            p.peso_unitario,
            COALESCE(abc_cedi.clase_abc_cedi, 'D') as clase_abc
        FROM productos p
        INNER JOIN demanda_regional dr ON p.id = dr.producto_id
        LEFT JOIN stock_cedi_caracas scc ON p.id = scc.producto_id
        LEFT JOIN stock_cedi_origen sco ON p.id = sco.producto_id
        LEFT JOIN stock_tiendas_region str ON p.id = str.producto_id
        LEFT JOIN abc_cedi_ccs abc_cedi ON p.id = abc_cedi.producto_id
        WHERE p.activo = true
          AND p.cedi_origen_id = %(cedi_origen)s
    """

    params: Dict[str, Any] = {
        'tiendas_ids': tiendas_ids,
        'cedi_origen': cedi_origen,
    }

    # Filtro de cuadrantes
    if filtros and filtros.get('cuadrantes'):
        query += " AND p.cuadrante = ANY(%(cuadrantes)s)"
        params['cuadrantes'] = filtros['cuadrantes']

    query += " ORDER BY COALESCE(dr.p75_regional, 0) DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]

    # 3. Cargar configuración de cobertura por categoría (perecederos)
    config_cobertura_categoria = {}
    try:
        cursor.execute("""
            SELECT categoria_normalizada, dias_cobertura_a, dias_cobertura_b,
                   dias_cobertura_c, dias_cobertura_d
            FROM config_cobertura_categoria
            WHERE activo = true
        """)
        for cat_row in cursor.fetchall():
            cat_norm = (cat_row[0] or '').strip().upper()
            config_cobertura_categoria[cat_norm] = {
                'A': cat_row[1] if cat_row[1] is not None else 7,
                'B': cat_row[2] if cat_row[2] is not None else 14,
                'C': cat_row[3] if cat_row[3] is not None else 21,
                'D': cat_row[4] if cat_row[4] is not None else 30
            }
    except Exception as e:
        logger.warning(f"No se pudo cargar config cobertura por categoría (CEDI CCS): {e}")

    cursor.close()

    # 4. Calcular productos
    productos = []
    for row in rows:
        row_dict = dict(zip(columns, row))

        producto_id = row_dict['producto_id']
        codigo = row_dict['codigo']
        descripcion = row_dict['descripcion']
        categoria = row_dict['categoria'] or ''
        cuadrante = row_dict['cuadrante']
        unidades_por_bulto = int(row_dict['unidades_por_bulto']) or 1
        p75_regional = float(row_dict['p75_regional'])
        sigma_regional = float(row_dict['sigma_regional'] or p75_regional * 0.3)
        demanda_maxima = float(row_dict['demanda_maxima_regional'] or p75_regional * 2)
        stock_cedi_ccs = float(row_dict['stock_cedi_caracas'])
        stock_cedi_orig = float(row_dict['stock_cedi_origen'])
        stock_tiendas = float(row_dict['stock_tiendas_region'])
        cedi_origen_id = row_dict['cedi_origen_id'] or cedi_origen
        clase_abc = row_dict['clase_abc'] or 'D'
        num_tiendas = int(row_dict['num_tiendas'])

        # Determinar override de cobertura para perecederos
        dias_cobertura_override = None
        cat_upper = categoria.strip().upper()
        es_fruver = cedi_origen_id == 'cedi_verde' or 'FRUVER' in cat_upper or 'FRUT' in cat_upper or 'VERDUR' in cat_upper
        es_panaderia = 'PANAD' in cat_upper or 'PAN ' in cat_upper or cat_upper.startswith('PAN')

        if es_fruver or es_panaderia:
            # Perecederos: cobertura fija de 1 día (misma lógica que inter-CEDI)
            dias_cobertura_override = 1
        elif cat_upper in config_cobertura_categoria:
            dias_cobertura_override = config_cobertura_categoria[cat_upper].get(clase_abc)

        # Usar calcular_inventario_simple con config de cedi_caracas
        resultado = calcular_inventario_simple(
            demanda_p75=p75_regional,
            sigma_demanda=sigma_regional,
            demanda_maxima=demanda_maxima if demanda_maxima > 0 else p75_regional * 2,
            unidades_por_bulto=unidades_por_bulto,
            stock_actual=stock_cedi_ccs,  # Stock en CEDI Caracas
            stock_cedi=stock_cedi_orig,   # Stock en CEDI origen (Valencia)
            clase_abc=clase_abc,
            es_generador_trafico=False,
            dias_cobertura_override=dias_cobertura_override,
            config_tienda=config_tienda_abc
        )

        punto_reorden_unid = resultado.punto_reorden_unid
        stock_maximo_unid = resultado.stock_maximo_unid
        stock_seguridad_unid = resultado.stock_seguridad_unid
        cantidad_sugerida_unid = resultado.cantidad_sugerida_unid
        cantidad_sugerida_bultos = resultado.cantidad_sugerida_bultos
        tiene_sobrestock = resultado.tiene_sobrestock

        # Calcular días de stock
        if p75_regional > 0:
            dias_stock = stock_cedi_ccs / p75_regional
            dias_stock_tiendas = stock_tiendas / p75_regional
            dias_ss = stock_seguridad_unid / p75_regional
            dias_rop = punto_reorden_unid / p75_regional
            dias_max = stock_maximo_unid / p75_regional
        else:
            dias_stock = 999 if stock_cedi_ccs > 0 else 0
            dias_stock_tiendas = 999 if stock_tiendas > 0 else 0
            dias_ss = 0
            dias_rop = 0
            dias_max = 0

        # Determinar si necesita reposición
        necesita_reposicion = (
            cantidad_sugerida_bultos > 0 and
            stock_cedi_orig > 0 and
            not tiene_sobrestock
        )

        if not necesita_reposicion:
            cantidad_sugerida_unid = 0
            cantidad_sugerida_bultos = 0

        # Calcular prioridad (misma lógica que inter-CEDI)
        if dias_stock <= dias_ss:
            prioridad = 1  # Crítico
        elif dias_stock <= dias_rop:
            prioridad = 2  # Urgente
        else:
            prioridad = 3  # Normal

        productos.append({
            'producto_id': producto_id,
            'codigo_producto': codigo,
            'descripcion_producto': descripcion,
            'categoria': categoria,
            'cuadrante': cuadrante,
            'clasificacion_abc': clase_abc,
            'unidades_por_bulto': unidades_por_bulto,
            'prom_p75_unid': p75_regional,
            'prom_20dias_unid': p75_regional,
            'stock_tienda': stock_cedi_ccs,  # Para DPD+U: "stock_tienda" = stock en CEDI CCS
            'stock_cedi_origen': stock_cedi_orig,
            'dias_stock': round(dias_stock, 2),
            'dias_ss': round(dias_ss, 1),
            'dias_rop': round(dias_rop, 1),
            'dias_max': round(dias_max, 1),
            'stock_seguridad_unid': stock_seguridad_unid,
            'punto_reorden_unid': punto_reorden_unid,
            'stock_maximo_unid': stock_maximo_unid,
            'cantidad_necesaria_unid': cantidad_sugerida_unid,
            'cantidad_sugerida_bultos': int(cantidad_sugerida_bultos),
            'cantidad_sugerida_unid': cantidad_sugerida_unid,
            'es_sugerido': necesita_reposicion,
            'transito_bultos': 0,
            'transito_desglose': [],
            # Campos extra para la vista inter-CEDI en el frontend
            'cedi_origen_id': cedi_origen_id,
            'stock_cedi_caracas': stock_cedi_ccs,
            'dias_cobertura_ccs': round(dias_stock, 1),
            'stock_tiendas_region': stock_tiendas,
            'dias_cobertura_tiendas': round(dias_stock_tiendas, 1),
            'prioridad': prioridad,
            'num_tiendas_region': num_tiendas,
            'codigo_barras': row_dict.get('codigo_barras'),
            'peso_kg': float(row_dict['peso_unitario']) if row_dict.get('peso_unitario') else None,
        })

    logger.info(f"📦 CEDI Caracas ({cedi_origen}): {len(productos)} productos, {sum(1 for p in productos if p['es_sugerido'])} sugeridos")
    return productos


def _calcular_cedi_caracas_worker(
    cedi_origen: str,
    dias_cobertura: int,
    filtros: Optional[Dict[str, Any]],
) -> Tuple[str, List[Dict]]:
    """Worker para CEDI Caracas: calcula demanda regional en thread separado."""
    _ts = _time.time()
    with get_db_connection_resilient() as thread_conn:
        loop = asyncio.new_event_loop()
        config = loop.run_until_complete(
            cargar_config_tienda(thread_conn, 'cedi_caracas', set_global=False)
        )
        productos = loop.run_until_complete(
            obtener_productos_cedi_caracas(
                thread_conn,
                cedi_origen,
                dias_cobertura,
                filtros=filtros,
                config_tienda_abc=config
            )
        )
        loop.close()
    logger.info(f"⏱️ CEDI Caracas worker ({cedi_origen}): {len(productos)} productos en {_time.time()-_ts:.1f}s")
    return ('cedi_caracas', productos)


# =====================================================================================
# WORKER: Cálculo por tienda en thread separado (para paralelismo)
# =====================================================================================

def _calcular_tienda_worker(
    cedi_origen: str,
    tienda_id: str,
    tienda_nombre: str,
    dias_cobertura: int,
    filtros: Optional[Dict[str, Any]],
    idx: int,
    total: int
) -> Tuple[str, List[Dict]]:
    """
    Worker sincrónico que calcula productos para una tienda.
    Cada worker obtiene su propia conexión a BD (thread-safe).
    """
    _ts = _time.time()
    with get_db_connection_resilient() as thread_conn:
        # Cargar config tienda (sin set global - thread-safe)
        loop = asyncio.new_event_loop()
        config = loop.run_until_complete(
            cargar_config_tienda(thread_conn, tienda_id, set_global=False)
        )
        productos = loop.run_until_complete(
            obtener_productos_tienda(
                thread_conn,
                cedi_origen,
                tienda_id,
                dias_cobertura,
                filtros=filtros,
                config_tienda_abc=config
            )
        )
        loop.close()
    logger.info(f"⏱️ Tienda {idx}/{total} {tienda_id}: {len(productos)} productos en {_time.time()-_ts:.1f}s")
    return (tienda_id, productos)


# =====================================================================================
# ENDPOINT: CALCULAR PEDIDOS MULTI-TIENDA
# =====================================================================================

@router.post("/calcular")
async def calcular_pedidos_multitienda(
    request: CalcularMultiTiendaRequest,
    cuadrantes: Optional[List[str]] = Query(None, description="Filtrar por cuadrantes"),
    conn: Any = Depends(get_db)
):
    """
    Calcula pedidos para múltiples tiendas desde un CEDI.

    1. Calcula el pedido sugerido para cada tienda
    2. Detecta productos con conflicto (stock CEDI < suma de necesidades)
    3. Aplica algoritmo DPD+U para distribuir stock limitado
    4. Retorna pedidos ajustados y lista de conflictos

    El paso 2 (resolución de conflictos) en el frontend solo aparece si:
    - Hay más de 1 tienda seleccionada
    - Existen conflictos de stock
    """
    try:
        # Obtener nombre del CEDI
        cursor = conn.cursor()
        cursor.execute("SELECT nombre FROM ubicaciones WHERE id = %s", (request.cedi_origen,))
        cedi_row = cursor.fetchone()
        cedi_nombre = cedi_row[0] if cedi_row else request.cedi_origen
        cursor.close()

        # Obtener configuración DPD+U
        config_dpdu = await obtener_config_dpdu(conn)

        # Calcular productos para cada tienda EN PARALELO
        # Cada tienda se procesa en un thread separado con su propia conexión a BD
        _t0 = _time.time()
        total_tiendas = len(request.tiendas_destino)
        filtros = {'cuadrantes': cuadrantes} if cuadrantes else None

        # Usar ThreadPoolExecutor para paralelizar (psycopg2 es sync, funciona en threads)
        # max_workers=6 para no saturar las conexiones de BD
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor(max_workers=min(6, total_tiendas)) as executor:
            futures = [
                loop.run_in_executor(
                    executor,
                    _calcular_tienda_worker,
                    request.cedi_origen,
                    tienda.tienda_id,
                    tienda.tienda_nombre,
                    request.dias_cobertura,
                    filtros,
                    i + 1,
                    total_tiendas
                )
                for i, tienda in enumerate(request.tiendas_destino)
            ]
            results = await asyncio.gather(*futures)

        productos_por_tienda: Dict[str, List[Dict]] = dict(results)
        logger.info(f"⏱️ Total cálculo tiendas (paralelo): {_time.time()-_t0:.1f}s | {total_tiendas} tiendas")

        # Si se incluye CEDI Caracas, calcular su demanda regional
        incluir_cedi_ccs = request.incluir_cedi_caracas
        if incluir_cedi_ccs:
            _t_ccs = _time.time()
            # Ejecutar en thread separado (igual que tiendas) para evitar
            # conflicto con el event loop de FastAPI
            cedi_ccs_id, cedi_ccs_productos = await loop.run_in_executor(
                None,  # default executor
                _calcular_cedi_caracas_worker,
                request.cedi_origen, request.dias_cobertura, filtros
            )
            productos_por_tienda[cedi_ccs_id] = cedi_ccs_productos
            logger.info(f"⏱️ CEDI Caracas incluido: {len(cedi_ccs_productos)} productos en {_time.time()-_t_ccs:.1f}s")

        # Consolidar todos los productos únicos
        _tc = _time.time()
        total_productos_raw = sum(len(p) for p in productos_por_tienda.values())
        logger.info(f"⏱️ Consolidando {total_productos_raw} productos de {len(productos_por_tienda)} tiendas...")
        todos_productos: Dict[str, Dict] = {}
        # Pre-build tienda name lookup
        _tienda_nombres = {t.tienda_id: t.tienda_nombre for t in request.tiendas_destino}
        if incluir_cedi_ccs:
            _tienda_nombres['cedi_caracas'] = 'CEDI CARACAS'
        for tienda_id, productos in productos_por_tienda.items():
            for prod in productos:
                codigo = prod['codigo_producto']
                if codigo not in todos_productos:
                    todos_productos[codigo] = {
                        'codigo_producto': codigo,
                        'descripcion_producto': prod['descripcion_producto'],
                        'categoria': prod['categoria'],
                        'clasificacion_abc': prod['clasificacion_abc'],
                        'unidades_por_bulto': prod['unidades_por_bulto'],
                        'stock_cedi': prod['stock_cedi_origen'],
                        'tiendas': {}
                    }
                todos_productos[codigo]['tiendas'][tienda_id] = {
                    'tienda_nombre': _tienda_nombres.get(tienda_id, tienda_id),
                    'clasificacion_abc': prod['clasificacion_abc'],  # ABC específico de esta tienda
                    'demanda_p75': prod['prom_p75_unid'],
                    'stock_tienda': prod['stock_tienda'],
                    'dias_stock': prod['dias_stock'],
                    'cantidad_necesaria': prod['cantidad_necesaria_unid'],
                    'cantidad_sugerida_bultos': prod['cantidad_sugerida_bultos'],
                    'transito_bultos': prod.get('transito_bultos', 0),
                    'transito_desglose': prod.get('transito_desglose', []),
                }

        logger.info(f"⏱️ Consolidación: {len(todos_productos)} productos únicos en {_time.time()-_tc:.1f}s")

        # Detectar conflictos y aplicar DPD+U
        _td = _time.time()
        conflictos = []
        productos_sin_conflicto = 0

        for codigo, prod_data in todos_productos.items():
            stock_cedi = prod_data['stock_cedi']
            tiendas_data = prod_data['tiendas']

            # Solo hay conflicto si más de 1 tienda necesita el producto
            if len(tiendas_data) <= 1:
                productos_sin_conflicto += 1
                continue

            # Calcular necesidad total
            necesidad_total = sum(t['cantidad_necesaria'] for t in tiendas_data.values())

            # ¿Es conflicto?
            es_conflicto = stock_cedi < necesidad_total

            total_destinos = len(request.tiendas_destino) + (1 if incluir_cedi_ccs else 0)
            if es_conflicto and total_destinos > 1:
                # Preparar datos para DPD+U — solo tiendas que necesitan reposición
                # Excluir tiendas con cantidad_necesaria = 0 (stock por encima del ROP)
                # para no asignarles stock que no necesitan, robándolo de tiendas urgentes
                datos_tiendas = [
                    DatosTiendaProducto(
                        tienda_id=tid,
                        tienda_nombre=tdata['tienda_nombre'],
                        demanda_p75=tdata['demanda_p75'],
                        stock_actual=tdata['stock_tienda'],
                        dias_stock=tdata['dias_stock'],
                        cantidad_necesaria=tdata['cantidad_necesaria']
                    )
                    for tid, tdata in tiendas_data.items()
                    if tdata['cantidad_necesaria'] > 0
                ]

                # Si después de filtrar no quedan tiendas con necesidad, no hay conflicto real
                if len(datos_tiendas) <= 1:
                    productos_sin_conflicto += 1
                    continue

                # Aplicar DPD+U
                asignaciones = calcular_distribucion_dpdu(
                    stock_cedi=stock_cedi,
                    unidades_por_bulto=prod_data['unidades_por_bulto'],
                    datos_tiendas=datos_tiendas,
                    config=config_dpdu
                )

                # Crear distribución DPD+U con datos de tránsito
                distribucion_con_transito = []
                for a in asignaciones:
                    tienda_data = tiendas_data.get(a.tienda_id, {})
                    distribucion_con_transito.append(AsignacionTiendaResponse(
                        tienda_id=a.tienda_id,
                        tienda_nombre=a.tienda_nombre,
                        abc=tienda_data.get('clasificacion_abc', 'D'),  # ABC específico de esta tienda
                        demanda_p75=a.demanda_p75,
                        stock_actual=a.stock_actual,
                        dias_stock=a.dias_stock,
                        cantidad_necesaria=a.cantidad_necesaria,
                        transito_bultos=tienda_data.get('transito_bultos', 0),
                        transito_desglose=tienda_data.get('transito_desglose'),
                        urgencia=a.urgencia,
                        pct_demanda=a.pct_demanda,
                        pct_urgencia=a.pct_urgencia,
                        peso_final=a.peso_final,
                        cantidad_asignada_unid=a.cantidad_asignada_unid,
                        cantidad_asignada_bultos=a.cantidad_asignada_bultos,
                        deficit_vs_necesidad=a.deficit_vs_necesidad,
                        cobertura_dias_resultante=a.cobertura_dias_resultante
                    ))

                conflictos.append(ConflictoProductoResponse(
                    codigo_producto=codigo,
                    descripcion_producto=prod_data['descripcion_producto'],
                    categoria=prod_data['categoria'],
                    clasificacion_abc=prod_data['clasificacion_abc'],
                    unidades_por_bulto=prod_data['unidades_por_bulto'],
                    stock_cedi_disponible=stock_cedi,
                    stock_cedi_bultos=math.floor(stock_cedi / prod_data['unidades_por_bulto']),
                    demanda_total_tiendas=sum(d.demanda_p75 for d in datos_tiendas),
                    necesidad_total_tiendas=necesidad_total,
                    es_conflicto=True,
                    distribucion_dpdu=distribucion_con_transito
                ))
            else:
                productos_sin_conflicto += 1

        logger.info(f"⏱️ Detección conflictos: {len(conflictos)} conflictos en {_time.time()-_td:.1f}s")

        # Construir pedidos por tienda con cantidades ajustadas
        _te = _time.time()
        productos_por_tienda_dict = productos_por_tienda  # Guardar referencia al dict
        pedidos_por_tienda = []

        # Pre-build conflict lookup dicts for O(1) access
        conflictos_dict: Dict[str, Any] = {c.codigo_producto: c for c in conflictos}
        conflictos_asignaciones: Dict[str, Dict[str, Any]] = {}
        for c in conflictos:
            conflictos_asignaciones[c.codigo_producto] = {a.tienda_id: a for a in c.distribucion_dpdu}

        for tienda in request.tiendas_destino:
            productos_tienda = productos_por_tienda.get(tienda.tienda_id, [])

            # Ajustar cantidades según DPD+U
            productos_ajustados = []
            productos_ajustados_count = 0
            total_bultos_tienda = 0
            total_unidades_tienda = 0
            total_sugeridos_tienda = 0

            for prod in productos_tienda:
                codigo = prod['codigo_producto']
                cantidad_original = prod['cantidad_sugerida_bultos']
                cantidad_final = cantidad_original
                ajustado = False

                # Buscar si este producto tiene conflicto (O(1) dict lookup)
                conflicto = conflictos_dict.get(codigo)
                if conflicto:
                    # Buscar asignación para esta tienda (O(1) dict lookup)
                    asignacion = conflictos_asignaciones[codigo].get(tienda.tienda_id)
                    if asignacion:
                        cantidad_final = asignacion.cantidad_asignada_bultos
                        ajustado = cantidad_final != cantidad_original
                        if ajustado:
                            productos_ajustados_count += 1

                prod_dict = {
                    'codigo_producto': codigo,
                    'descripcion_producto': prod['descripcion_producto'],
                    'categoria': prod['categoria'],
                    'clasificacion_abc': prod['clasificacion_abc'],
                    'cuadrante': prod.get('cuadrante'),
                    'unidades_por_bulto': prod['unidades_por_bulto'],
                    'cantidad_sugerida_unid': cantidad_final * prod['unidades_por_bulto'],
                    'cantidad_sugerida_bultos': cantidad_final,
                    'stock_tienda': prod['stock_tienda'],
                    'stock_cedi_origen': prod['stock_cedi_origen'],
                    'transito_bultos': prod.get('transito_bultos', 0),
                    'transito_desglose': prod.get('transito_desglose'),
                    'prom_p75_unid': prod['prom_p75_unid'],
                    'dias_stock': prod['dias_stock'],
                    'ajustado_por_dpdu': ajustado,
                    'cantidad_original_bultos': cantidad_original if ajustado else None,
                    'es_sugerido': prod.get('es_sugerido', True),
                    'peso_kg': prod.get('peso_kg'),
                }
                productos_ajustados.append(prod_dict)
                if prod_dict['es_sugerido']:
                    total_bultos_tienda += prod_dict['cantidad_sugerida_bultos']
                    total_unidades_tienda += prod_dict['cantidad_sugerida_unid']
                    total_sugeridos_tienda += 1

            pedidos_por_tienda.append({
                'tienda_id': tienda.tienda_id,
                'tienda_nombre': tienda.tienda_nombre,
                'productos': productos_ajustados,
                'total_productos': total_sugeridos_tienda,
                'total_bultos': total_bultos_tienda,
                'total_unidades': total_unidades_tienda,
                'productos_ajustados_dpdu': productos_ajustados_count,
            })

        # Agregar pedido de CEDI Caracas (si se incluyó)
        if incluir_cedi_ccs:
            productos_cedi_ccs = productos_por_tienda_dict.get('cedi_caracas', [])
            productos_ajustados_ccs = []
            ajustados_count_ccs = 0
            total_bultos_ccs = 0
            total_unidades_ccs = 0
            total_sugeridos_ccs = 0

            for prod in productos_cedi_ccs:
                codigo = prod['codigo_producto']
                cantidad_original = prod['cantidad_sugerida_bultos']
                cantidad_final = cantidad_original
                ajustado = False

                conflicto = conflictos_dict.get(codigo)
                if conflicto:
                    asignacion = conflictos_asignaciones[codigo].get('cedi_caracas')
                    if asignacion:
                        cantidad_final = asignacion.cantidad_asignada_bultos
                        ajustado = cantidad_final != cantidad_original
                        if ajustado:
                            ajustados_count_ccs += 1

                prod_dict = {
                    'codigo_producto': codigo,
                    'descripcion_producto': prod['descripcion_producto'],
                    'categoria': prod['categoria'],
                    'clasificacion_abc': prod['clasificacion_abc'],
                    'cuadrante': prod.get('cuadrante'),
                    'unidades_por_bulto': prod['unidades_por_bulto'],
                    'cantidad_sugerida_unid': cantidad_final * prod['unidades_por_bulto'],
                    'cantidad_sugerida_bultos': cantidad_final,
                    'stock_tienda': prod['stock_tienda'],
                    'stock_cedi_origen': prod['stock_cedi_origen'],
                    'transito_bultos': prod.get('transito_bultos', 0),
                    'transito_desglose': prod.get('transito_desglose'),
                    'prom_p75_unid': prod['prom_p75_unid'],
                    'dias_stock': prod['dias_stock'],
                    'ajustado_por_dpdu': ajustado,
                    'cantidad_original_bultos': cantidad_original if ajustado else None,
                    'es_sugerido': prod.get('es_sugerido', True),
                    # Campos extra para vista inter-CEDI en frontend
                    'cedi_origen_id': prod.get('cedi_origen_id'),
                    'stock_cedi_caracas': prod.get('stock_cedi_caracas'),
                    'dias_cobertura_ccs': prod.get('dias_cobertura_ccs'),
                    'stock_tiendas_region': prod.get('stock_tiendas_region'),
                    'dias_cobertura_tiendas': prod.get('dias_cobertura_tiendas'),
                    'prioridad': prod.get('prioridad'),
                    'num_tiendas_region': prod.get('num_tiendas_region'),
                    'codigo_barras': prod.get('codigo_barras'),
                    'peso_kg': prod.get('peso_kg'),
                }
                productos_ajustados_ccs.append(prod_dict)
                if prod_dict['es_sugerido']:
                    total_bultos_ccs += prod_dict['cantidad_sugerida_bultos']
                    total_unidades_ccs += prod_dict['cantidad_sugerida_unid']
                    total_sugeridos_ccs += 1

            pedidos_por_tienda.append({
                'tienda_id': 'cedi_caracas',
                'tienda_nombre': 'CEDI CARACAS',
                'productos': productos_ajustados_ccs,
                'total_productos': total_sugeridos_ccs,
                'total_bultos': total_bultos_ccs,
                'total_unidades': total_unidades_ccs,
                'productos_ajustados_dpdu': ajustados_count_ccs,
                'es_cedi': True,
            })

        # Ordenar conflictos por stock_cedi (menor primero = más crítico)
        conflictos.sort(key=lambda c: c.stock_cedi_disponible)

        total_prods_response = sum(len(p['productos']) for p in pedidos_por_tienda)
        logger.info(f"⏱️ Ensamblaje pedidos: {total_prods_response} productos en {_time.time()-_te:.1f}s")
        logger.info(f"⏱️ TOTAL calcular: {_time.time()-_t0:.1f}s | {len(request.tiendas_destino)} tiendas, {len(todos_productos)} productos únicos, {len(conflictos)} conflictos, {total_prods_response} items en respuesta")

        # Serialize conflictos (Pydantic models) to dicts once
        _ts = _time.time()
        conflictos_dicts = [c.model_dump() for c in conflictos]

        response_data = {
            'cedi_origen': request.cedi_origen,
            'cedi_origen_nombre': cedi_nombre,
            'fecha_calculo': datetime.now(VENEZUELA_TZ).isoformat(),
            'dias_cobertura': request.dias_cobertura,
            'config_dpdu': {
                'peso_demanda': config_dpdu.peso_demanda,
                'peso_urgencia': config_dpdu.peso_urgencia,
                'dias_minimo_urgencia': config_dpdu.dias_minimo_urgencia,
            },
            'conflictos': conflictos_dicts,
            'total_conflictos': len(conflictos),
            'productos_sin_conflicto': productos_sin_conflicto,
            'pedidos_por_tienda': pedidos_por_tienda,
            'resumen': {
                'total_tiendas': len(request.tiendas_destino),
                'total_productos_unicos': len(todos_productos),
                'total_conflictos': len(conflictos),
                'total_bultos': sum(p['total_bultos'] for p in pedidos_por_tienda),
                'incluye_cedi_caracas': incluir_cedi_ccs,
            },
        }
        logger.info(f"⏱️ Serialización: {_time.time()-_ts:.1f}s")

        return JSONResponse(content=response_data)

    except Exception as e:
        logger.error(f"Error calculando pedidos multi-tienda: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


# =====================================================================================
# ENDPOINT: GUARDAR PEDIDOS EN LOTE
# =====================================================================================

@router.post("/guardar-lote", response_model=GuardarMultiTiendaResponse)
async def guardar_pedidos_lote(
    request: GuardarMultiTiendaRequest,
    conn: Any = Depends(get_db_write)
):
    """
    Guarda múltiples pedidos en una sola transacción.

    - Genera un grupo_pedido_id para agrupar los pedidos
    - Números de pedido correlacionados: PS-XXXXX-1, PS-XXXXX-2, etc.
    - Rollback automático si falla alguno
    """
    try:
        cursor = conn.cursor()

        # Generar ID de grupo
        grupo_id = f"GRUPO-{uuid.uuid4().hex[:8].upper()}"

        # Obtener siguiente número base de pedido
        cursor.execute("SELECT MAX(numero_pedido) FROM pedidos_sugeridos WHERE numero_pedido LIKE 'PS-%'")
        ultimo = cursor.fetchone()[0]
        if ultimo:
            try:
                # Extraer número base (puede ser PS-00001 o PS-00001-1)
                partes = ultimo.replace('PS-', '').split('-')
                num_base = int(partes[0]) + 1
            except:
                num_base = 1
        else:
            num_base = 1

        fecha_pedido = request.fecha_pedido or datetime.now(VENEZUELA_TZ).strftime('%Y-%m-%d')
        pedidos_creados = []

        for idx, pedido in enumerate(request.pedidos, 1):
            pedido_id = f"ped_{uuid.uuid4().hex[:12]}"
            numero_pedido = f"PS-{num_base:05d}-{idx}"

            # Filtrar solo productos incluidos
            productos_incluidos = [p for p in pedido.productos if p.incluido]

            if not productos_incluidos:
                continue

            total_productos = len(productos_incluidos)
            total_bultos = sum(p.cantidad_pedida_bultos for p in productos_incluidos)
            total_unidades = sum(p.cantidad_pedida_unidades for p in productos_incluidos)

            # Insertar pedido
            cursor.execute("""
                INSERT INTO pedidos_sugeridos (
                    id, numero_pedido, version,
                    cedi_origen_id, cedi_origen_nombre,
                    tienda_destino_id, tienda_destino_nombre,
                    fecha_pedido, fecha_creacion,
                    estado, tipo_pedido, prioridad,
                    dias_cobertura,
                    total_productos, total_lineas, total_bultos, total_unidades,
                    grupo_pedido_id, orden_en_grupo,
                    usuario_creador
                ) VALUES (
                    %s, %s, 1,
                    %s, %s,
                    %s, %s,
                    %s, (CURRENT_TIMESTAMP AT TIME ZONE 'UTC'),
                    'borrador', %s, %s,
                    %s,
                    %s, %s, %s, %s,
                    %s, %s,
                    %s
                )
            """, (
                pedido_id, numero_pedido,
                request.cedi_origen_id, request.cedi_origen_nombre,
                pedido.tienda_destino_id, pedido.tienda_destino_nombre,
                fecha_pedido,
                request.tipo_pedido, request.prioridad,
                request.dias_cobertura,
                total_productos, total_productos, total_bultos, total_unidades,
                grupo_id, idx,
                request.usuario_creador
            ))

            # Insertar detalle de productos
            for linea_num, prod in enumerate(productos_incluidos, start=1):
                cursor.execute("""
                    INSERT INTO pedidos_sugeridos_detalle (
                        pedido_id, linea_numero, codigo_producto, descripcion_producto,
                        categoria, clasificacion_abc, cuadrante_producto,
                        cantidad_bultos, cantidad_sugerida_bultos,
                        cantidad_pedida_bultos, cantidad_pedida_unidades, total_unidades,
                        stock_tienda, stock_cedi_origen,
                        prom_ventas_5dias_unid, prom_ventas_20dias_unid,
                        prom_p75_unid,
                        stock_minimo, stock_maximo, punto_reorden,
                        razon_pedido
                    ) VALUES (
                        %s, %s, %s, %s,
                        %s, %s, %s,
                        %s, %s,
                        %s, %s, %s,
                        %s, %s,
                        %s, %s,
                        %s,
                        %s, %s, %s,
                        %s
                    )
                """, (
                    pedido_id, linea_num, prod.codigo_producto, prod.descripcion_producto,
                    prod.categoria, prod.clasificacion_abc, prod.cuadrante,
                    prod.unidades_por_bulto, prod.cantidad_pedida_bultos,
                    prod.cantidad_pedida_bultos, prod.cantidad_pedida_unidades, prod.cantidad_pedida_unidades,
                    prod.stock_tienda, prod.stock_cedi_origen,
                    prod.prom_ventas_5dias_unid, prod.prom_ventas_20dias_unid,
                    prod.prom_p75_unid,
                    prod.stock_minimo, prod.stock_maximo, prod.punto_reorden,
                    prod.razon_ajuste_dpdu if prod.ajustado_por_dpdu else None
                ))

            # Insertar en historial
            cursor.execute("""
                INSERT INTO pedidos_sugeridos_historial (
                    pedido_id, estado_anterior, estado_nuevo, motivo_cambio
                ) VALUES (%s, NULL, 'borrador', 'Pedido creado (multi-tienda)')
            """, (pedido_id,))

            pedidos_creados.append(PedidoGuardadoInfo(
                pedido_id=pedido_id,
                numero_pedido=numero_pedido,
                tienda_id=pedido.tienda_destino_id,
                tienda_nombre=pedido.tienda_destino_nombre,
                total_productos=total_productos,
                total_bultos=int(total_bultos),
                estado='borrador'
            ))

        conn.commit()
        cursor.close()

        return GuardarMultiTiendaResponse(
            success=True,
            grupo_pedido_id=grupo_id,
            pedidos_creados=pedidos_creados,
            total_pedidos=len(pedidos_creados),
            mensaje=f"Se crearon {len(pedidos_creados)} pedidos exitosamente"
        )

    except Exception as e:
        conn.rollback()
        logger.error(f"Error guardando pedidos multi-tienda: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


# =====================================================================================
# ENDPOINT: OBTENER CONFIGURACIÓN DPD+U
# =====================================================================================

@router.get("/config-dpdu", response_model=ConfigDPDUResponse)
async def get_config_dpdu(conn: Any = Depends(get_db)):
    """Obtiene la configuración actual del algoritmo DPD+U."""
    try:
        config = await obtener_config_dpdu(conn)
        return ConfigDPDUResponse(
            peso_demanda=config.peso_demanda,
            peso_urgencia=config.peso_urgencia,
            dias_minimo_urgencia=config.dias_minimo_urgencia
        )
    except Exception as e:
        logger.error(f"Error obteniendo config DPD+U: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


@router.get("/{pedido_id}/exportar-excel")
async def exportar_pedido_excel_multitienda(
    pedido_id: str,
    conn: Any = Depends(get_db)
):
    """
    Exporta pedido multi-tienda a Excel con columna Cuadrante.
    Similar a single-tienda con indicador de ajustes DPD+U.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        from fastapi.responses import Response

        cursor = conn.cursor()

        # Query similar a single-tienda, agregar grupo_pedido_id y razon_pedido
        cursor.execute("""
            SELECT numero_pedido, tienda_destino_nombre, cedi_origen_nombre,
                   fecha_pedido, estado, dias_cobertura, grupo_pedido_id
            FROM pedidos_sugeridos
            WHERE id = %s
        """, [pedido_id])

        pedido_row = cursor.fetchone()
        if not pedido_row:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        numero_pedido, tienda_nombre, cedi_nombre, fecha_pedido, estado, dias_cobertura, grupo_id = pedido_row

        cursor.execute("""
            SELECT
                codigo_producto, descripcion_producto, categoria,
                clasificacion_abc, cuadrante_producto, cantidad_bultos,
                stock_tienda, cantidad_sugerida_bultos, cantidad_pedida_bultos,
                total_unidades, razon_pedido
            FROM pedidos_sugeridos_detalle
            WHERE pedido_id = %s AND incluido = true
            ORDER BY clasificacion_abc, descripcion_producto
        """, [pedido_id])

        rows = cursor.fetchall()
        cursor.close()

        if not rows:
            raise HTTPException(status_code=404, detail="No hay productos")

        # Crear Excel (código similar a single-tienda)
        wb = openpyxl.Workbook()
        ws = wb.active

        # Título con indicador de grupo
        ws.merge_cells('A1:K1')
        titulo = f"Pedido Multi-Tienda: {numero_pedido}"
        if grupo_id:
            titulo += f" (Grupo: {grupo_id})"
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Info
        ws.merge_cells('A2:K2')
        ws['A2'] = f"Tienda: {tienda_nombre} | CEDI: {cedi_nombre} | Fecha: {fecha_pedido}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # Headers (agregar columna Observación)
        headers = [
            'Código', 'Descripción', 'Categoría', 'ABC', 'Cuadrante',
            'Unid/Bulto', 'Stock', 'Sugerido', 'Pedido', 'Total', 'Observación'
        ]

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Colores ABC
        abc_colors = {'A': 'E8F5E9', 'B': 'FFF9C4', 'C': 'FFE0B2', 'D': 'F5F5F5'}

        # Datos
        for row_idx, row_data in enumerate(rows, 5):
            abc = row_data[3] or 'D'
            razon = row_data[10] or ''
            observacion = 'Ajustado DPD+U' if 'dpdu' in razon.lower() or 'conflicto' in razon.lower() else ''

            data = [
                row_data[0], row_data[1], row_data[2], abc,
                row_data[4] or 'NO ESPECIFICADO',
                int(row_data[5]) if row_data[5] else 1,
                float(row_data[6]) if row_data[6] else 0,
                float(row_data[7]) if row_data[7] else 0,
                float(row_data[8]) if row_data[8] else 0,
                float(row_data[9]) if row_data[9] else 0,
                observacion
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill(start_color=abc_colors.get(abc, 'FFFFFF'),
                                       end_color=abc_colors.get(abc, 'FFFFFF'),
                                       fill_type='solid')
                if 6 <= col_idx <= 10:
                    cell.alignment = Alignment(horizontal='right')

        # Anchos
        column_widths = [12, 35, 18, 6, 15, 10, 10, 12, 12, 12, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{numero_pedido.replace('/', '-')}_multitienda.xlsx"

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
