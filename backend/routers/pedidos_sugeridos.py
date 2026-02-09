"""
Router para el workflow completo de Pedidos Sugeridos con Devoluciones
Migrado a PostgreSQL - Diciembre 2025

Incluye:
- Calcular pedido sugerido con devoluciones
- Guardar pedido en estado borrador
- Enviar para aprobación de gerente
- Aprobar/rechazar por gerente (con comentarios)
- Finalizar pedido y generar Excel
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from typing import List, Optional, Any
import uuid
from datetime import datetime, date, timezone
from zoneinfo import ZoneInfo
from decimal import Decimal
import logging
import math

from models.pedidos_sugeridos import (
    # Request/Response models
    CalcularPedidoRequest,
    CalcularPedidoResponse,
    GuardarPedidoRequest,
    PedidoGuardadoResponse,
    ActualizarEstadoPedidoRequest,
    CrearComentarioRequest,
    # Entity models
    ProductoPedidoSugeridoCalculado,
    ProductoDevolucionSugeridaCalculada,
    PedidoSugeridoCompleto,
    PedidoSugeridoResumen,
    PedidoHistorial,
    PedidoComentario,
    # Enums
    EstadoPedido,
    # Verificar Llegada
    EstadoLlegada,
    ProductoLlegadaVerificacion,
    VerificarLlegadaResponse,
    RegistrarLlegadaRequest,
    RegistrarLlegadaResponse,
)
from db_manager import get_db_connection, get_db_connection_write, get_db_connection_resilient
from services.calculo_inventario_abc import (
    calcular_inventario_simple,
    set_config_tienda,
    ConfigTiendaABC,
    LEAD_TIME_DEFAULT
)

logger = logging.getLogger(__name__)


def safe_float(value, default=0.0):
    """Safely convert a value to float, handling invalid strings like 'Bulto'"""
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


router = APIRouter(prefix="/api/pedidos-sugeridos", tags=["Pedidos Sugeridos - Workflow"])


def get_db():
    """Get database connection (read-only, resilient to replica conflicts)"""
    with get_db_connection_resilient() as conn:
        yield conn


def get_db_write():
    """Get database connection (read-write para escrituras)"""
    with get_db_connection_write() as conn:
        yield conn


# =====================================================================================
# ESTADO CONSTANTES (para compatibilidad con PostgreSQL)
# =====================================================================================

ESTADO_BORRADOR = "borrador"
ESTADO_PENDIENTE_APROBACION_GERENTE = "pendiente_aprobacion_gerente"
ESTADO_APROBADO_GERENTE = "aprobado_gerente"
ESTADO_RECHAZADO_GERENTE = "rechazado_gerente"

# Timezone de Venezuela (UTC-4)
VENEZUELA_TZ = ZoneInfo("America/Caracas")


def to_venezuela_aware(dt: datetime) -> datetime:
    """
    Convierte un datetime de UTC a hora Venezuela (UTC-4).

    PostgreSQL RDS está configurado en UTC, por lo que CURRENT_TIMESTAMP
    devuelve hora UTC. psycopg2 lo retorna como naive datetime (sin tzinfo).

    Esta función asume que el datetime naive está en UTC y lo convierte a Venezuela.
    """
    if dt is None:
        return None
    # Si no tiene timezone, asumimos que está en UTC (viene de PostgreSQL RDS)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    # Convertir a hora Venezuela
    return dt.astimezone(VENEZUELA_TZ)


# =====================================================================================
# LISTAR PEDIDOS
# =====================================================================================

@router.get("/", response_model=List[PedidoSugeridoResumen])
async def listar_pedidos(
    estado: Optional[str] = None,
    tienda_id: Optional[str] = None,
    conn: Any = Depends(get_db)
):
    """
    Lista pedidos con filtros opcionales

    Filtros:
    - estado: borrador, pendiente_aprobacion_gerente, aprobado_gerente, finalizado
    - tienda_id: filtrar por tienda destino
    """
    try:
        cursor = conn.cursor()

        # Query con conteo de productos y bultos por clasificación ABC + indicadores de llegada
        # Tolerancia: completo >= 97%, parcial > 0%, no_llego = 0%
        query = """
            SELECT
                ps.id, ps.numero_pedido, ps.fecha_pedido, ps.fecha_creacion,
                ps.cedi_origen_nombre, ps.tienda_destino_nombre,
                ps.estado, ps.prioridad, ps.tipo_pedido,
                ps.total_productos, ps.total_lineas, ps.total_bultos, ps.total_unidades, ps.total_peso_kg,
                ps.fecha_entrega_solicitada, ps.fecha_aprobacion, ps.fecha_recepcion,
                ps.usuario_creador,
                COALESCE(EXTRACT(DAY FROM AGE(CURRENT_DATE, ps.fecha_creacion::DATE)), 0)::INTEGER as dias_desde_creacion,
                COALESCE(abc_counts.productos_a, 0) as productos_a,
                COALESCE(abc_counts.productos_b, 0) as productos_b,
                COALESCE(abc_counts.productos_c, 0) as productos_c,
                COALESCE(abc_counts.productos_d, 0) as productos_d,
                COALESCE(abc_counts.bultos_a, 0) as bultos_a,
                COALESCE(abc_counts.bultos_b, 0) as bultos_b,
                COALESCE(abc_counts.bultos_c, 0) as bultos_c,
                COALESCE(abc_counts.bultos_d, 0) as bultos_d,
                -- Peso total calculado
                COALESCE(abc_counts.peso_total_calculado, ps.total_peso_kg, 0) as peso_total_kg,
                -- Indicadores de llegada
                COALESCE(llegada_stats.llegada_completos, 0) as llegada_completos,
                COALESCE(llegada_stats.llegada_parciales, 0) as llegada_parciales,
                COALESCE(llegada_stats.llegada_no_llegaron, 0) as llegada_no_llegaron,
                COALESCE(llegada_stats.tiene_verificacion, FALSE) as tiene_verificacion
            FROM pedidos_sugeridos ps
            LEFT JOIN LATERAL (
                SELECT
                    COUNT(*) FILTER (WHERE UPPER(d.clasificacion_abc) = 'A') as productos_a,
                    COUNT(*) FILTER (WHERE UPPER(d.clasificacion_abc) = 'B') as productos_b,
                    COUNT(*) FILTER (WHERE UPPER(d.clasificacion_abc) = 'C') as productos_c,
                    COUNT(*) FILTER (WHERE UPPER(d.clasificacion_abc) IN ('D', '') OR d.clasificacion_abc IS NULL) as productos_d,
                    SUM(d.cantidad_pedida_bultos) FILTER (WHERE UPPER(d.clasificacion_abc) = 'A') as bultos_a,
                    SUM(d.cantidad_pedida_bultos) FILTER (WHERE UPPER(d.clasificacion_abc) = 'B') as bultos_b,
                    SUM(d.cantidad_pedida_bultos) FILTER (WHERE UPPER(d.clasificacion_abc) = 'C') as bultos_c,
                    SUM(d.cantidad_pedida_bultos) FILTER (WHERE UPPER(d.clasificacion_abc) IN ('D', '') OR d.clasificacion_abc IS NULL) as bultos_d,
                    -- Calcular peso total en kg desde productos
                    SUM(COALESCE(d.peso_unitario_kg, p.peso_unitario, 0) * d.cantidad_pedida_unidades) as peso_total_calculado
                FROM pedidos_sugeridos_detalle d
                LEFT JOIN productos p ON p.codigo = d.codigo_producto
                WHERE d.pedido_id = ps.id AND d.incluido = true
            ) abc_counts ON true
            LEFT JOIN LATERAL (
                SELECT
                    -- Completo: recibido >= 97% de pedido
                    COUNT(*) FILTER (WHERE COALESCE(cantidad_recibida_bultos, 0) >= cantidad_pedida_bultos * 0.97 AND cantidad_pedida_bultos > 0) as llegada_completos,
                    -- Parcial: recibido > 0 pero < 97%
                    COUNT(*) FILTER (WHERE COALESCE(cantidad_recibida_bultos, 0) > 0 AND COALESCE(cantidad_recibida_bultos, 0) < cantidad_pedida_bultos * 0.97 AND cantidad_pedida_bultos > 0) as llegada_parciales,
                    -- No llegó: recibido = 0 o NULL
                    COUNT(*) FILTER (WHERE (COALESCE(cantidad_recibida_bultos, 0) = 0) AND cantidad_pedida_bultos > 0) as llegada_no_llegaron,
                    -- Tiene verificación si algún producto tiene cantidad_recibida > 0
                    (SUM(COALESCE(cantidad_recibida_bultos, 0)) > 0) as tiene_verificacion
                FROM pedidos_sugeridos_detalle
                WHERE pedido_id = ps.id AND incluido = true
            ) llegada_stats ON true
            WHERE 1=1
        """

        params = []
        if estado:
            query += " AND ps.estado = %s"
            params.append(estado)
        if tienda_id:
            query += " AND ps.tienda_destino_id = %s"
            params.append(tienda_id)

        query += " ORDER BY ps.fecha_creacion DESC LIMIT 100"

        cursor.execute(query, params if params else None)
        result = cursor.fetchall()
        cursor.close()

        pedidos = []
        for row in result:
            # Calculate porcentaje_avance
            estado_pedido = row[6]
            if estado_pedido == EstadoPedido.RECIBIDO:
                porcentaje = 100
            elif estado_pedido == EstadoPedido.EN_TRANSITO:
                porcentaje = 80
            elif estado_pedido == EstadoPedido.EN_PREPARACION:
                porcentaje = 60
            elif estado_pedido == EstadoPedido.APROBADO:
                porcentaje = 50
            elif estado_pedido == EstadoPedido.SOLICITADO:
                porcentaje = 30
            elif estado_pedido == ESTADO_BORRADOR:
                porcentaje = 10
            else:
                porcentaje = 0

            # Peso total calculado (row[27])
            peso_total_kg_calculado = safe_float(row[27]) if row[27] else None

            # Calcular porcentajes de llegada (shifted by 1 due to peso_total_kg)
            llegada_completos = row[28] or 0
            llegada_parciales = row[29] or 0
            llegada_no_llegaron = row[30] or 0
            tiene_verificacion = row[31] or False
            total_verificados = llegada_completos + llegada_parciales + llegada_no_llegaron

            if total_verificados > 0:
                pct_completos = (llegada_completos / total_verificados) * 100
                pct_parciales = (llegada_parciales / total_verificados) * 100
                pct_no_llegaron = (llegada_no_llegaron / total_verificados) * 100
            else:
                pct_completos = pct_parciales = pct_no_llegaron = 0.0

            pedidos.append(PedidoSugeridoResumen(
                id=row[0],
                numero_pedido=row[1],
                fecha_pedido=row[2],
                fecha_creacion=to_venezuela_aware(row[3]),
                cedi_origen_nombre=row[4] or "CEDI",
                tienda_destino_nombre=row[5] or "Tienda",
                estado=row[6],
                prioridad=row[7] or "normal",
                tipo_pedido=row[8] or "reposicion",
                total_productos=row[9] or 0,
                total_lineas=row[10] or 0,
                total_bultos=safe_float(row[11]),
                total_unidades=safe_float(row[12]),
                total_peso_kg=peso_total_kg_calculado,
                fecha_entrega_solicitada=row[14],
                fecha_aprobacion=to_venezuela_aware(row[15]),
                fecha_recepcion=to_venezuela_aware(row[16]),
                usuario_creador=row[17] or "sistema",
                dias_desde_creacion=row[18],
                porcentaje_avance=porcentaje,
                productos_a=row[19] or 0,
                productos_b=row[20] or 0,
                productos_c=row[21] or 0,
                productos_d=row[22] or 0,
                bultos_a=safe_float(row[23]),
                bultos_b=safe_float(row[24]),
                bultos_c=safe_float(row[25]),
                bultos_d=safe_float(row[26]),
                # Indicadores de llegada
                llegada_completos=llegada_completos,
                llegada_parciales=llegada_parciales,
                llegada_no_llegaron=llegada_no_llegaron,
                llegada_pct_completos=round(pct_completos, 1),
                llegada_pct_parciales=round(pct_parciales, 1),
                llegada_pct_no_llegaron=round(pct_no_llegaron, 1),
                tiene_verificacion_llegada=tiene_verificacion
            ))

        return pedidos

    except Exception as e:
        logger.error(f"❌ Error listando pedidos: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error listando pedidos: {str(e)}")


# =====================================================================================
# GUARDAR PEDIDO (estado: borrador)
# =====================================================================================

@router.post("/", response_model=PedidoGuardadoResponse)
async def crear_pedido(
    request: GuardarPedidoRequest,
    conn: Any = Depends(get_db_write)
):
    """
    Alias de /guardar para compatibilidad con frontend
    POST /api/pedidos-sugeridos
    """
    return await guardar_pedido(request, conn)


@router.post("/guardar", response_model=PedidoGuardadoResponse)
async def guardar_pedido(
    request: GuardarPedidoRequest,
    conn: Any = Depends(get_db_write)
):
    """
    Guarda pedido en estado BORRADOR
    """
    try:
        logger.info(f"💾 Guardando pedido: {request.tienda_destino_nombre}")
        cursor = conn.cursor()

        # Generar IDs
        pedido_id = str(uuid.uuid4())

        # Generar número de pedido
        cursor.execute("SELECT MAX(numero_pedido) FROM pedidos_sugeridos WHERE numero_pedido LIKE 'PS-%%'")
        result = cursor.fetchone()
        if result and result[0]:
            ultimo = int(result[0].split('-')[1])
            numero_pedido = f"PS-{ultimo + 1:05d}"
        else:
            numero_pedido = "PS-00001"

        # Filtrar productos incluidos
        productos_incluidos = [p for p in request.productos if p.incluido]
        devoluciones_incluidas = [d for d in request.devoluciones if d.incluido]

        # Calcular totales productos a recibir
        total_productos = len(productos_incluidos)
        total_bultos = sum(Decimal(str(p.cantidad_pedida_bultos)) for p in productos_incluidos)
        total_unidades = sum(Decimal(str(p.total_unidades)) for p in productos_incluidos)

        # Calcular totales devoluciones
        tiene_devoluciones = len(devoluciones_incluidas) > 0
        total_productos_devolucion = len(devoluciones_incluidas)
        total_bultos_devolucion = sum(Decimal(str(d.devolucion_confirmada_bultos)) for d in devoluciones_incluidas)
        total_unidades_devolucion = sum(Decimal(str(d.total_unidades_devolver)) for d in devoluciones_incluidas)

        # Insertar pedido principal
        cursor.execute("""
            INSERT INTO pedidos_sugeridos (
                id, numero_pedido,
                cedi_origen_id, cedi_origen_nombre,
                tienda_destino_id, tienda_destino_nombre,
                fecha_pedido, fecha_entrega_solicitada,
                estado, requiere_aprobacion,
                total_productos, total_bultos, total_unidades,
                tiene_devoluciones, total_productos_devolucion,
                total_bultos_devolucion, total_unidades_devolucion,
                dias_cobertura, tipo_pedido, prioridad,
                observaciones, notas_picking, notas_entrega,
                usuario_creador, fecha_creacion, version
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, (CURRENT_TIMESTAMP AT TIME ZONE 'UTC'), 1)
        """, [
            pedido_id, numero_pedido,
            request.cedi_origen_id, request.cedi_origen_nombre,
            request.tienda_destino_id, request.tienda_destino_nombre,
            request.fecha_pedido or date.today(), request.fecha_entrega_solicitada,
            ESTADO_BORRADOR, request.requiere_aprobacion,
            total_productos, float(total_bultos), float(total_unidades),
            tiene_devoluciones, total_productos_devolucion,
            float(total_bultos_devolucion), float(total_unidades_devolucion),
            request.dias_cobertura, request.tipo_pedido, request.prioridad,
            request.observaciones, request.notas_picking, request.notas_entrega,
            "sistema"
        ])

        # Insertar detalle de productos a RECIBIR
        for idx, producto in enumerate(productos_incluidos):
            detalle_id = str(uuid.uuid4())
            cursor.execute("""
                INSERT INTO pedidos_sugeridos_detalle (
                    id, pedido_id, linea_numero,
                    codigo_producto, codigo_barras, descripcion_producto,
                    categoria, grupo, subgrupo, marca, modelo, presentacion,
                    cuadrante_producto, cantidad_bultos,
                    cantidad_sugerida_unidades, cantidad_sugerida_bultos,
                    cantidad_pedida_unidades, cantidad_pedida_bultos,
                    total_unidades,
                    prom_ventas_5dias_unid, prom_ventas_8sem_unid,
                    stock_tienda, stock_cedi_origen,
                    stock_minimo, stock_maximo, punto_reorden,
                    razon_pedido, incluido, observaciones,
                    fecha_creacion
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            """, [
                detalle_id, pedido_id, idx + 1,
                producto.codigo_producto, producto.codigo_barras, producto.descripcion_producto,
                producto.categoria, producto.grupo, producto.subgrupo,
                producto.marca, producto.modelo, producto.presentacion,
                producto.cuadrante_producto, float(producto.cantidad_bultos),
                float(producto.cantidad_sugerida_unidades), float(producto.cantidad_sugerida_bultos),
                float(producto.cantidad_pedida_unidades), float(producto.cantidad_pedida_bultos),
                float(producto.total_unidades),
                float(producto.prom_ventas_5dias_unid), float(producto.prom_ventas_8sem_unid),
                float(producto.stock_tienda), float(producto.stock_cedi_origen),
                float(producto.stock_minimo), float(producto.stock_maximo), float(producto.punto_reorden),
                producto.razon_pedido, producto.incluido, producto.observaciones
            ])

        # Insertar devoluciones
        for idx, devolucion in enumerate(devoluciones_incluidas):
            devolucion_id = str(uuid.uuid4())
            cursor.execute("""
                INSERT INTO pedidos_sugeridos_devoluciones (
                    id, pedido_id, linea_numero,
                    codigo_producto, codigo_barras, descripcion_producto,
                    categoria, grupo, subgrupo, marca, presentacion,
                    cuadrante_producto, cantidad_bultos,
                    stock_actual_tienda, stock_maximo, stock_optimo,
                    exceso_unidades, exceso_bultos,
                    devolucion_sugerida_unidades, devolucion_sugerida_bultos,
                    devolucion_confirmada_unidades, devolucion_confirmada_bultos,
                    total_unidades_devolver,
                    razon_devolucion, prioridad_devolucion,
                    dias_sin_venta, prom_ventas_30dias, dias_cobertura_actual,
                    incluido, observaciones,
                    fecha_creacion
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            """, [
                devolucion_id, pedido_id, idx + 1,
                devolucion.codigo_producto, devolucion.codigo_barras, devolucion.descripcion_producto,
                devolucion.categoria, devolucion.grupo, devolucion.subgrupo,
                devolucion.marca, devolucion.presentacion,
                devolucion.cuadrante_producto, float(devolucion.cantidad_bultos),
                float(devolucion.stock_actual_tienda), float(devolucion.stock_maximo),
                float(devolucion.stock_optimo) if devolucion.stock_optimo else None,
                float(devolucion.exceso_unidades), float(devolucion.exceso_bultos),
                float(devolucion.devolucion_sugerida_unidades), float(devolucion.devolucion_sugerida_bultos),
                float(devolucion.devolucion_confirmada_unidades), float(devolucion.devolucion_confirmada_bultos),
                float(devolucion.total_unidades_devolver),
                devolucion.razon_devolucion, devolucion.prioridad_devolucion,
                devolucion.dias_sin_venta,
                float(devolucion.prom_ventas_30dias) if devolucion.prom_ventas_30dias else None,
                float(devolucion.dias_cobertura_actual) if devolucion.dias_cobertura_actual else None,
                devolucion.incluido, devolucion.observaciones
            ])

        # Registrar en historial
        historial_id = str(uuid.uuid4())
        cursor.execute("""
            INSERT INTO pedidos_sugeridos_historial (
                id, pedido_id, estado_anterior, estado_nuevo,
                motivo_cambio, usuario, fecha_cambio
            ) VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        """, [
            historial_id, pedido_id, None, ESTADO_BORRADOR,
            "Pedido creado", "sistema"
        ])

        conn.commit()
        cursor.close()

        logger.info(f"✅ Pedido {numero_pedido} guardado exitosamente")

        return PedidoGuardadoResponse(
            id=pedido_id,
            numero_pedido=numero_pedido,
            estado=ESTADO_BORRADOR,
            total_productos=total_productos,
            total_bultos=float(total_bultos),
            tiene_devoluciones=tiene_devoluciones,
            total_productos_devolucion=total_productos_devolucion,
            total_bultos_devolucion=float(total_bultos_devolucion),
            fecha_creacion=datetime.now(),
            mensaje=f"Pedido {numero_pedido} guardado exitosamente en estado borrador"
        )

    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error guardando pedido: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error guardando pedido: {str(e)}")


# =====================================================================================
# OBTENER PEDIDO COMPLETO
# =====================================================================================

@router.get("/{pedido_id}", response_model=PedidoSugeridoCompleto)
async def obtener_pedido(
    pedido_id: str,
    conn: Any = Depends(get_db)
):
    """
    Obtiene pedido completo con todos los productos y devoluciones
    """
    try:
        cursor = conn.cursor()

        # Obtener pedido principal
        cursor.execute("""
            SELECT
                id, numero_pedido, fecha_pedido, fecha_creacion,
                cedi_origen_id, cedi_origen_nombre,
                tienda_destino_id, tienda_destino_nombre,
                estado, prioridad, tipo_pedido,
                total_productos, total_bultos, total_unidades,
                tiene_devoluciones, total_productos_devolucion,
                total_bultos_devolucion, total_unidades_devolucion,
                dias_cobertura, observaciones,
                usuario_creador, fecha_modificacion
            FROM pedidos_sugeridos
            WHERE id = %s
        """, [pedido_id])
        pedido_row = cursor.fetchone()

        if not pedido_row:
            cursor.close()
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        # Obtener productos del pedido
        cursor.execute("""
            SELECT
                codigo_producto, descripcion_producto,
                cantidad_bultos, cantidad_pedida_bultos, total_unidades,
                cantidad_sugerida_bultos, clasificacion_abc, razon_pedido,
                prom_ventas_8sem_unid, prom_ventas_8sem_bultos,
                stock_tienda, stock_total, incluido
            FROM pedidos_sugeridos_detalle
            WHERE pedido_id = %s
            ORDER BY linea_numero
        """, [pedido_id])
        productos_rows = cursor.fetchall()
        cursor.close()

        # Construir lista de productos
        productos = []
        for idx, prod_row in enumerate(productos_rows):
            cantidad_bultos = safe_float(prod_row[2], default=1.0)
            cantidad_pedida_bultos = safe_float(prod_row[3])

            productos.append({
                "id": f"{pedido_id}-{idx+1}",
                "pedido_id": pedido_id,
                "linea_numero": idx + 1,
                "codigo_producto": prod_row[0],
                "descripcion_producto": prod_row[1],
                "cantidad_bultos": cantidad_bultos,
                "cantidad_sugerida_bultos": safe_float(prod_row[5]),
                "cantidad_sugerida_unidades": safe_float(prod_row[5]) * cantidad_bultos,
                "clasificacion_abc": prod_row[6],
                "razon_pedido": prod_row[7] or "",
                "prom_ventas_8sem_unid": safe_float(prod_row[8]),
                "prom_ventas_8sem_bultos": safe_float(prod_row[9]),
                "stock_tienda": safe_float(prod_row[10]),
                "stock_total": safe_float(prod_row[11]),
                "cantidad_pedida_bultos": cantidad_pedida_bultos,
                "cantidad_pedida_unidades": cantidad_pedida_bultos * cantidad_bultos,
                "total_unidades": safe_float(prod_row[4]),
                "incluido": prod_row[12] if prod_row[12] is not None else True,
                "fecha_creacion": pedido_row[3],
            })

        # Construir respuesta completa
        return {
            "id": pedido_row[0],
            "numero_pedido": pedido_row[1],
            "fecha_pedido": pedido_row[2],
            "fecha_creacion": to_venezuela_aware(pedido_row[3]),
            "cedi_origen_nombre": pedido_row[5] or "CEDI",
            "tienda_destino_nombre": pedido_row[7] or "Tienda",
            "estado": pedido_row[8],
            "prioridad": pedido_row[9] or "normal",
            "tipo_pedido": pedido_row[10] or "reposicion",
            "total_productos": pedido_row[11] or 0,
            "total_lineas": len(productos),
            "total_bultos": float(pedido_row[12]) if pedido_row[12] else 0.0,
            "total_unidades": float(pedido_row[13]) if pedido_row[13] else 0.0,
            "total_peso_kg": None,
            "fecha_entrega_solicitada": None,
            "fecha_aprobacion": None,
            "fecha_recepcion": None,
            "usuario_creador": pedido_row[20] or "sistema",
            "dias_desde_creacion": None,
            "porcentaje_avance": None,
            "cedi_origen_id": pedido_row[4],
            "tienda_destino_id": pedido_row[6],
            "numero_guia": None,
            "numero_orden_compra": None,
            "numero_picking": None,
            "sub_estado": None,
            "requiere_aprobacion": pedido_row[8] == ESTADO_PENDIENTE_APROBACION_GERENTE,
            "total_volumen_m3": None,
            "tiene_devoluciones": pedido_row[14] or False,
            "total_productos_devolucion": pedido_row[15] or 0,
            "total_bultos_devolucion": float(pedido_row[16]) if pedido_row[16] else 0.0,
            "total_unidades_devolucion": float(pedido_row[17]) if pedido_row[17] else 0.0,
            "dias_cobertura": pedido_row[18] or 3,
            "requiere_refrigeracion": False,
            "requiere_congelacion": False,
            "paleta_asignada": None,
            "observaciones": pedido_row[19],
            "notas_picking": None,
            "notas_entrega": None,
            "notas_recepcion": None,
            "usuario_aprobador": None,
            "usuario_picker": None,
            "usuario_receptor": None,
            "fecha_modificacion": to_venezuela_aware(pedido_row[21]),
            "fecha_inicio_picking": None,
            "fecha_fin_picking": None,
            "fecha_despacho": None,
            "fecha_cancelacion": None,
            "fecha_entrega_real": None,
            "version": 1,
            "pedido_padre_id": None,
            "porcentaje_cumplimiento": None,
            "tiempo_preparacion_horas": None,
            "productos": productos,
            "devoluciones": []
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error obteniendo pedido: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error obteniendo pedido: {str(e)}")


# =====================================================================================
# ELIMINAR PEDIDO
# =====================================================================================

@router.delete("/{pedido_id}")
async def eliminar_pedido(
    pedido_id: str,
    conn: Any = Depends(get_db_write)
):
    """
    Elimina un pedido y todos sus detalles asociados.
    Solo se pueden eliminar pedidos en estado BORRADOR.
    """
    try:
        cursor = conn.cursor()

        # Verificar que el pedido existe y está en estado borrador
        cursor.execute("""
            SELECT id, numero_pedido, estado, tienda_destino_nombre
            FROM pedidos_sugeridos
            WHERE id = %s
        """, [pedido_id])
        pedido = cursor.fetchone()

        if not pedido:
            cursor.close()
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        numero_pedido = pedido[1]
        estado = pedido[2]
        tienda = pedido[3]

        # Solo permitir eliminar pedidos en borrador
        if estado != ESTADO_BORRADOR:
            cursor.close()
            raise HTTPException(
                status_code=400,
                detail=f"Solo se pueden eliminar pedidos en estado borrador. Estado actual: {estado}"
            )

        # Eliminar en cascada (las FK tienen ON DELETE CASCADE)
        # Pero por seguridad eliminamos explícitamente
        cursor.execute("DELETE FROM pedidos_sugeridos_comentarios WHERE pedido_id = %s", [pedido_id])
        cursor.execute("DELETE FROM pedidos_sugeridos_historial WHERE pedido_id = %s", [pedido_id])
        cursor.execute("DELETE FROM pedidos_sugeridos_devoluciones WHERE pedido_id = %s", [pedido_id])
        cursor.execute("DELETE FROM pedidos_sugeridos_detalle WHERE pedido_id = %s", [pedido_id])
        cursor.execute("DELETE FROM pedidos_sugeridos WHERE id = %s", [pedido_id])

        conn.commit()
        cursor.close()

        logger.info(f"🗑️ Pedido {numero_pedido} ({tienda}) eliminado exitosamente")

        return {
            "success": True,
            "mensaje": f"Pedido {numero_pedido} eliminado exitosamente",
            "pedido_id": pedido_id,
            "numero_pedido": numero_pedido
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error eliminando pedido: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error eliminando pedido: {str(e)}")


# =====================================================================================
# ENVIAR PARA APROBACIÓN DE GERENTE
# =====================================================================================

@router.post("/{pedido_id}/enviar-aprobacion")
async def enviar_para_aprobacion(
    pedido_id: str,
    conn: Any = Depends(get_db_write)
):
    """
    Cambia estado de BORRADOR → PENDIENTE_APROBACION_GERENTE
    """
    try:
        cursor = conn.cursor()

        # Verificar pedido existe y está en estado correcto
        cursor.execute("""
            SELECT id, numero_pedido, estado, tienda_destino_nombre
            FROM pedidos_sugeridos
            WHERE id = %s
        """, [pedido_id])
        pedido = cursor.fetchone()

        if not pedido:
            cursor.close()
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        if pedido[2] != ESTADO_BORRADOR:
            cursor.close()
            raise HTTPException(
                status_code=400,
                detail=f"Pedido debe estar en estado BORRADOR. Estado actual: {pedido[2]}"
            )

        # Actualizar estado
        cursor.execute("""
            UPDATE pedidos_sugeridos
            SET estado = %s,
                fecha_modificacion = CURRENT_TIMESTAMP
            WHERE id = %s
        """, [ESTADO_PENDIENTE_APROBACION_GERENTE, pedido_id])

        # Registrar en historial
        historial_id = str(uuid.uuid4())
        cursor.execute("""
            INSERT INTO pedidos_sugeridos_historial (
                id, pedido_id, estado_anterior, estado_nuevo,
                motivo_cambio, usuario, fecha_cambio
            ) VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        """, [
            historial_id, pedido_id,
            ESTADO_BORRADOR,
            ESTADO_PENDIENTE_APROBACION_GERENTE,
            "Enviado para aprobación del gerente",
            "sistema"
        ])

        conn.commit()
        cursor.close()

        logger.info(f"✅ Pedido {pedido[1]} enviado para aprobación")

        return {
            "success": True,
            "message": f"Pedido {pedido[1]} enviado para aprobación del gerente de {pedido[3]}",
            "pedido_id": pedido_id,
            "numero_pedido": pedido[1],
            "estado_nuevo": ESTADO_PENDIENTE_APROBACION_GERENTE
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error enviando pedido para aprobación: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error enviando pedido: {str(e)}")


# =====================================================================================
# APROBAR PEDIDO (GERENTE)
# =====================================================================================

@router.post("/{pedido_id}/aprobar")
async def aprobar_pedido(
    pedido_id: str,
    comentario_general: Optional[str] = None,
    conn: Any = Depends(get_db_write)
):
    """
    Gerente aprueba el pedido
    """
    try:
        cursor = conn.cursor()

        # Verificar pedido existe y está en estado correcto
        cursor.execute("""
            SELECT id, numero_pedido, estado, tienda_destino_nombre
            FROM pedidos_sugeridos
            WHERE id = %s
        """, [pedido_id])
        pedido = cursor.fetchone()

        if not pedido:
            cursor.close()
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        if pedido[2] != ESTADO_PENDIENTE_APROBACION_GERENTE:
            cursor.close()
            raise HTTPException(
                status_code=400,
                detail=f"Pedido debe estar en estado PENDIENTE_APROBACION_GERENTE. Estado actual: {pedido[2]}"
            )

        # Actualizar estado
        cursor.execute("""
            UPDATE pedidos_sugeridos
            SET estado = %s,
                usuario_aprobador_gerente = %s,
                fecha_aprobacion_gerente = CURRENT_TIMESTAMP,
                fecha_modificacion = CURRENT_TIMESTAMP
            WHERE id = %s
        """, [ESTADO_APROBADO_GERENTE, "sistema", pedido_id])

        # Registrar en historial
        historial_id = str(uuid.uuid4())
        cursor.execute("""
            INSERT INTO pedidos_sugeridos_historial (
                id, pedido_id, estado_anterior, estado_nuevo,
                motivo_cambio, usuario, fecha_cambio
            ) VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        """, [
            historial_id, pedido_id,
            ESTADO_PENDIENTE_APROBACION_GERENTE,
            ESTADO_APROBADO_GERENTE,
            comentario_general or "Aprobado por gerente",
            "sistema"
        ])

        conn.commit()
        cursor.close()

        logger.info(f"✅ Pedido {pedido[1]} aprobado por gerente")

        return {
            "success": True,
            "message": f"Pedido {pedido[1]} aprobado exitosamente",
            "pedido_id": pedido_id,
            "numero_pedido": pedido[1],
            "estado_nuevo": ESTADO_APROBADO_GERENTE
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error aprobando pedido: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error aprobando pedido: {str(e)}")


# =====================================================================================
# RECHAZAR PEDIDO (GERENTE)
# =====================================================================================

@router.post("/{pedido_id}/rechazar")
async def rechazar_pedido(
    pedido_id: str,
    motivo: str,
    conn: Any = Depends(get_db_write)
):
    """
    Gerente rechaza el pedido
    """
    try:
        cursor = conn.cursor()

        # Verificar pedido existe y está en estado correcto
        cursor.execute("""
            SELECT id, numero_pedido, estado, tienda_destino_nombre
            FROM pedidos_sugeridos
            WHERE id = %s
        """, [pedido_id])
        pedido = cursor.fetchone()

        if not pedido:
            cursor.close()
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        if pedido[2] != ESTADO_PENDIENTE_APROBACION_GERENTE:
            cursor.close()
            raise HTTPException(
                status_code=400,
                detail=f"Pedido debe estar en estado PENDIENTE_APROBACION_GERENTE. Estado actual: {pedido[2]}"
            )

        # Actualizar estado
        cursor.execute("""
            UPDATE pedidos_sugeridos
            SET estado = %s,
                usuario_aprobador_gerente = %s,
                fecha_modificacion = CURRENT_TIMESTAMP
            WHERE id = %s
        """, [ESTADO_RECHAZADO_GERENTE, "sistema", pedido_id])

        # Registrar en historial
        historial_id = str(uuid.uuid4())
        cursor.execute("""
            INSERT INTO pedidos_sugeridos_historial (
                id, pedido_id, estado_anterior, estado_nuevo,
                motivo_cambio, usuario, fecha_cambio
            ) VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        """, [
            historial_id, pedido_id,
            ESTADO_PENDIENTE_APROBACION_GERENTE,
            ESTADO_RECHAZADO_GERENTE,
            motivo,
            "sistema"
        ])

        conn.commit()
        cursor.close()

        logger.info(f"❌ Pedido {pedido[1]} rechazado por gerente: {motivo}")

        return {
            "success": True,
            "message": f"Pedido {pedido[1]} rechazado",
            "pedido_id": pedido_id,
            "numero_pedido": pedido[1],
            "estado_nuevo": ESTADO_RECHAZADO_GERENTE,
            "motivo": motivo
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error rechazando pedido: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error rechazando pedido: {str(e)}")


# =====================================================================================
# DEBUG ENDPOINT - Ver schema de tabla
# =====================================================================================

@router.get("/debug/schema")
async def get_table_schema(conn: Any = Depends(get_db)):
    """
    Endpoint temporal para inspeccionar el esquema de la tabla pedidos_sugeridos
    """
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_name = 'pedidos_sugeridos'
            ORDER BY ordinal_position
        """)
        columns = cursor.fetchall()
        cursor.close()

        return {
            "table": "pedidos_sugeridos",
            "columns": [{"name": col[0], "type": col[1]} for col in columns]
        }
    except Exception as e:
        return {"error": str(e)}


# =====================================================================================
# CREAR PEDIDO V2.0 (Método Nivel Objetivo ABC-XYZ)
# =====================================================================================

from pydantic import BaseModel

class ProductoPedidoV2(BaseModel):
    """Producto para pedido v2.0 con datos de nivel objetivo"""
    producto_id: str
    cantidad_pedida: float
    cantidad_sugerida: float
    nivel_objetivo: float
    stock_actual: float
    inventario_en_transito: float
    stock_seguridad: float
    demanda_ciclo: float
    matriz_abc_xyz: str
    prioridad: int


class CrearPedidoV2Request(BaseModel):
    """Request para crear pedido v2.0"""
    cedi_origen_id: str
    tienda_destino_id: str
    fecha_pedido: date
    tipo_pedido: str = "sugerido_v2"
    metodo_calculo: str = "NIVEL_OBJETIVO_V2"
    productos: List[ProductoPedidoV2]


# =====================================================================================
# CALCULAR PRODUCTOS SUGERIDOS (PostgreSQL)
# =====================================================================================

class CalcularProductosRequest(BaseModel):
    """Request para calcular productos sugeridos"""
    cedi_origen: str
    tienda_destino: str
    dias_cobertura: int = 3


class ProductoCalculado(BaseModel):
    """Producto calculado para pedido sugerido"""
    codigo_producto: str
    codigo_barras: Optional[str] = None
    descripcion_producto: str
    categoria: Optional[str] = None
    grupo: Optional[str] = None
    subgrupo: Optional[str] = None
    marca: Optional[str] = None
    presentacion: Optional[str] = None
    cantidad_bultos: float = 1.0
    unidad_pedido: str = "Bulto"  # Unidad de pedido: Bulto, Blister, Cesta, KG, UND, etc.
    peso_unidad: float = 1000.0
    cuadrante_producto: Optional[str] = None
    # Ventas
    prom_ventas_5dias_unid: float = 0.0
    prom_ventas_20dias_unid: float = 0.0
    prom_top3_unid: float = 0.0  # Promedio TOP 3 días
    prom_p75_unid: float = 0.0   # Percentil 75
    prom_mismo_dia_unid: float = 0.0
    prom_ventas_8sem_unid: float = 0.0
    prom_ventas_8sem_bultos: float = 0.0
    prom_ventas_3dias_unid: float = 0.0
    prom_ventas_3dias_bultos: float = 0.0
    prom_mismo_dia_bultos: float = 0.0
    # Inventario
    stock_tienda: float = 0.0
    stock_en_transito: float = 0.0
    stock_total: float = 0.0
    stock_total_bultos: float = 0.0
    stock_dias_cobertura: float = 0.0
    stock_cedi_seco: float = 0.0
    stock_cedi_frio: float = 0.0
    stock_cedi_verde: float = 0.0
    stock_cedi_origen: float = 0.0
    # Configuración
    clasificacion_abc: Optional[str] = None
    clase_efectiva: Optional[str] = None  # Clase usada para calculo (puede diferir por generador trafico)
    es_generador_trafico: bool = False    # Si es generador de trafico
    stock_minimo: float = 0.0
    stock_maximo: float = 0.0
    stock_seguridad: float = 0.0
    punto_reorden: float = 0.0
    cantidad_sugerida_unid: float = 0.0
    cantidad_sugerida_bultos: float = 0.0
    cantidad_ajustada_bultos: float = 0.0
    razon_pedido: str = ""
    metodo_calculo: str = ""              # estadistico o padre_prudente
    # Sobrestock
    tiene_sobrestock: bool = False
    exceso_unidades: float = 0.0
    exceso_bultos: int = 0
    dias_exceso: float = 0.0
    # Límites de inventario (capacidad máxima y mínimo exhibición)
    capacidad_maxima_configurada: Optional[float] = None  # Límite superior configurado en unidades
    cantidad_antes_ajuste_capacidad: Optional[float] = None  # Cantidad original antes de ajustar por capacidad
    ajustado_por_capacidad: bool = False  # True si se ajustó por límite de capacidad máxima
    tipo_restriccion_capacidad: Optional[str] = None  # congelador, refrigerador, etc.
    notas_capacidad: Optional[str] = None  # Notas de la configuración de capacidad
    # Mínimo de exhibición
    minimo_exhibicion_configurado: Optional[float] = None  # Mínimo para que producto se vea bien
    ajustado_por_minimo_exhibicion: bool = False  # True si se elevó por mínimo de exhibición
    # Warnings de sanity checks
    warnings_calculo: List[str] = []

    # === CAMPOS V2 (Cobertura Real por Día de Semana) ===
    # Promedios por día de semana (0=Dom, 1=Lun, ..., 6=Sáb)
    v2_prom_dow: List[float] = []  # Array de 7 elementos con promedio de cada día
    # Demanda real del período (suma de demanda de los días específicos a cubrir)
    v2_demanda_periodo: float = 0.0  # Suma de demanda real del período de cobertura
    # Cantidad sugerida V2 (basada en demanda real del período)
    v2_cantidad_sugerida_unid: float = 0.0
    v2_cantidad_sugerida_bultos: float = 0.0
    v2_diferencia_bultos: float = 0.0  # V2 - V1 (positivo = V2 pide más)
    # Simulación de cobertura día a día (con la cantidad V2)
    v2_cobertura_dias: List[dict] = []  # [{dia: "Jue", fecha: "20 Dic", demanda: 3.5, stock_final: 19.5, cobertura_pct: 100}, ...]
    v2_dias_cobertura_real: float = 0.0  # Días reales que cubre (puede ser 3.7 días)
    v2_primer_dia_riesgo: Optional[str] = None  # Primer día donde cobertura < 100%
    # Info del período
    v2_dia_pedido: Optional[str] = None  # Día en que se hace el pedido (ej: "Mié")
    v2_dia_llegada: Optional[str] = None  # Día en que llega el pedido (ej: "Jue")
    v2_fecha_pedido: Optional[str] = None  # Fecha del pedido (ej: "18 Dic")
    v2_fecha_llegada: Optional[str] = None  # Fecha de llegada (ej: "19 Dic")
    v2_dias_cobertura_config: int = 0  # Días de cobertura según config ABC de tienda
    v2_lead_time_config: float = 1.5  # Lead time según config de tienda
    # Histórico detallado por día de semana para el modal
    v2_historico_dow: List[dict] = []  # [{dow: 0, nombre: "Dom", promedio: 11.9, dias_con_data: 4, fechas: ["1 Dic", "8 Dic", ...], ventas: [12, 10, ...]}]


@router.post("/calcular", response_model=List[ProductoCalculado])
async def calcular_productos_sugeridos(
    request: CalcularProductosRequest,
    cuadrantes: Optional[List[str]] = Query(None, description="Filtrar por cuadrantes"),
    conn: Any = Depends(get_db)
):
    """
    Calcula productos sugeridos para un pedido basado en:
    - Maestro de productos (PostgreSQL)
    - Ventas históricas de la tienda destino
    - Inventario actual en tienda y CEDI origen
    - Configuración ABC por tienda (si existe)

    Retorna lista de productos con cantidades sugeridas.
    """
    try:
        logger.info(f"📦 Calculando productos sugeridos: {request.cedi_origen} → {request.tienda_destino}")
        cursor = conn.cursor()

        # 1. Cargar configuración ABC de la tienda (si existe)
        try:
            cursor.execute("""
                SELECT lead_time_override, dias_cobertura_a, dias_cobertura_b,
                       dias_cobertura_c, clase_d_dias_cobertura
                FROM config_parametros_abc_tienda
                WHERE tienda_id = %s AND activo = true
            """, [request.tienda_destino])
            config_row = cursor.fetchone()

            if config_row:
                config_tienda = ConfigTiendaABC(
                    lead_time=float(config_row[0]) if config_row[0] else LEAD_TIME_DEFAULT,
                    dias_cobertura_a=int(config_row[1]) if config_row[1] else 7,
                    dias_cobertura_b=int(config_row[2]) if config_row[2] else 14,
                    dias_cobertura_c=int(config_row[3]) if config_row[3] else 21,
                    dias_cobertura_d=int(config_row[4]) if config_row[4] else 30
                )
                set_config_tienda(config_tienda)
                logger.info(f"📋 Config tienda aplicada: LT={config_tienda.lead_time}, A={config_tienda.dias_cobertura_a}d, B={config_tienda.dias_cobertura_b}d, C={config_tienda.dias_cobertura_c}d, D={config_tienda.dias_cobertura_d}d")
            else:
                config_tienda = ConfigTiendaABC()  # Usar defaults
                set_config_tienda(config_tienda)
                logger.info(f"📋 Usando configuración ABC por defecto para {request.tienda_destino}")
        except Exception as e:
            logger.warning(f"No se pudo cargar config tienda: {e}. Usando defaults.")
            config_tienda = ConfigTiendaABC()  # Usar defaults
            set_config_tienda(config_tienda)

        # 1.5. Cargar umbrales ABC para log informativo
        # NOTA: La clasificación ABC ahora viene de productos_abc_tienda (tabla cache)
        # que se recalcula diariamente a las 4:00 AM por recalcular_abc_cache.py
        umbral_a, umbral_b, umbral_c = 50, 200, 800  # Defaults
        try:
            cursor.execute("""
                SELECT id, valor_numerico
                FROM config_inventario_global
                WHERE categoria = 'abc_umbrales_ranking' AND activo = true
            """)
            for row in cursor.fetchall():
                if row[0] == 'abc_umbral_a' and row[1]:
                    umbral_a = int(row[1])
                elif row[0] == 'abc_umbral_b' and row[1]:
                    umbral_b = int(row[1])
                elif row[0] == 'abc_umbral_c' and row[1]:
                    umbral_c = int(row[1])
            logger.info(f"📋 Umbrales ABC (desde cache): A≤{umbral_a}, B≤{umbral_b}, C≤{umbral_c}, D>{umbral_c}")
        except Exception as e:
            logger.warning(f"No se pudieron cargar umbrales ABC: {e}. Usando defaults.")

        # 1.6. Cargar configuración de cobertura por categoría (desde BD, no hardcodeado)
        config_cobertura_categoria = {}
        try:
            cursor.execute("""
                SELECT categoria_normalizada, dias_cobertura_a, dias_cobertura_b,
                       dias_cobertura_c, dias_cobertura_d
                FROM config_cobertura_categoria
                WHERE activo = true
            """)
            for row in cursor.fetchall():
                cat_norm = (row[0] or '').strip().upper()
                config_cobertura_categoria[cat_norm] = {
                    'A': row[1] if row[1] is not None else 7,
                    'B': row[2] if row[2] is not None else 14,
                    'C': row[3] if row[3] is not None else 21,
                    'D': row[4] if row[4] is not None else 30
                }
            if config_cobertura_categoria:
                logger.info(f"🥬 Coberturas por categoría cargadas: {config_cobertura_categoria}")
        except Exception as e:
            logger.warning(f"No se pudo cargar config cobertura por categoría: {e}")

        # 1.7. Cargar límites de inventario (capacidad máxima y mínimo exhibición) para la tienda destino
        limites_inventario = {}  # {producto_codigo: {capacidad_maxima, minimo_exhibicion, tipo, notas}}
        try:
            cursor.execute("""
                SELECT producto_codigo, capacidad_maxima_unidades, minimo_exhibicion_unidades,
                       tipo_restriccion, notas
                FROM capacidad_almacenamiento_producto
                WHERE tienda_id = %s AND activo = true
            """, [request.tienda_destino])
            for row in cursor.fetchall():
                limites_inventario[row[0]] = {
                    'capacidad_maxima': safe_float(row[1]) if row[1] else None,
                    'minimo_exhibicion': safe_float(row[2]) if row[2] else None,
                    'tipo_restriccion': row[3] or 'espacio_fisico',
                    'notas': row[4]
                }
            if limites_inventario:
                n_max = sum(1 for v in limites_inventario.values() if v['capacidad_maxima'])
                n_min = sum(1 for v in limites_inventario.values() if v['minimo_exhibicion'])
                logger.info(f"📦 Límites de inventario cargados: {n_max} con capacidad máx, {n_min} con mínimo exhibición")
        except Exception as e:
            logger.warning(f"No se pudo cargar límites de inventario: {e}")

        # 1.8. Cargar productos excluidos para la tienda destino
        codigos_excluidos = set()
        try:
            # Usar savepoint para no romper la transacción si la tabla no existe
            cursor.execute("SAVEPOINT check_excluidos")
            cursor.execute("""
                SELECT codigo_producto
                FROM productos_excluidos_tienda
                WHERE tienda_id = %s AND activo = TRUE
            """, [request.tienda_destino])
            codigos_excluidos = {row[0] for row in cursor.fetchall()}
            cursor.execute("RELEASE SAVEPOINT check_excluidos")
            if codigos_excluidos:
                logger.info(f"🚫 Productos excluidos cargados: {len(codigos_excluidos)} productos no aparecerán en sugerencias")
        except Exception as e:
            # Rollback al savepoint para recuperar la transacción
            try:
                cursor.execute("ROLLBACK TO SAVEPOINT check_excluidos")
            except:
                pass
            logger.warning(f"No se pudo cargar productos excluidos: {e}")

        # 2. Obtener la región de la tienda destino y tiendas de referencia
        cursor.execute("""
            SELECT region FROM ubicaciones WHERE id = %s
        """, [request.tienda_destino])
        region_row = cursor.fetchone()
        tienda_region = region_row[0] if region_row and region_row[0] else 'VALENCIA'

        # Obtener tiendas de la misma región (excluyendo la tienda destino y CEDIs)
        cursor.execute("""
            SELECT id, nombre FROM ubicaciones
            WHERE region = %s
              AND id != %s
              AND tipo = 'tienda'
              AND activo = true
            ORDER BY nombre
        """, [tienda_region, request.tienda_destino])
        tiendas_referencia = cursor.fetchall()
        tiendas_ref_ids = [t[0] for t in tiendas_referencia]
        tiendas_ref_nombres = {t[0]: t[1] for t in tiendas_referencia}

        logger.info(f"📍 Región: {tienda_region}, Tiendas referencia: {[t[1] for t in tiendas_referencia]}")

        # 3. Query principal para calcular productos sugeridos desde PostgreSQL
        # NOTA: Esta consulta está diseñada para funcionar con tiendas nuevas que tienen
        # pocos días de historial. Si hay menos de 20 días, usa los datos disponibles.
        # NUEVO: Incluye P75 de tiendas de referencia para productos sin ventas locales.
        query = """
            WITH ventas_diarias_disponibles AS (
                -- Ventas últimos 30 días (optimizado para tiendas con millones de registros como Bosque)
                -- IMPORTANTE: Excluir día actual (incompleto) para no sesgar promedios
                -- NOTA: Para tienda_18 (PARAISO) se excluye 2025-12-06 (inauguración con ventas atípicas)
                SELECT
                    producto_id,
                    fecha_venta::date as fecha,
                    SUM(cantidad_vendida) as total_dia
                FROM ventas
                WHERE ubicacion_id = %s
                  AND fecha_venta::date >= CURRENT_DATE - INTERVAL '30 days'  -- Solo últimos 30 días (performance: reduce 8.7M → 260K filas para Bosque)
                  AND fecha_venta::date < CURRENT_DATE  -- Excluir hoy (día incompleto)
                  AND NOT (ubicacion_id = 'tienda_18' AND fecha_venta::date = '2025-12-06')  -- Excluir inauguración Paraíso
                GROUP BY producto_id, fecha_venta::date
            ),
            ventas_20dias AS (
                -- Estadisticas de los ultimos 20 dias (o los dias disponibles si hay menos)
                SELECT
                    producto_id,
                    COUNT(DISTINCT fecha) as dias_con_venta,
                    SUM(total_dia) as total_vendido,
                    AVG(total_dia) as prom_diario
                FROM ventas_diarias_disponibles
                WHERE fecha >= CURRENT_DATE - INTERVAL '20 days'
                GROUP BY producto_id
            ),
            ventas_5dias AS (
                SELECT
                    producto_id,
                    AVG(total_dia) as prom_diario_5d
                FROM ventas_diarias_disponibles
                WHERE fecha >= CURRENT_DATE - INTERVAL '5 days'
                GROUP BY producto_id
            ),
            -- TOP3: promedio de los 3 días con más ventas (o menos si no hay 3 dias)
            ranked_days AS (
                SELECT
                    producto_id,
                    total_dia,
                    ROW_NUMBER() OVER (PARTITION BY producto_id ORDER BY total_dia DESC) as rn
                FROM ventas_diarias_disponibles
                WHERE fecha >= CURRENT_DATE - INTERVAL '20 days'
            ),
            top3_ventas AS (
                SELECT
                    producto_id,
                    AVG(total_dia) as prom_top3
                FROM ranked_days
                WHERE rn <= 3
                GROUP BY producto_id
            ),
            -- P75: Percentil 75 (si solo hay 1 dia, usa ese valor como P75)
            percentil_75 AS (
                SELECT
                    producto_id,
                    CASE
                        WHEN COUNT(*) = 1 THEN MAX(total_dia)  -- Solo 1 dia: usar ese valor
                        ELSE PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY total_dia)
                    END as p75
                FROM ventas_diarias_disponibles
                WHERE fecha >= CURRENT_DATE - INTERVAL '20 days'
                GROUP BY producto_id
            ),
            -- Estadisticas para calculo ABC (sigma_demanda y demanda_maxima)
            estadisticas_30d AS (
                SELECT
                    producto_id,
                    CASE
                        WHEN COUNT(*) <= 1 THEN 0  -- Con 1 dia no hay desviacion
                        ELSE COALESCE(STDDEV(total_dia), 0)
                    END as sigma_demanda,
                    COALESCE(MAX(total_dia), 0) as demanda_maxima
                FROM ventas_diarias_disponibles
                WHERE fecha >= CURRENT_DATE - INTERVAL '30 days'
                GROUP BY producto_id
            ),
            -- ABC desde tabla cache (productos_abc_tienda)
            -- Pre-calculado diariamente a las 4:00 AM por recalcular_abc_cache.py
            -- Usa ranking por CANTIDAD vendida en los últimos 30 días POR TIENDA
            abc_tienda_cache AS (
                SELECT
                    producto_id,
                    clase_abc as clase_abc_valor,
                    rank_cantidad,
                    cantidad_30d as cantidad_total,
                    venta_30d as venta_total
                FROM productos_abc_tienda
                WHERE ubicacion_id = %s
            ),
            inv_tienda AS (
                SELECT
                    producto_id,
                    SUM(cantidad) as stock_tienda
                FROM inventario_actual
                WHERE ubicacion_id = %s
                GROUP BY producto_id
            ),
            inv_cedi AS (
                SELECT
                    producto_id,
                    SUM(cantidad) as stock_cedi
                FROM inventario_actual
                WHERE ubicacion_id = %s
                GROUP BY producto_id
            ),
            -- P75 de tiendas de referencia (misma región) para productos SIN ventas locales
            -- Esto permite sugerir "envíos de prueba" basados en demanda de tiendas similares
            ventas_referencia AS (
                SELECT
                    producto_id,
                    fecha_venta::date as fecha,
                    ubicacion_id,
                    SUM(cantidad_vendida) as total_dia
                FROM ventas
                WHERE ubicacion_id = ANY(%s)  -- Tiendas de referencia (misma región)
                  AND fecha_venta::date < CURRENT_DATE
                  AND fecha_venta >= CURRENT_DATE - INTERVAL '30 days'
                GROUP BY producto_id, fecha_venta::date, ubicacion_id
            ),
            p75_referencia AS (
                -- P75 promedio de las tiendas de referencia
                SELECT
                    producto_id,
                    AVG(p75_tienda) as p75_ref,
                    STRING_AGG(DISTINCT ubicacion_id, ',' ORDER BY ubicacion_id) as tiendas_con_venta
                FROM (
                    SELECT
                        producto_id,
                        ubicacion_id,
                        PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY total_dia) as p75_tienda
                    FROM ventas_referencia
                    GROUP BY producto_id, ubicacion_id
                ) sub
                GROUP BY producto_id
            )
            SELECT
                p.codigo as codigo_producto,
                p.codigo_barras,
                COALESCE(p.nombre, p.descripcion) as descripcion_producto,
                p.categoria,
                p.grupo,
                p.subgrupo,
                p.marca,
                p.presentacion,
                p.cuadrante,
                COALESCE(p.unidades_por_bulto, 1) as cantidad_bultos,
                COALESCE(p.unidad_pedido, 'Bulto') as unidad_pedido,
                COALESCE(p.peso_unitario * 1000, 1000.0) as peso_unidad,  -- Convertir kg a gramos
                -- Ventas
                COALESCE(v5.prom_diario_5d, 0) as prom_ventas_5dias_unid,
                COALESCE(v20.prom_diario, 0) as prom_ventas_20dias_unid,
                COALESCE(v20.dias_con_venta, 0) as dias_con_venta,
                COALESCE(v20.total_vendido, 0) as total_vendido_20d,
                -- Inventario
                COALESCE(it.stock_tienda, 0) as stock_tienda,
                COALESCE(ic.stock_cedi, 0) as stock_cedi_origen,
                -- TOP3 y P75
                COALESCE(t3.prom_top3, 0) as prom_top3,
                COALESCE(p75.p75, 0) as prom_p75,
                -- ABC por valor (Pareto)
                abc.clase_abc_valor,
                -- Estadisticas 30 dias para calculo inventario ABC
                COALESCE(est30.sigma_demanda, 0) as sigma_demanda,
                COALESCE(est30.demanda_maxima, 0) as demanda_maxima,
                -- Generador de trafico (GAP > 400)
                COALESCE(p.es_generador_trafico, false) as es_generador_trafico,
                -- NUEVO: P75 de referencia (tiendas misma región)
                COALESCE(p75ref.p75_ref, 0) as p75_referencia,
                p75ref.tiendas_con_venta as tiendas_referencia
            FROM productos p
            LEFT JOIN ventas_20dias v20 ON p.codigo = v20.producto_id
            LEFT JOIN ventas_5dias v5 ON p.codigo = v5.producto_id
            LEFT JOIN top3_ventas t3 ON p.codigo = t3.producto_id
            LEFT JOIN percentil_75 p75 ON p.codigo = p75.producto_id
            LEFT JOIN abc_tienda_cache abc ON p.codigo = abc.producto_id
            LEFT JOIN estadisticas_30d est30 ON p.codigo = est30.producto_id
            LEFT JOIN inv_tienda it ON p.codigo = it.producto_id
            LEFT JOIN inv_cedi ic ON p.codigo = ic.producto_id
            LEFT JOIN p75_referencia p75ref ON p.codigo = p75ref.producto_id
            WHERE p.activo = true
                AND (v20.total_vendido > 0 OR it.stock_tienda > 0 OR ic.stock_cedi > 0)
        """

        # Construir parámetros base
        params = [
            request.tienda_destino,  # ventas_diarias_disponibles
            request.tienda_destino,  # abc_tienda_cache (tabla cache ABC por tienda)
            request.tienda_destino,  # inv_tienda
            request.cedi_origen,     # inv_cedi
            tiendas_ref_ids if tiendas_ref_ids else ['__NONE__']  # tiendas referencia (evitar error si vacío)
        ]

        # Agregar filtro de cuadrantes si se proporciona
        if cuadrantes and len(cuadrantes) > 0:
            query += " AND p.cuadrante = ANY(%s)"
            params.append(cuadrantes)

        # Agregar ORDER BY al final
        query += " ORDER BY COALESCE(v20.total_vendido, 0) DESC"

        # Ejecutar query con parámetros
        cursor.execute(query, params)

        rows = cursor.fetchall()

        logger.info(f"📊 Encontrados {len(rows)} productos con datos")

        # ====================================================================
        # QUERY V3: Obtener promedios por día de semana para cada producto
        # ====================================================================
        # Esto permite calcular cobertura real día a día
        # DOW: 0=Domingo, 1=Lunes, ..., 6=Sábado
        cursor_dow = conn.cursor()
        cursor_dow.execute("""
            WITH ventas_por_dow AS (
                SELECT
                    producto_id,
                    EXTRACT(DOW FROM fecha_venta::date) as dow,
                    fecha_venta::date as fecha,
                    SUM(cantidad_vendida) as total_dia
                FROM ventas
                WHERE ubicacion_id = %s
                  AND fecha_venta::date >= CURRENT_DATE - INTERVAL '30 days'
                  AND fecha_venta::date < CURRENT_DATE
                GROUP BY producto_id, fecha_venta::date
            )
            SELECT
                producto_id,
                dow,
                AVG(total_dia) as prom_dow
            FROM ventas_por_dow
            GROUP BY producto_id, dow
            ORDER BY producto_id, dow
        """, [request.tienda_destino])

        # Construir diccionario de promedios por DOW para cada producto
        # Formato: {producto_id: [prom_dom, prom_lun, prom_mar, prom_mie, prom_jue, prom_vie, prom_sab]}
        promedios_dow_dict: dict = {}
        for row_dow in cursor_dow.fetchall():
            prod_id = row_dow[0]
            dow = int(row_dow[1])
            prom = float(row_dow[2]) if row_dow[2] else 0.0
            if prod_id not in promedios_dow_dict:
                promedios_dow_dict[prod_id] = [0.0] * 7  # 7 días
            promedios_dow_dict[prod_id][dow] = prom
        cursor_dow.close()

        # ====================================================================
        # QUERY V3b: Obtener detalle histórico por DOW (para modal explicativo)
        # ====================================================================
        cursor_dow_detail = conn.cursor()
        cursor_dow_detail.execute("""
            SELECT
                producto_id,
                EXTRACT(DOW FROM fecha_venta::date) as dow,
                TO_CHAR(fecha_venta::date, 'DD Mon') as fecha_str,
                fecha_venta::date as fecha,
                SUM(cantidad_vendida) as total_dia
            FROM ventas
            WHERE ubicacion_id = %s
              AND fecha_venta::date >= CURRENT_DATE - INTERVAL '30 days'
              AND fecha_venta::date < CURRENT_DATE
            GROUP BY producto_id, fecha_venta::date
            ORDER BY producto_id, fecha_venta::date
        """, [request.tienda_destino])

        # Construir diccionario de detalle histórico por DOW
        # Formato: {producto_id: {dow: [{fecha: "18 Dic", venta: 25}, ...]}}
        historico_dow_dict: dict = {}
        for row_detail in cursor_dow_detail.fetchall():
            prod_id = row_detail[0]
            dow = int(row_detail[1])
            fecha_str = row_detail[2]
            venta = float(row_detail[4]) if row_detail[4] else 0.0
            if prod_id not in historico_dow_dict:
                historico_dow_dict[prod_id] = {i: [] for i in range(7)}
            historico_dow_dict[prod_id][dow].append({
                "fecha": fecha_str,
                "venta": venta
            })
        cursor_dow_detail.close()

        # Obtener el día actual para calcular cobertura real
        from datetime import datetime, timedelta
        fecha_actual = datetime.now().date()
        dia_actual = datetime.now().weekday()  # 0=Lunes, 6=Domingo en Python
        # Convertir a DOW de PostgreSQL (0=Domingo, 6=Sábado)
        dow_actual = (dia_actual + 1) % 7

        # Mapeo de DOW a nombre del día
        NOMBRES_DIA = ['Dom', 'Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb']

        logger.info(f"📅 Día actual: {NOMBRES_DIA[dow_actual]} (DOW={dow_actual}), productos con DOW data: {len(promedios_dow_dict)}")

        # ====================================================================
        # PRE-CÁLCULO: Mediana de P75 por categoría (para envío prueba)
        # ====================================================================
        # Para productos sin ventas en NINGUNA tienda, usamos la demanda
        # típica de su categoría como proxy. Esto permite sugerir una
        # cantidad proporcional al tipo de producto en vez de un fijo de 1 bulto.
        from collections import defaultdict
        import statistics as stats_module
        p75_por_categoria: dict = defaultdict(list)
        for row in rows:
            categoria = (row[3] or '').strip().upper()
            p75_local = safe_float(row[19])
            if p75_local > 0 and categoria:
                p75_por_categoria[categoria].append(p75_local)

        p75_mediana_categoria = {
            cat: stats_module.median(vals) for cat, vals in p75_por_categoria.items() if vals
        }
        # Fallback global: mediana de todos los P75 > 0
        all_p75 = [v for vals in p75_por_categoria.values() for v in vals]
        p75_fallback_global = stats_module.median(all_p75) if all_p75 else 1.0

        logger.info(f"📊 P75 mediana por categoría: {len(p75_mediana_categoria)} categorías, fallback global: {p75_fallback_global:.1f}")

        productos = []
        productos_sin_venta_con_ref = 0  # Contador para log
        productos_excluidos_count = 0  # Contador de productos excluidos
        for row in rows:
            codigo = row[0]

            # Saltar productos excluidos para esta tienda
            if codigo in codigos_excluidos:
                productos_excluidos_count += 1
                continue

            categoria_producto = (row[3] or '').strip().upper()  # Categoría normalizada para buscar config

            # Convertir cantidad_bultos de manera segura
            cantidad_bultos = safe_float(row[9], default=1.0)
            if cantidad_bultos <= 0:
                cantidad_bultos = 1.0

            unidad_pedido = row[10] or 'Bulto'  # Unidad de pedido: Bulto, Blister, Cesta, etc.
            unidades_por_bulto = int(cantidad_bultos) if cantidad_bultos > 0 else 1
            prom_20d = safe_float(row[13])
            stock_tienda = safe_float(row[16])
            stock_cedi = safe_float(row[17])
            prom_top3 = safe_float(row[18])
            prom_p75 = safe_float(row[19])
            clase_abc_valor = row[20]  # ABC por ranking de cantidad de SQL
            sigma_demanda = safe_float(row[21])
            demanda_maxima = safe_float(row[22])
            es_generador_trafico = bool(row[23]) if row[23] else False
            # P75 de tiendas de referencia
            p75_referencia = safe_float(row[24])
            tiendas_referencia_str = row[25]  # String con IDs de tiendas separados por coma

            # Clasificacion ABC: usar ABC por ranking del SQL si esta disponible
            venta_diaria_bultos = prom_20d / cantidad_bultos if cantidad_bultos > 0 else 0
            if clase_abc_valor:
                clasificacion = clase_abc_valor
            elif venta_diaria_bultos > 0:
                clasificacion = 'D'  # Default para productos sin clasificacion ABC pero con ventas
            else:
                clasificacion = '-'

            # Stock total en bultos para referencia
            stock_total_bultos = stock_tienda / cantidad_bultos if cantidad_bultos > 0 else 0

            # ========================================================================
            # LÓGICA DE REFERENCIA REGIONAL Y ENVÍO PRUEBA
            # ========================================================================
            # Dos casos diferentes cuando no hay ventas locales suficientes:
            #
            # 1. REFERENCIA REGIONAL: Producto con pocas ventas locales pero con
            #    P75 significativo en tiendas de referencia. Se usa ese P75 para
            #    calcular normalmente con la fórmula ABC.
            #
            # 2. ENVÍO PRUEBA: Producto SIN ventas locales (P75=0) pero que SÍ se
            #    vende en otras tiendas. Se sugiere mínimo 1 bulto para "probar"
            #    el producto en la tienda, independiente del cálculo matemático.
            # ========================================================================

            es_envio_prueba = False      # Sin ventas locales -> probar con 1 bulto mínimo
            usa_referencia_regional = False  # Pocas ventas -> usar P75 de referencia para cálculo
            p75_usado = prom_p75
            nombre_tienda_ref = None

            # Obtener nombre de la primera tienda de referencia (para mensajes)
            if tiendas_referencia_str:
                primera_tienda_id = tiendas_referencia_str.split(',')[0]
                nombre_tienda_ref = tiendas_ref_nombres.get(primera_tienda_id, primera_tienda_id)

            # Caso 1: SIN ventas locales (P75 = 0) pero con referencia -> ENVÍO PRUEBA
            sin_ventas_locales = prom_p75 == 0
            if sin_ventas_locales and p75_referencia > 0 and stock_cedi > 0:
                es_envio_prueba = True
                p75_usado = p75_referencia  # Usar referencia para el cálculo base
                clasificacion = 'D'  # Conservador (clase D para productos nuevos/prueba)
                productos_sin_venta_con_ref += 1

            # Caso 1b: SIN ventas en NINGUNA tienda pero con stock en CEDI -> ENVÍO PRUEBA
            # Regla: "Todo producto que tengamos en el CEDI debería estar en cada tienda"
            # Usa la mediana de P75 de la categoría como proxy de demanda
            elif sin_ventas_locales and p75_referencia == 0 and stock_cedi > 0:
                es_envio_prueba = True
                p75_usado = p75_mediana_categoria.get(categoria_producto, p75_fallback_global)
                clasificacion = 'D'
                productos_sin_venta_con_ref += 1
                logger.debug(f"📦 {codigo}: Envío prueba sin ref, P75 categoría '{categoria_producto}' = {p75_usado:.1f}")

            # Caso 2: POCAS ventas locales pero P75 regional mucho mayor -> REFERENCIA REGIONAL
            # (usar P75 de referencia para calcular, sin forzar mínimos)
            pocas_ventas_locales = prom_p75 > 0 and prom_p75 < 1 and p75_referencia > prom_p75 * 3
            if pocas_ventas_locales and p75_referencia > 0 and stock_cedi > 0:
                usa_referencia_regional = True
                p75_usado = p75_referencia
                # Mantener clasificación original, no forzar a C

            # Usar nuevo modulo de calculo ABC si hay demanda (local o de referencia)
            # Incluir clase D para no dejar productos sin sugerencia (evita ciclo: sin stock → sin venta → sin demanda → sin sugerencia)
            if p75_usado > 0 and clasificacion in ('A', 'B', 'C', 'D'):
                # Para envío prueba o referencia regional, estimar sigma si no hay datos locales
                sigma_usada = sigma_demanda
                if (es_envio_prueba or usa_referencia_regional) and sigma_demanda == 0:
                    sigma_usada = p75_usado * 0.3  # Estimar 30% de variabilidad

                # Determinar si hay override de días por categoría (perecederos)
                # Buscar coincidencia exacta o parcial (categoría del producto contiene la key)
                dias_override = None
                categoria_match = None

                # Primero buscar coincidencia exacta
                if categoria_producto in config_cobertura_categoria:
                    categoria_match = categoria_producto
                else:
                    # Buscar coincidencia parcial: si la categoría del producto CONTIENE alguna key configurada
                    # Ejemplo: producto tiene "FRUVER Y VERDURAS" y config tiene "FRUVER"
                    for cat_config in config_cobertura_categoria.keys():
                        if cat_config and cat_config in categoria_producto:
                            categoria_match = cat_config
                            break

                if categoria_match:
                    config_cat = config_cobertura_categoria[categoria_match]
                    dias_override = config_cat.get(clasificacion, None)
                    if dias_override is not None:
                        logger.debug(f"🥬 {codigo} ({categoria_producto} -> {categoria_match}): Clase {clasificacion} -> {dias_override} días cobertura")
                elif categoria_producto and config_cobertura_categoria:
                    # Log para diagnóstico: categoría del producto no encontrada en config
                    logger.debug(f"⚠️ Categoría '{categoria_producto}' no encontrada en config_cobertura_categoria. Keys disponibles: {list(config_cobertura_categoria.keys())}")

                resultado = calcular_inventario_simple(
                    demanda_p75=p75_usado,  # Usar P75 local o de referencia
                    sigma_demanda=sigma_usada,
                    demanda_maxima=demanda_maxima if demanda_maxima > 0 else p75_usado * 2,
                    unidades_por_bulto=unidades_por_bulto,
                    stock_actual=stock_tienda,
                    stock_cedi=stock_cedi,
                    clase_abc=clasificacion,
                    es_generador_trafico=es_generador_trafico,
                    dias_cobertura_override=dias_override
                )

                # Log diagnóstico para categorías perecederas (ej: FRUVER)
                if categoria_match and dias_override is not None:
                    logger.info(
                        f"🥬 PERECEDERO {codigo}: cat={categoria_producto}→{categoria_match}, "
                        f"clase={clasificacion}, dias_override={dias_override}, "
                        f"P75={p75_usado:.1f}, stock={stock_tienda:.0f}, "
                        f"MAX={resultado.stock_maximo_unid:.1f}, SUG={resultado.cantidad_sugerida_bultos}"
                    )

                # ====================================================================
                # CÁLCULO V2: Cobertura Real por Día de Semana
                # ====================================================================
                # V2 calcula cuánto REALMENTE se necesita para cubrir el período,
                # sumando la demanda específica de cada día de la semana.
                #
                # Ejemplo: Pedido el Miércoles (llega Jueves PM) para cubrir 3 días (clase A):
                # - Necesita cubrir: Vie + Sáb + Dom
                # - Si Sábado vende 2x que Viernes, V2 lo considera
                #
                v2_prom_dow = promedios_dow_dict.get(codigo, [0.0] * 7)
                v2_cobertura_dias = []
                v2_dias_cobertura_real = 0.0
                v2_primer_dia_riesgo = None
                v2_demanda_periodo = 0.0

                # Días de cobertura según config_tienda por clase ABC (o override por categoría)
                # Prioridad: dias_override (por categoría) > config_tienda por clase
                dias_cobertura_por_clase = {
                    'A': config_tienda.dias_cobertura_a,
                    'B': config_tienda.dias_cobertura_b,
                    'C': config_tienda.dias_cobertura_c,
                    'D': config_tienda.dias_cobertura_d,
                    '-': config_tienda.dias_cobertura_a  # Default a clase A si no tiene clasificación
                }
                dias_cobertura_abc = dias_override if dias_override is not None else dias_cobertura_por_clase.get(clasificacion, config_tienda.dias_cobertura_a)

                # Convertir a bultos para mostrar
                unid_por_bulto = unidades_por_bulto if unidades_por_bulto > 0 else 1

                # Lead time desde config_tienda (default 1.5 días → llega mañana PM)
                lead_time_dias = config_tienda.lead_time
                # Si lead_time es 1 o 1.5, llega mañana; primer día que cubre es pasado mañana
                dias_hasta_llegada = int(lead_time_dias) if lead_time_dias >= 1 else 1
                dias_hasta_cobertura = dias_hasta_llegada + 1  # Día siguiente a la llegada

                # dow_actual viene de Python weekday() convertido: 0=Dom, 1=Lun, ..., 6=Sáb
                dow_llegada = (dow_actual + dias_hasta_llegada) % 7
                dow_primer_dia_cubrir = (dow_actual + dias_hasta_cobertura) % 7

                # Calcular fechas exactas
                fecha_llegada = fecha_actual + timedelta(days=dias_hasta_llegada)
                fecha_primer_dia_cubrir = fecha_actual + timedelta(days=dias_hasta_cobertura)

                MESES_ES = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

                v2_dia_pedido = NOMBRES_DIA[dow_actual]
                v2_dia_llegada = NOMBRES_DIA[dow_llegada]
                v2_fecha_pedido = f"{fecha_actual.day} {MESES_ES[fecha_actual.month - 1]}"
                v2_fecha_llegada = f"{fecha_llegada.day} {MESES_ES[fecha_llegada.month - 1]}"
                v2_dias_cobertura_config = int(dias_cobertura_abc)
                v2_lead_time_config = lead_time_dias

                # ============================================================
                # PASO 1: Calcular demanda total del período (días específicos)
                # ============================================================
                demanda_total_periodo = 0.0
                dias_a_cubrir = []

                for i in range(int(dias_cobertura_abc)):
                    dow_dia = (dow_primer_dia_cubrir + i) % 7
                    nombre_dia = NOMBRES_DIA[dow_dia]
                    fecha_dia = fecha_primer_dia_cubrir + timedelta(days=i)
                    fecha_dia_str = f"{fecha_dia.day} {MESES_ES[fecha_dia.month - 1]}"

                    # Demanda de este día específico
                    demanda_dia = v2_prom_dow[dow_dia] if v2_prom_dow and v2_prom_dow[dow_dia] > 0 else prom_20d

                    # Si no hay datos de DOW para ningún día, usar prom_20d
                    if demanda_dia == 0 and (not v2_prom_dow or sum(v2_prom_dow) == 0):
                        demanda_dia = prom_20d

                    dias_a_cubrir.append({
                        "dow": dow_dia,
                        "nombre": nombre_dia,
                        "fecha": fecha_dia_str,
                        "demanda": demanda_dia
                    })
                    demanda_total_periodo += demanda_dia

                v2_demanda_periodo = demanda_total_periodo

                # Construir histórico detallado por DOW para este producto
                v2_historico_dow = []
                historico_producto = historico_dow_dict.get(codigo, {})
                for dow in range(7):
                    datos_dow = historico_producto.get(dow, [])
                    promedio = v2_prom_dow[dow] if v2_prom_dow else 0.0
                    v2_historico_dow.append({
                        "dow": dow,
                        "nombre": NOMBRES_DIA[dow],
                        "promedio": round(promedio, 1),
                        "dias_con_data": len(datos_dow),
                        "detalle": datos_dow  # [{fecha: "18 Dic", venta: 25}, ...]
                    })

                # ============================================================
                # PASO 2: Calcular cantidad sugerida V2
                # ============================================================
                # Necesitamos: demanda_total_periodo + stock_seguridad - stock_actual
                # Usamos el mismo stock de seguridad que V1 para consistencia
                stock_seguridad_v2 = resultado.stock_seguridad_unid

                # Cantidad necesaria = demanda del período + seguridad - lo que ya tenemos
                cantidad_necesaria_unid = demanda_total_periodo + stock_seguridad_v2 - stock_tienda
                cantidad_necesaria_unid = max(0, cantidad_necesaria_unid)  # No puede ser negativo

                # Redondear a bultos completos (hacia arriba)
                v2_cantidad_sugerida_bultos = math.ceil(cantidad_necesaria_unid / unid_por_bulto) if unid_por_bulto > 0 else 0
                v2_cantidad_sugerida_unid = v2_cantidad_sugerida_bultos * unid_por_bulto

                # Diferencia vs V1
                v2_diferencia_bultos = v2_cantidad_sugerida_bultos - resultado.cantidad_sugerida_bultos

                # ============================================================
                # PASO 3: Simular cobertura día a día CON la cantidad V2
                # ============================================================
                stock_disponible_v2 = stock_tienda + v2_cantidad_sugerida_unid
                stock_restante = stock_disponible_v2
                dias_cubiertos = 0.0

                for dia_info in dias_a_cubrir:
                    demanda_dia = dia_info["demanda"]
                    nombre_dia = dia_info["nombre"]
                    dow_dia = dia_info["dow"]
                    fecha_dia_str = dia_info["fecha"]

                    # Calcular cobertura para este día
                    if demanda_dia > 0:
                        cobertura_pct = min(100, (stock_restante / demanda_dia) * 100)
                        stock_despues = max(0, stock_restante - demanda_dia)
                    else:
                        cobertura_pct = 100
                        stock_despues = stock_restante

                    # Determinar estado del día
                    if cobertura_pct >= 100:
                        estado = "ok"
                        dias_cubiertos += 1.0
                    elif cobertura_pct >= 50:
                        estado = "riesgo"
                        dias_cubiertos += cobertura_pct / 100
                        if v2_primer_dia_riesgo is None:
                            v2_primer_dia_riesgo = nombre_dia
                    else:
                        estado = "quiebre"
                        dias_cubiertos += cobertura_pct / 100
                        if v2_primer_dia_riesgo is None:
                            v2_primer_dia_riesgo = nombre_dia

                    v2_cobertura_dias.append({
                        "dia": nombre_dia,
                        "fecha": fecha_dia_str,
                        "dow": dow_dia,
                        "demanda_unid": round(demanda_dia, 1),
                        "demanda_bultos": round(demanda_dia / unid_por_bulto, 2),
                        "stock_antes": round(stock_restante, 1),
                        "stock_despues": round(stock_despues, 1),
                        "cobertura_pct": round(cobertura_pct, 0),
                        "estado": estado
                    })

                    stock_restante = stock_despues

                v2_dias_cobertura_real = round(dias_cubiertos, 1)

                # SOBRESTOCK: Si hay sobrestock, NO sugerir nada (prioridad máxima)
                # Esto evita que lógicas posteriores (mínimo exhibición, envío prueba) eleven la cantidad
                cantidad_sugerida_bultos_final = resultado.cantidad_sugerida_bultos
                cantidad_sugerida_unid_final = resultado.cantidad_sugerida_unid

                if resultado.tiene_sobrestock:
                    # Forzar cantidad a 0 - hay sobrestock, no pedir nada
                    cantidad_sugerida_bultos_final = 0
                    cantidad_sugerida_unid_final = 0.0
                    logger.debug(f"🔴 {codigo}: Sobrestock detectado - cantidad forzada a 0")
                elif es_envio_prueba and resultado.cantidad_sugerida_bultos == 0:
                    # ENVÍO PRUEBA: Garantizar mínimo 1 bulto (solo si NO hay sobrestock)
                    cantidad_sugerida_bultos_final = 1
                    cantidad_sugerida_unid_final = float(unidades_por_bulto)
                    logger.info(f"📦 {codigo}: Envío prueba forzado a 1 bulto ({unidades_por_bulto} unid)")

                # ====================================================================
                # LÍMITES DE INVENTARIO: MÍNIMO EXHIBICIÓN Y CAPACIDAD MÁXIMA
                # ====================================================================
                # 1. Mínimo de exhibición: Si el producto necesita una cantidad mínima
                #    para "verse bien" en el mostrador, elevar la sugerencia si es menor.
                # 2. Capacidad máxima: Si tiene límite de capacidad (congelador, etc.),
                #    no exceder el espacio disponible.
                limite_info = limites_inventario.get(codigo)
                cantidad_antes_capacidad = None
                ajustado_por_capacidad = False
                tipo_restriccion_cap = None
                notas_capacidad = None
                minimo_exhibicion_aplicado = None

                if limite_info and not resultado.tiene_sobrestock:
                    # NOTA: No aplicar límites de inventario si hay sobrestock
                    # (ya tenemos demasiado stock, no tiene sentido elevar por exhibición)
                    tipo_restriccion_cap = limite_info['tipo_restriccion']
                    notas_capacidad = limite_info['notas']

                    # MÍNIMO DE EXHIBICIÓN: elevar cantidad si es necesario para exhibición
                    minimo_exhibicion = limite_info.get('minimo_exhibicion')
                    if minimo_exhibicion and minimo_exhibicion > 0:
                        # Calcular cuántas unidades necesitamos para alcanzar el mínimo de exhibición
                        unidades_necesarias_exhibicion = max(0, minimo_exhibicion - stock_tienda)

                        if unidades_necesarias_exhibicion > cantidad_sugerida_unid_final:
                            cantidad_antes_minimo = cantidad_sugerida_unid_final
                            cantidad_sugerida_unid_final = unidades_necesarias_exhibicion
                            cantidad_sugerida_bultos_final = math.ceil(unidades_necesarias_exhibicion / unidades_por_bulto) if unidades_por_bulto > 0 else 0
                            minimo_exhibicion_aplicado = minimo_exhibicion

                            logger.info(
                                f"📊 {codigo}: Elevado por mínimo exhibición. "
                                f"Original: {cantidad_antes_minimo:.0f} → Elevado: {cantidad_sugerida_unid_final:.0f} unid "
                                f"(Mín exhibición: {minimo_exhibicion:.0f}, Stock: {stock_tienda:.0f})"
                            )

                    # CAPACIDAD MÁXIMA: no exceder el espacio disponible
                    capacidad_maxima = limite_info.get('capacidad_maxima')
                    if capacidad_maxima and capacidad_maxima > 0:
                        # Espacio disponible = capacidad máxima - stock actual
                        espacio_disponible = max(0, capacidad_maxima - stock_tienda)

                        # Si la cantidad sugerida excede el espacio disponible, ajustar
                        if cantidad_sugerida_unid_final > espacio_disponible:
                            cantidad_antes_capacidad = cantidad_sugerida_unid_final
                            cantidad_sugerida_unid_final = espacio_disponible
                            cantidad_sugerida_bultos_final = math.ceil(espacio_disponible / unidades_por_bulto) if unidades_por_bulto > 0 else 0
                            ajustado_por_capacidad = True

                            # Nombre legible del tipo de restricción
                            tipo_legible = {
                                'congelador': 'congelador',
                                'refrigerador': 'refrigerador',
                                'anaquel': 'anaquel',
                                'piso': 'espacio en piso',
                                'exhibidor': 'exhibidor',
                                'espacio_fisico': 'espacio físico'
                            }.get(tipo_restriccion_cap, tipo_restriccion_cap)

                            logger.info(
                                f"⚠️ {codigo}: Ajustado por capacidad de {tipo_legible}. "
                                f"Original: {cantidad_antes_capacidad:.0f} → Ajustado: {cantidad_sugerida_unid_final:.0f} unid "
                                f"(Cap. máx: {capacidad_maxima:.0f}, Stock: {stock_tienda:.0f}, Disponible: {espacio_disponible:.0f})"
                            )

                # Determinar razon del pedido y método de cálculo
                if resultado.tiene_sobrestock:
                    razon = "Sobrestock - No pedir"
                    metodo_usado = resultado.metodo_usado
                elif es_envio_prueba:
                    # Producto SIN ventas locales -> Envío de prueba (mínimo 1 bulto)
                    if nombre_tienda_ref:
                        razon = f"Envío prueba (ref: {nombre_tienda_ref})"
                    else:
                        razon = "Envío prueba"
                    metodo_usado = "envio_prueba"
                elif usa_referencia_regional:
                    # Producto con POCAS ventas locales -> usar P75 de referencia
                    if nombre_tienda_ref:
                        razon = f"P75 ref: {nombre_tienda_ref}"
                    else:
                        razon = "P75 referencia regional"
                    metodo_usado = "referencia_regional"
                else:
                    razon = ""  # Usuario puede agregar notas manualmente
                    metodo_usado = resultado.metodo_usado

                # Calcular dias cobertura actual
                stock_dias = stock_tienda / p75_usado if p75_usado > 0 else 999

                # Agregar warnings explicativos
                warnings = resultado.warnings.copy() if resultado.warnings else []
                if es_envio_prueba:
                    warnings.append(f"Sin ventas locales. P75 de {nombre_tienda_ref or 'región'}: {p75_usado:.1f} unid/día")
                elif usa_referencia_regional:
                    warnings.append(f"Ventas locales bajas. Usando P75 de {nombre_tienda_ref or 'región'}: {p75_usado:.1f} unid/día")

                # Warning de mínimo de exhibición (importante para visibilidad)
                if minimo_exhibicion_aplicado:
                    warnings.append(
                        f"📊 MÍNIMO EXHIBICIÓN: Elevado para alcanzar {minimo_exhibicion_aplicado:.0f} unidades "
                        f"(mínimo para que el producto se vea bien en exhibición). "
                        f"Stock actual: {stock_tienda:.0f}"
                    )

                # Warning de capacidad limitada (muy importante para visibilidad)
                if ajustado_por_capacidad and cantidad_antes_capacidad is not None:
                    tipo_legible = {
                        'congelador': 'congelador',
                        'refrigerador': 'refrigerador',
                        'anaquel': 'anaquel',
                        'piso': 'espacio en piso',
                        'exhibidor': 'exhibidor',
                        'espacio_fisico': 'espacio físico'
                    }.get(tipo_restriccion_cap, tipo_restriccion_cap)

                    cap_maxima = limite_info['capacidad_maxima']
                    espacio_disp = max(0, cap_maxima - stock_tienda)

                    warnings.append(
                        f"⚠️ CAPACIDAD LIMITADA: Ajustado de {cantidad_antes_capacidad:.0f} a {cantidad_sugerida_unid_final:.0f} unidades "
                        f"por capacidad máxima de {tipo_legible} ({cap_maxima:.0f} unid). "
                        f"Stock actual: {stock_tienda:.0f} | Espacio disponible: {espacio_disp:.0f}"
                    )

                productos.append(ProductoCalculado(
                    codigo_producto=codigo,
                    codigo_barras=row[1],
                    descripcion_producto=row[2] or codigo,
                    categoria=row[3],
                    grupo=row[4],
                    subgrupo=row[5],
                    marca=row[6],
                    presentacion=row[7],
                    cuadrante_producto=row[8],
                    cantidad_bultos=cantidad_bultos,
                    unidad_pedido=unidad_pedido,
                    peso_unidad=safe_float(row[11], default=1000.0),
                    prom_ventas_5dias_unid=safe_float(row[12]),
                    prom_ventas_20dias_unid=prom_20d,
                    prom_top3_unid=prom_top3,
                    prom_p75_unid=p75_usado,  # Usar P75 local o de referencia
                    prom_ventas_8sem_unid=prom_20d,  # Aproximacion
                    prom_ventas_8sem_bultos=venta_diaria_bultos,
                    stock_tienda=stock_tienda,
                    stock_en_transito=0.0,
                    stock_total=stock_tienda,
                    stock_total_bultos=stock_total_bultos,
                    stock_dias_cobertura=stock_dias,
                    stock_cedi_origen=stock_cedi,
                    clasificacion_abc=clasificacion,
                    clase_efectiva=resultado.clase_efectiva,
                    es_generador_trafico=es_generador_trafico,
                    stock_minimo=resultado.stock_minimo_unid,
                    stock_maximo=resultado.stock_maximo_unid,
                    stock_seguridad=resultado.stock_seguridad_unid,
                    punto_reorden=resultado.punto_reorden_unid,
                    cantidad_sugerida_unid=cantidad_sugerida_unid_final,
                    cantidad_sugerida_bultos=float(cantidad_sugerida_bultos_final),
                    cantidad_ajustada_bultos=float(cantidad_sugerida_bultos_final),
                    razon_pedido=razon,
                    metodo_calculo=metodo_usado,
                    tiene_sobrestock=resultado.tiene_sobrestock,
                    exceso_unidades=resultado.exceso_unidades,
                    exceso_bultos=resultado.exceso_bultos,
                    dias_exceso=resultado.dias_exceso,
                    # Límites de inventario
                    capacidad_maxima_configurada=limite_info['capacidad_maxima'] if limite_info else None,
                    cantidad_antes_ajuste_capacidad=cantidad_antes_capacidad,
                    ajustado_por_capacidad=ajustado_por_capacidad,
                    tipo_restriccion_capacidad=tipo_restriccion_cap,
                    notas_capacidad=notas_capacidad,
                    minimo_exhibicion_configurado=limite_info['minimo_exhibicion'] if limite_info else None,
                    ajustado_por_minimo_exhibicion=minimo_exhibicion_aplicado is not None,
                    warnings_calculo=warnings,
                    # === CAMPOS V2 (Cobertura Real por Día) ===
                    v2_prom_dow=v2_prom_dow,
                    v2_demanda_periodo=round(v2_demanda_periodo, 1),
                    v2_cantidad_sugerida_unid=v2_cantidad_sugerida_unid,
                    v2_cantidad_sugerida_bultos=float(v2_cantidad_sugerida_bultos),
                    v2_diferencia_bultos=float(v2_diferencia_bultos),
                    v2_cobertura_dias=v2_cobertura_dias,
                    v2_dias_cobertura_real=v2_dias_cobertura_real,
                    v2_primer_dia_riesgo=v2_primer_dia_riesgo,
                    v2_dia_pedido=v2_dia_pedido,
                    v2_dia_llegada=v2_dia_llegada,
                    v2_fecha_pedido=v2_fecha_pedido,
                    v2_fecha_llegada=v2_fecha_llegada,
                    v2_dias_cobertura_config=v2_dias_cobertura_config,
                    v2_lead_time_config=v2_lead_time_config,
                    v2_historico_dow=v2_historico_dow
                ))
            else:
                # Producto sin demanda o sin clasificacion - no sugerir pedido
                productos.append(ProductoCalculado(
                    codigo_producto=codigo,
                    codigo_barras=row[1],
                    descripcion_producto=row[2] or codigo,
                    categoria=row[3],
                    grupo=row[4],
                    subgrupo=row[5],
                    marca=row[6],
                    presentacion=row[7],
                    cantidad_bultos=cantidad_bultos,
                    unidad_pedido=unidad_pedido,
                    peso_unidad=safe_float(row[11], default=1000.0),
                    prom_ventas_5dias_unid=safe_float(row[12]),
                    prom_ventas_20dias_unid=prom_20d,
                    prom_top3_unid=prom_top3,
                    prom_p75_unid=prom_p75,
                    prom_ventas_8sem_unid=prom_20d,
                    prom_ventas_8sem_bultos=venta_diaria_bultos,
                    stock_tienda=stock_tienda,
                    stock_en_transito=0.0,
                    stock_total=stock_tienda,
                    stock_total_bultos=stock_total_bultos,
                    stock_dias_cobertura=999,
                    stock_cedi_origen=stock_cedi,
                    clasificacion_abc=clasificacion,
                    clase_efectiva=clasificacion,
                    es_generador_trafico=es_generador_trafico,
                    stock_minimo=0.0,
                    stock_maximo=0.0,
                    stock_seguridad=0.0,
                    punto_reorden=0.0,
                    cantidad_sugerida_unid=0.0,
                    cantidad_sugerida_bultos=0.0,
                    cantidad_ajustada_bultos=0.0,
                    razon_pedido="Sin demanda historica",
                    metodo_calculo="",
                    tiene_sobrestock=False,
                    exceso_unidades=0.0,
                    exceso_bultos=0,
                    dias_exceso=0.0,
                    warnings_calculo=[]
                ))

        logger.info(f"✅ Calculados {len(productos)} productos sugeridos")
        if productos_sin_venta_con_ref > 0:
            logger.info(f"📦 Incluidos {productos_sin_venta_con_ref} productos sin ventas locales (envíos de prueba basados en región {tienda_region})")
        if productos_excluidos_count > 0:
            logger.info(f"🚫 Excluidos {productos_excluidos_count} productos de la lista de exclusiones")
        return productos

    except Exception as e:
        import traceback
        logger.error(f"❌ Error calculando productos sugeridos: {str(e)}")
        logger.error(f"Traceback completo:\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error calculando productos: {str(e)}")


@router.post("/crear-v2", response_model=PedidoGuardadoResponse)
async def crear_pedido_v2(
    request: CrearPedidoV2Request,
    conn: Any = Depends(get_db_write)
):
    """
    Crea un pedido sugerido usando el método v2.0 (Nivel Objetivo ABC-XYZ)
    """
    try:
        logger.info(f"🆕 Creando pedido v2.0: Destino={request.tienda_destino_id}")
        cursor = conn.cursor()

        # Generar IDs
        pedido_id = str(uuid.uuid4())

        # Generar número de pedido
        cursor.execute("SELECT MAX(numero_pedido) FROM pedidos_sugeridos WHERE numero_pedido LIKE 'PS-%%'")
        result = cursor.fetchone()
        if result and result[0]:
            ultimo = int(result[0].split('-')[1])
            numero_pedido = f"PS-{ultimo + 1:05d}"
        else:
            numero_pedido = "PS-00001"

        # Obtener nombres de ubicaciones
        cursor.execute("SELECT nombre FROM ubicaciones WHERE id = %s", [request.cedi_origen_id])
        cedi_nombre = cursor.fetchone()
        cursor.execute("SELECT nombre FROM ubicaciones WHERE id = %s", [request.tienda_destino_id])
        tienda_nombre = cursor.fetchone()

        if not cedi_nombre or not tienda_nombre:
            cursor.close()
            raise HTTPException(status_code=404, detail="Ubicación no encontrada")

        # Calcular totales
        total_productos = len(request.productos)
        total_unidades = sum(p.cantidad_pedida for p in request.productos)

        # Insertar pedido principal
        cursor.execute("""
            INSERT INTO pedidos_sugeridos (
                id, numero_pedido,
                cedi_origen_id, cedi_origen_nombre,
                tienda_destino_id, tienda_destino_nombre,
                fecha_pedido,
                estado, requiere_aprobacion,
                total_productos, total_unidades,
                tipo_pedido, prioridad,
                usuario_creador, fecha_creacion,
                version, metodo_calculo
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, (CURRENT_TIMESTAMP AT TIME ZONE 'UTC'), 2, %s)
        """, [
            pedido_id, numero_pedido,
            request.cedi_origen_id, cedi_nombre[0],
            request.tienda_destino_id, tienda_nombre[0],
            request.fecha_pedido,
            ESTADO_BORRADOR, False,
            total_productos, float(total_unidades),
            request.tipo_pedido, "normal",
            "sistema",
            request.metodo_calculo
        ])

        # Insertar detalles de productos
        for idx, producto in enumerate(request.productos):
            detalle_id = str(uuid.uuid4())

            # Obtener información adicional del producto
            cursor.execute("""
                SELECT
                    codigo,
                    COALESCE(nombre, descripcion) as descripcion,
                    categoria,
                    marca
                FROM productos
                WHERE codigo = %s
            """, [producto.producto_id])
            prod_info = cursor.fetchone()

            if not prod_info:
                logger.warning(f"Producto {producto.producto_id} no encontrado, saltando...")
                continue

            codigo, descripcion, categoria, marca = prod_info

            cursor.execute("""
                INSERT INTO pedidos_sugeridos_detalle (
                    id, pedido_id, linea_numero,
                    codigo_producto, descripcion_producto,
                    categoria, marca,
                    cuadrante_producto,
                    cantidad_sugerida_unidades,
                    cantidad_pedida_unidades,
                    total_unidades,
                    stock_tienda,
                    stock_minimo,
                    incluido,
                    fecha_creacion
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            """, [
                detalle_id, pedido_id, idx + 1,
                codigo, descripcion,
                categoria, marca,
                producto.matriz_abc_xyz,
                float(producto.cantidad_sugerida),
                float(producto.cantidad_pedida),
                float(producto.cantidad_pedida),
                float(producto.stock_actual),
                float(producto.nivel_objetivo),
                True
            ])

        conn.commit()
        cursor.close()

        logger.info(f"✅ Pedido v2.0 creado: {numero_pedido} con {total_productos} productos")

        return PedidoGuardadoResponse(
            id=pedido_id,
            numero_pedido=numero_pedido,
            estado=ESTADO_BORRADOR,
            total_productos=total_productos,
            total_bultos=0.0,
            fecha_creacion=datetime.now(),
            mensaje=f"Pedido {numero_pedido} creado exitosamente"
        )

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error creando pedido v2.0: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creando pedido: {str(e)}")


# =====================================================================================
# VERIFICAR LLEGADA - Detectar incrementos de inventario
# =====================================================================================

@router.get("/{pedido_id}/verificar-llegada", response_model=VerificarLlegadaResponse)
async def verificar_llegada(
    pedido_id: str,
    conn: Any = Depends(get_db)
):
    """
    Verifica si los productos de un pedido llegaron a la tienda.

    Detecta incrementos de inventario desde la fecha del pedido hasta ahora,
    sumando todos los incrementos positivos (llegadas) e ignorando decrementos (ventas).

    Funciona en cualquier estado del pedido (borrador, aprobado, etc.)
    """
    try:
        cursor = conn.cursor()

        # 1. Obtener información del pedido
        cursor.execute("""
            SELECT
                p.id, p.numero_pedido, p.fecha_pedido,
                p.tienda_destino_id, p.tienda_destino_nombre
            FROM pedidos_sugeridos p
            WHERE p.id = %s
        """, [pedido_id])

        pedido_row = cursor.fetchone()
        if not pedido_row:
            raise HTTPException(status_code=404, detail=f"Pedido {pedido_id} no encontrado")

        _, numero_pedido, fecha_pedido, tienda_destino_id, tienda_destino_nombre = pedido_row

        # 2. Obtener productos del pedido con sus cantidades, factor de conversión, clasificación ABC y stocks
        # La clasificación ABC viene de productos_abc_tienda (específica por tienda)
        # El factor de conversión (unidades_por_bulto) viene de la tabla productos
        # Los stocks vienen de inventario_actual para tienda, cedi_caracas y cedi_verde
        # IMPORTANTE: inventario_historico.producto_id almacena el CÓDIGO del producto, no productos.id
        cursor.execute("""
            SELECT
                d.codigo_producto,
                d.descripcion_producto,
                COALESCE(d.cantidad_pedida_bultos, 0) as cantidad_pedida_bultos,
                COALESCE(d.cantidad_pedida_unidades, d.total_unidades, 0) as cantidad_pedida_unidades,
                COALESCE(d.cantidad_recibida_bultos, 0) as cantidad_recibida_bultos,
                d.codigo_producto as producto_id_historico,
                COALESCE(p.unidades_por_bulto, 1) as unidades_x_bulto,
                'bultos' as unidad,
                COALESCE(abc.clase_abc, 'D') as clasificacion_abc,
                COALESCE(inv_tienda.cantidad, 0) as stock_tienda,
                COALESCE(inv_caracas.cantidad, 0) as stock_cedi_caracas,
                COALESCE(inv_verde.cantidad, 0) as stock_cedi_verde
            FROM pedidos_sugeridos_detalle d
            LEFT JOIN productos p ON d.codigo_producto = p.codigo
            LEFT JOIN productos_abc_tienda abc ON abc.producto_id = p.id AND abc.ubicacion_id = %s
            LEFT JOIN inventario_actual inv_tienda ON inv_tienda.producto_id = d.codigo_producto AND inv_tienda.ubicacion_id = %s
            LEFT JOIN inventario_actual inv_caracas ON inv_caracas.producto_id = d.codigo_producto AND inv_caracas.ubicacion_id = 'cedi_caracas'
            LEFT JOIN inventario_actual inv_verde ON inv_verde.producto_id = d.codigo_producto AND inv_verde.ubicacion_id = 'cedi_verde'
            WHERE d.pedido_id = %s AND d.incluido = true
            ORDER BY d.linea_numero
        """, [tienda_destino_id, tienda_destino_id, pedido_id])

        productos_pedido = cursor.fetchall()

        if not productos_pedido:
            raise HTTPException(status_code=404, detail="El pedido no tiene productos")

        # 3. Procesar cada producto
        productos_verificados = []
        productos_completos = 0
        productos_parciales = 0
        productos_no_llegaron = 0
        hay_nuevos_incrementos = False

        # Tolerancia del 3% para considerar llegada "completa"
        TOLERANCIA_COMPLETO = Decimal('97')

        for prod_row in productos_pedido:
            (codigo_producto, descripcion, cant_pedida_bultos, cant_pedida_unidades,
             cant_ya_guardada, producto_id_historico, unidades_x_bulto, unidad, clasificacion_abc,
             stock_tienda, stock_cedi_caracas, stock_cedi_verde) = prod_row
            unidades_x_bulto = Decimal(str(unidades_x_bulto or 1))
            unidad = unidad or 'bultos'
            clasificacion_abc = clasificacion_abc or 'D'
            stock_tienda = Decimal(str(stock_tienda or 0))
            stock_cedi_caracas = Decimal(str(stock_cedi_caracas or 0))
            stock_cedi_verde = Decimal(str(stock_cedi_verde or 0))

            # Query para detectar incrementos desde fecha_pedido
            # IMPORTANTE: inventario_historico.producto_id usa el CÓDIGO del producto
            # Suma todos los incrementos positivos entre snapshots consecutivos
            cursor.execute("""
                WITH snapshots AS (
                    SELECT
                        fecha_snapshot,
                        SUM(cantidad) as cantidad,
                        LAG(SUM(cantidad)) OVER (ORDER BY fecha_snapshot) as cantidad_anterior
                    FROM inventario_historico
                    WHERE producto_id = %s
                      AND ubicacion_id = %s
                      AND fecha_snapshot >= %s
                    GROUP BY fecha_snapshot
                    ORDER BY fecha_snapshot
                ),
                incrementos AS (
                    SELECT
                        fecha_snapshot,
                        cantidad,
                        cantidad_anterior,
                        CASE
                            WHEN cantidad_anterior IS NOT NULL AND cantidad > cantidad_anterior
                            THEN cantidad - cantidad_anterior
                            ELSE 0
                        END as incremento
                    FROM snapshots
                )
                SELECT
                    COALESCE(SUM(incremento), 0) as total_llegadas,
                    MIN(CASE WHEN incremento > 0 THEN fecha_snapshot END) as fecha_primer_incremento,
                    (SELECT cantidad FROM snapshots ORDER BY fecha_snapshot ASC LIMIT 1) as snapshot_inicial,
                    (SELECT cantidad FROM snapshots ORDER BY fecha_snapshot DESC LIMIT 1) as snapshot_final,
                    COUNT(*) as num_snapshots
                FROM incrementos
            """, [producto_id_historico, tienda_destino_id, fecha_pedido])

            result = cursor.fetchone()
            total_llegadas_unidades, fecha_primer_incremento, snapshot_inicial, snapshot_final, num_snapshots = result

            total_llegadas_unidades = Decimal(str(total_llegadas_unidades or 0))
            cant_ya_guardada = Decimal(str(cant_ya_guardada or 0))
            cant_pedida_bultos = Decimal(str(cant_pedida_bultos or 0))
            cant_pedida_unidades = Decimal(str(cant_pedida_unidades or 0))

            # Convertir llegadas de unidades a bultos
            total_llegadas_bultos = total_llegadas_unidades / unidades_x_bulto

            # Calcular nuevo incremento en bultos (lo que no se ha guardado aún)
            nuevo_incremento = max(Decimal(0), total_llegadas_bultos - cant_ya_guardada)

            # Calcular porcentaje basado en bultos pedidos
            if cant_pedida_bultos > 0:
                porcentaje = (total_llegadas_bultos / cant_pedida_bultos) * 100
            else:
                porcentaje = Decimal(0)

            # Determinar estado (con tolerancia del 3%)
            # Si no hay datos de inventario, asumimos inventario cero
            # Por lo tanto, si no hay incrementos detectados = no llegó
            tiene_datos = num_snapshots > 1 if num_snapshots else False

            if porcentaje >= TOLERANCIA_COMPLETO:
                estado = EstadoLlegada.COMPLETO
                productos_completos += 1
                mensaje = "Llegada completa"
            elif porcentaje > 0:
                estado = EstadoLlegada.PARCIAL
                productos_parciales += 1
                mensaje = f"Llegada parcial ({porcentaje:.0f}%)"
            else:
                # Si no hay datos de inventario o porcentaje es 0, es "No llegó"
                # Asumimos que si no hay histórico, el inventario era cero
                estado = EstadoLlegada.NO_LLEGO
                productos_no_llegaron += 1
                if not tiene_datos:
                    mensaje = "Sin histórico de inventario (asumido como no llegó)"
                else:
                    mensaje = "No se detectó llegada"

            if nuevo_incremento > 0:
                hay_nuevos_incrementos = True

            productos_verificados.append(ProductoLlegadaVerificacion(
                codigo_producto=codigo_producto,
                descripcion_producto=descripcion or "",
                cantidad_pedida_bultos=cant_pedida_bultos,
                cantidad_pedida_unidades=cant_pedida_unidades,
                unidad=unidad,  # bultos, cestas, blister, etc.
                unidades_x_bulto=int(unidades_x_bulto),  # Factor de conversión
                clasificacion_abc=clasificacion_abc,  # A, B, C, D
                total_llegadas_detectadas=total_llegadas_bultos.quantize(Decimal('0.01')),
                cantidad_ya_guardada=cant_ya_guardada,
                nuevo_incremento=nuevo_incremento.quantize(Decimal('0.01')),
                porcentaje_llegada=min(porcentaje, Decimal(150)),  # Cap at 150%
                estado_llegada=estado,
                tiene_datos=tiene_datos,
                mensaje=mensaje,
                stock_tienda=stock_tienda,
                stock_cedi_caracas=stock_cedi_caracas,
                stock_cedi_verde=stock_cedi_verde,
                snapshot_inicial=Decimal(str(snapshot_inicial)) if snapshot_inicial else None,
                snapshot_final=Decimal(str(snapshot_final)) if snapshot_final else None,
                fecha_primer_incremento=fecha_primer_incremento
            ))

        cursor.close()

        # Calcular porcentaje global = productos completos / total productos
        # Ya no excluimos "sin_datos" porque ahora se tratan como "no_llego"
        total_productos = len(productos_verificados)
        if total_productos > 0:
            porcentaje_global = Decimal(productos_completos) / Decimal(total_productos) * 100
        else:
            porcentaje_global = Decimal(0)

        return VerificarLlegadaResponse(
            pedido_id=pedido_id,
            numero_pedido=numero_pedido,
            tienda_destino_id=tienda_destino_id,
            tienda_destino_nombre=tienda_destino_nombre,
            fecha_pedido=fecha_pedido,
            fecha_verificacion=datetime.now(),
            productos=productos_verificados,
            total_productos=total_productos,
            productos_completos=productos_completos,
            productos_parciales=productos_parciales,
            productos_no_llegaron=productos_no_llegaron,
            productos_sin_datos=0,  # Ya no usamos esta categoría
            porcentaje_cumplimiento_global=porcentaje_global.quantize(Decimal('0.1')),
            tiene_datos_suficientes=True,  # Siempre tenemos datos (asumimos 0 si no hay histórico)
            hay_nuevos_incrementos=hay_nuevos_incrementos,
            mensaje="Verificación completada"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error verificando llegada: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error verificando llegada: {str(e)}")


# =====================================================================================
# REGISTRAR LLEGADA - Guardar las llegadas detectadas
# =====================================================================================

@router.post("/{pedido_id}/registrar-llegada", response_model=RegistrarLlegadaResponse)
async def registrar_llegada(
    pedido_id: str,
    request: RegistrarLlegadaRequest,
    conn: Any = Depends(get_db_write)
):
    """
    Registra/guarda las llegadas detectadas en el pedido.

    Actualiza el campo cantidad_recibida_bultos de cada producto,
    sumando la cantidad_llegada al valor existente.
    """
    try:
        cursor = conn.cursor()

        # 1. Verificar que el pedido existe
        cursor.execute("""
            SELECT numero_pedido FROM pedidos_sugeridos WHERE id = %s
        """, [pedido_id])

        pedido_row = cursor.fetchone()
        if not pedido_row:
            raise HTTPException(status_code=404, detail=f"Pedido {pedido_id} no encontrado")

        numero_pedido = pedido_row[0]

        # 2. Actualizar cada producto
        productos_actualizados = 0

        for producto in request.productos:
            # Sumar la cantidad_llegada a lo que ya existe
            cursor.execute("""
                UPDATE pedidos_sugeridos_detalle
                SET
                    cantidad_recibida_bultos = COALESCE(cantidad_recibida_bultos, 0) + %s,
                    cantidad_recibida_unidades = COALESCE(cantidad_recibida_unidades, 0) + %s,
                    fecha_modificacion = CURRENT_TIMESTAMP
                WHERE pedido_id = %s AND codigo_producto = %s
            """, [
                float(producto.cantidad_llegada),
                float(producto.cantidad_llegada),  # También actualizamos unidades
                pedido_id,
                producto.codigo_producto
            ])

            if cursor.rowcount > 0:
                productos_actualizados += 1

        conn.commit()
        cursor.close()

        logger.info(f"✅ Llegadas registradas para pedido {numero_pedido}: {productos_actualizados} productos")

        return RegistrarLlegadaResponse(
            pedido_id=pedido_id,
            numero_pedido=numero_pedido,
            productos_actualizados=productos_actualizados,
            mensaje=f"Llegadas registradas exitosamente para {productos_actualizados} productos"
        )

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error registrando llegada: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error registrando llegada: {str(e)}")


@router.get("/{pedido_id}/exportar-excel")
async def exportar_pedido_excel(
    pedido_id: str,
    conn: Any = Depends(get_db)
):
    """
    Exporta pedido single-tienda a Excel con columna Cuadrante.

    Formato:
    - Código | Descripción | Categoría | ABC | Cuadrante | Stock | Sugerido | Pedido
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        from fastapi.responses import Response

        cursor = conn.cursor()

        # Obtener encabezado del pedido
        cursor.execute("""
            SELECT numero_pedido, tienda_destino_nombre, cedi_origen_nombre,
                   fecha_pedido, estado, dias_cobertura
            FROM pedidos_sugeridos
            WHERE id = %s
        """, [pedido_id])

        pedido_row = cursor.fetchone()
        if not pedido_row:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")

        numero_pedido, tienda_nombre, cedi_nombre, fecha_pedido, estado, dias_cobertura = pedido_row

        # Obtener productos del pedido
        cursor.execute("""
            SELECT
                codigo_producto,
                descripcion_producto,
                categoria,
                clasificacion_abc,
                cuadrante_producto,
                cantidad_bultos,
                stock_tienda,
                cantidad_sugerida_bultos,
                cantidad_pedida_bultos,
                total_unidades
            FROM pedidos_sugeridos_detalle
            WHERE pedido_id = %s AND incluido = true
            ORDER BY clasificacion_abc, descripcion_producto
        """, [pedido_id])

        rows = cursor.fetchall()
        cursor.close()

        if not rows:
            raise HTTPException(status_code=404, detail="No hay productos en el pedido")

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Pedido {numero_pedido[:20]}"

        # Título
        ws.merge_cells('A1:J1')
        ws['A1'] = f"Pedido: {numero_pedido}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Información del pedido
        ws.merge_cells('A2:J2')
        ws['A2'] = f"Tienda: {tienda_nombre} | CEDI: {cedi_nombre} | Fecha: {fecha_pedido} | Cobertura: {dias_cobertura}d"
        ws['A2'].font = Font(size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Headers
        headers = [
            'Código', 'Descripción', 'Categoría', 'ABC', 'Cuadrante',
            'Unid/Bulto', 'Stock', 'Sugerido', 'Pedido', 'Total Unid'
        ]

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Colores por clasificación ABC
        abc_colors = {
            'A': 'E8F5E9',  # Verde claro
            'B': 'FFF9C4',  # Amarillo claro
            'C': 'FFE0B2',  # Naranja claro
            'D': 'F5F5F5'   # Gris claro
        }

        # Datos
        for row_idx, row_data in enumerate(rows, 5):
            abc = row_data[3] or 'D'
            fill_color = abc_colors.get(abc, 'FFFFFF')
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            data = [
                row_data[0],   # Código
                row_data[1],   # Descripción
                row_data[2],   # Categoría
                abc,           # ABC
                row_data[4] or 'NO ESPECIFICADO',  # Cuadrante
                int(row_data[5]) if row_data[5] else 1,
                float(row_data[6]) if row_data[6] else 0,
                float(row_data[7]) if row_data[7] else 0,
                float(row_data[8]) if row_data[8] else 0,
                float(row_data[9]) if row_data[9] else 0
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                if col_idx >= 6:  # Números: alinear derecha
                    cell.alignment = Alignment(horizontal='right')

        # Ajustar anchos
        column_widths = [12, 40, 20, 6, 15, 10, 12, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{numero_pedido.replace('/', '-')}.xlsx"

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")
    except Exception as e:
        logger.error(f"Error exportando: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
